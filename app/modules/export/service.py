"""
app/modules/export/service.py
════════════════════════════════════════════════════════════════════════════════
Enterprise-grade Excel export engine — MAKTech Financial Flow  (v4)

════════════════════════════════════════════════════════════════════════════════
ARCHITECTURE
════════════════════════════════════════════════════════════════════════════════

  resolve_date_range()       → ExportQueryParams → (date_from | None, date_to | None, label)
  _prisma_date_filter()      → date range → Prisma-safe datetime dict (or None for ALL)
  _get_latest_exchange_rate()→ reads most recent rate from DollarExchange rows
  _xl()                      → openpyxl Workbook factory — full enterprise styling
  _build_kpi_summary()       → writes KPI Summary tab for dashboard export
  _copy_sheet_into()         → merges a single-sheet wb into a multi-sheet wb
  export_<module>()          → fetch DB → build workbook → (bytes, filename)

════════════════════════════════════════════════════════════════════════════════
v4 CHANGE LOG
════════════════════════════════════════════════════════════════════════════════

  ExportPeriod.ALL   → NEW: period=all exports the complete dataset with no
                        date filter applied. resolve_date_range() returns
                        (None, None, "All Time"). All find_many() calls omit
                        the date WHERE clause entirely, returning every record.
                        Filename suffix becomes "all_time".

════════════════════════════════════════════════════════════════════════════════
v3 CHANGE LOG
════════════════════════════════════════════════════════════════════════════════

  FiverrEntry        + activeOrderAmount column
  FiverrOrder        + afterFiverr column  (amount × 0.80)
  UpworkOrder        + afterUpwork column  (amount × 0.90)
  OutsideOrder       + orderSheet column; all amounts now labelled ($)
  HrExpense          + remarks column
  PMAK               → SPLIT into Sheet "PMAK" (ledger) + "PMAK Inhouse"
                        all amounts labelled ($)
  CardSharing        → NOW INCLUDED in dashboard + KPI summary (was missing)
                        uses accountId FK → resolves to payoneerAccountName
                        cardNo/cardCvc intentionally excluded from export
  DailyRate model    → REMOVED; rate read from DollarExchange.rate directly
  Dashboard          → 13 sheets total (was 11)

════════════════════════════════════════════════════════════════════════════════
DASHBOARD EXPORT SHEETS  (13 total)
════════════════════════════════════════════════════════════════════════════════

   1. KPI Summary
   2. Fiverr Entries        (+ activeOrderAmount)
   3. Fiverr Orders         (+ afterFiverr)
   4. Upwork Entries
   5. Upwork Orders         (+ afterUpwork)
   6. Payoneer
   7. PMAK                  (ledger only — no buyer/seller)
   8. PMAK Inhouse          (buyer/seller deal tracking — NEW)
   9. Outside Orders        (+ orderSheet; all $ labels)
  10. Dollar Exchange        (BDT column here only)
  11. HR Expense             (+ remarks)
  12. Inventory
  13. Card Sharing           

════════════════════════════════════════════════════════════════════════════════
STYLING SYSTEM
════════════════════════════════════════════════════════════════════════════════

  Brand palette:
    Navy header   #1A3C6E  White text, Bold 11pt
    Teal accent   #0E7490  KPI values, Bold 13pt
    Alt row       #EFF6FF  Light blue stripe
    Status green  #DCFCE7  COMPLETED / CLEARED / RECEIVED / ACTIVE
    Status amber  #FEF3C7  IN_PROGRESS / PENDING / ON_HOLD
    Status red    #FEE2E2  DUE / OVERDUE / CANCELLED / REJECTED
    Totals row    #1A3C6E  White bold — mirrors header

  Features: freeze panes, auto column width, number formats (#,##0.00),
            status-conditional fills, SUM formula totals row, tab colours,
            metadata block, grid-lines hidden.
"""

from __future__ import annotations

import io
import calendar
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from prisma import Prisma

from app.shared.constants import ExportPeriod
from .schema import ExportQueryParams


# ═══════════════════════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ═══════════════════════════════════════════════════════════════════════════════

class _C:
    NAVY        = "1A3C6E"
    TEAL        = "0E7490"
    TEAL_LIGHT  = "CFFAFE"
    ALT_ROW     = "EFF6FF"
    WHITE       = "FFFFFF"
    BLACK       = "0F172A"
    GREY_BORDER = "CBD5E1"
    META_BG     = "F1F5F9"

    GREEN_BG    = "DCFCE7";  GREEN_TEXT  = "166534"
    AMBER_BG    = "FEF3C7";  AMBER_TEXT  = "92400E"
    RED_BG      = "FEE2E2";  RED_TEXT    = "991B1B"
    BLUE_BG     = "DBEAFE";  BLUE_TEXT   = "1E40AF"

    # 13 tab colours — one per dashboard sheet
    TABS = [
        "1A3C6E",  # 0  KPI Summary
        "0E7490",  # 1  Fiverr Entries
        "065F46",  # 2  Fiverr Orders
        "1D4ED8",  # 3  Upwork Entries
        "2563EB",  # 4  Upwork Orders
        "7C3AED",  # 5  Payoneer
        "B45309",  # 6  PMAK (Ledger)
        "92400E",  # 7  PMAK Inhouse   ← NEW
        "DC2626",  # 8  Outside Orders
        "0369A1",  # 9  Dollar Exchange
        "047857",  # 10 HR Expense
        "6D28D9",  # 11 Inventory
        "5B21B6",  # 12 Card Sharing   ← WAS MISSING
    ]


# ═══════════════════════════════════════════════════════════════════════════════
#  STYLE OBJECTS
# ═══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color=_C.GREY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


_HEADER_FILL   = _fill(_C.NAVY)
_TOTALS_FILL   = _fill(_C.NAVY)
_ALT_FILL      = _fill(_C.ALT_ROW)
_META_FILL     = _fill(_C.META_BG)
_KPI_CARD_FILL = _fill(_C.TEAL_LIGHT)

_HEADER_FONT   = Font(name="Calibri", bold=True, color=_C.WHITE,  size=11)
_TOTALS_FONT   = Font(name="Calibri", bold=True, color=_C.WHITE,  size=10)
_BODY_FONT     = Font(name="Calibri",             color=_C.BLACK,  size=10)
_BOLD_FONT     = Font(name="Calibri", bold=True,  color=_C.BLACK,  size=10)
_META_KEY_FONT = Font(name="Calibri", bold=True,  color=_C.NAVY,   size=9)
_META_VAL_FONT = Font(name="Calibri",             color=_C.BLACK,  size=9)
_KPI_LBL_FONT  = Font(name="Calibri", bold=True,  color=_C.NAVY,   size=10)
_KPI_VAL_FONT  = Font(name="Calibri", bold=True,  color=_C.TEAL,   size=13)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=False)
_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

_STATUS_STYLES: dict[str, tuple[str, str]] = {
    "COMPLETED":   (_C.GREEN_BG,  _C.GREEN_TEXT),
    "CLEARED":     (_C.GREEN_BG,  _C.GREEN_TEXT),
    "RECEIVED":    (_C.GREEN_BG,  _C.GREEN_TEXT),
    "ACTIVE":      (_C.GREEN_BG,  _C.GREEN_TEXT),
    "IN_PROGRESS": (_C.AMBER_BG,  _C.AMBER_TEXT),
    "PENDING":     (_C.AMBER_BG,  _C.AMBER_TEXT),
    "ON_HOLD":     (_C.AMBER_BG,  _C.AMBER_TEXT),
    "DUE":         (_C.RED_BG,    _C.RED_TEXT),
    "OVERDUE":     (_C.RED_BG,    _C.RED_TEXT),
    "CANCELLED":   (_C.RED_BG,    _C.RED_TEXT),
    "REJECTED":    (_C.RED_BG,    _C.RED_TEXT),
}

_STATUS_COLUMNS = {"orderStatus", "paymentStatus", "status"}

_CURRENCY_COLUMNS = {
    # Fiverr
    "availableWithdraw", "notCleared", "activeOrderAmount", "submitted",
    "withdrawn", "promotion", "afterFiverr",
    # Upwork
    "pending", "inReview", "workInProgress", "afterUpwork",
    # Orders (shared)
    "amount", "orderAmount", "receiveAmount", "dueAmount",
    # Ledgers
    "debit", "credit", "remainingBalance",
    # Dollar Exchange
    "rate", "totalBdt", "totalBdtLive",
    # Card Sharing
    "cardLimit", "cardPaymentReceive",
    # Inventory
    "unitPrice", "totalPrice",
}

_NUMBER_FMT  = "#,##0.00"
_INTEGER_FMT = "#,##0"
_DATE_FMT    = "yyyy-mm-dd"

_WRAP_COLUMNS = {
    "orderDetails", "paymentMethodDetails", "mailDetails",
    "remarks", "details", "clientLink", "orderSheet", "cardDetails",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  DATE RANGE RESOLVER
# ═══════════════════════════════════════════════════════════════════════════════

def resolve_date_range(
    params: ExportQueryParams,
) -> tuple[date | None, date | None, str]:
    """
    Resolve the export period into (date_from, date_to, human_label).

    Returns (None, None, "All Time") for ExportPeriod.ALL — callers must
    pass the result through _prisma_date_filter() which will return None,
    and then use _date_where() to build the correct Prisma where-clause.
    """
    today = date.today()

    # ExportPeriod.ALL — no date bounds; export entire dataset
    if params.period == ExportPeriod.ALL:
        return None, None, "All Time"

    # Explicit date override takes precedence over period (not applicable to ALL)
    if params.date_from and params.date_to:
        return params.date_from, params.date_to, f"{params.date_from} to {params.date_to}"

    period = params.period

    if period == ExportPeriod.DAILY:
        ref = params.export_date or today
        return ref, ref, ref.strftime("%d %B %Y")

    if period == ExportPeriod.WEEKLY:
        ref    = params.export_date or today
        monday = ref - timedelta(days=ref.weekday())
        sunday = monday + timedelta(days=6)
        return monday, sunday, f"Week {monday.strftime('%d %b')} – {sunday.strftime('%d %b %Y')}"

    if period == ExportPeriod.MONTHLY:
        y      = params.year  or today.year
        m      = params.month or today.month
        d_from = date(y, m, 1)
        d_to   = date(y, m, calendar.monthrange(y, m)[1])
        return d_from, d_to, date(y, m, 1).strftime("%B %Y")

    if period == ExportPeriod.YEARLY:
        y = params.year or today.year
        return date(y, 1, 1), date(y, 12, 31), str(y)

    raise ValueError(f"Unknown period: {period!r}")


def _prisma_date_filter(d_from: date | None, d_to: date | None) -> dict | None:
    """
    Convert a date range into a Prisma-safe datetime filter dict.

    Returns **None** when either bound is None (i.e. period=ALL), signalling
    that the caller should omit the date WHERE clause entirely.

    prisma-client-py requires datetime objects (not bare date) in filters.
    """
    if d_from is None or d_to is None:
        return None
    return {
        "gte": datetime.combine(d_from, time.min),
        "lte": datetime.combine(d_to,   time.max),
    }


def _date_where(df: dict | None) -> dict:
    """
    Build the Prisma ``where`` dict for a date filter.

    - If df is a populated dict  -> ``{"date": df}``  (period-filtered)
    - If df is None              -> ``{}``             (no filter -- ALL period)
    """
    return {"date": df} if df is not None else {}


async def _get_latest_exchange_rate(db: Prisma) -> tuple[float | None, str, str]:
    """
    Fetch the most recent BDT/USD rate from DollarExchange rows.
    DailyRate model has been removed — HR enters the rate on each transaction.
    """
    try:
        rec = await db.dollarexchange.find_first(order={"date": "desc"})
        if rec:
            rate  = float(rec.rate)
            d_lbl = rec.date.strftime("%d %b %Y") if hasattr(rec.date, "strftime") else str(rec.date)
            return rate, d_lbl, f"৳{rate:,.2f}/$"
    except Exception:
        pass
    return None, "N/A", "N/A"


# ═══════════════════════════════════════════════════════════════════════════════
#  CORE WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def _xl(
    sheet_title: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, Any]],
    meta: dict[str, str] | None = None,
    tab_color: str | None = None,
    totals: bool = True,
) -> openpyxl.Workbook:
    """
    Build a styled enterprise-grade .xlsx workbook.
    Returns the openpyxl Workbook object (not bytes) so _copy_sheet_into
    can merge it. Call .save(buf) externally to get bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False

    border  = _thin_border()
    row_ptr = 1

    # ── Metadata block ────────────────────────────────────────────────────────
    if meta:
        for key, val in meta.items():
            k_cell = ws.cell(row=row_ptr, column=1, value=key)
            v_cell = ws.cell(row=row_ptr, column=2, value=val)
            for cell, fnt in ((k_cell, _META_KEY_FONT), (v_cell, _META_VAL_FONT)):
                cell.font      = fnt
                cell.fill      = _META_FILL
                cell.alignment = _LEFT
            row_ptr += 1
        row_ptr += 1   # blank separator row

    # ── Column header row ─────────────────────────────────────────────────────
    header_row = row_ptr
    for col_i, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=row_ptr, column=col_i, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border    = border
    ws.row_dimensions[row_ptr].height = 24
    ws.freeze_panes = ws.cell(row=row_ptr + 1, column=1)
    row_ptr += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_start = row_ptr
    for row_i, row_data in enumerate(rows):
        alt = (row_i % 2 == 1)
        for col_i, (_, key) in enumerate(columns, start=1):
            raw = row_data.get(key)

            if isinstance(raw, Decimal):
                value = float(raw)
            elif isinstance(raw, datetime):
                value = raw.strftime("%Y-%m-%d")
            elif isinstance(raw, date):
                value = raw.strftime("%Y-%m-%d")
            elif isinstance(raw, list):
                value = "\n".join(str(v) for v in raw)
            elif raw is None:
                value = ""
            else:
                value = raw

            cell = ws.cell(row=row_ptr, column=col_i, value=value)
            cell.font      = _BODY_FONT
            cell.alignment = _LEFT
            cell.border    = border

            if alt:
                cell.fill = _ALT_FILL

            if key in _STATUS_COLUMNS and isinstance(value, str):
                style_key = value.upper()
                if style_key in _STATUS_STYLES:
                    bg, fg = _STATUS_STYLES[style_key]
                    cell.fill      = _fill(bg)
                    cell.font      = Font(name="Calibri", bold=True, color=fg, size=10)
                    cell.alignment = _CENTER

            if key in _CURRENCY_COLUMNS and isinstance(value, float):
                cell.number_format = _NUMBER_FMT
                cell.alignment     = _RIGHT
            elif key in {"activeOrders", "connects", "quantity"}:
                cell.number_format = _INTEGER_FMT
                cell.alignment     = _RIGHT
            elif key in _WRAP_COLUMNS:
                cell.alignment = _WRAP

        row_ptr += 1
    data_end = row_ptr - 1

    # ── Totals row ────────────────────────────────────────────────────────────
    if totals and rows:
        numeric_cols = [
            col_i
            for col_i, (_, key) in enumerate(columns, start=1)
            if key in _CURRENCY_COLUMNS
            and isinstance(rows[0].get(key), (Decimal, int, float))
            and not isinstance(rows[0].get(key), bool)
        ]
        if numeric_cols:
            tot_label = ws.cell(row=row_ptr, column=1, value="TOTAL")
            tot_label.font      = _TOTALS_FONT
            tot_label.fill      = _TOTALS_FILL
            tot_label.alignment = _LEFT
            tot_label.border    = border

            for col_i in range(2, len(columns) + 1):
                cell = ws.cell(row=row_ptr, column=col_i)
                if col_i in numeric_cols:
                    col_letter = get_column_letter(col_i)
                    cell.value         = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
                    cell.number_format = _NUMBER_FMT
                    cell.font          = _TOTALS_FONT
                    cell.alignment     = _RIGHT
                else:
                    cell.font = _TOTALS_FONT
                cell.fill   = _TOTALS_FILL
                cell.border = border

    # ── Auto column width ─────────────────────────────────────────────────────
    for col_i, (header, key) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_i)
        max_len = len(header)
        for row_data in rows:
            raw = row_data.get(key, "")
            cell_text = str(raw) if raw is not None else ""
            # For wrapped columns, only count first line
            if key in _WRAP_COLUMNS:
                cell_text = cell_text.split("\n")[0]
            max_len = max(max_len, len(cell_text))

        if key in _WRAP_COLUMNS:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        else:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    return wb


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    """Serialise a workbook to raw bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _copy_sheet_into(
    src_wb: openpyxl.Workbook,
    dst_wb: openpyxl.Workbook,
    sheet_name: str,
    tab_color: str | None = None,
) -> None:
    """Copy the active sheet from src_wb into dst_wb under sheet_name."""
    src_ws = src_wb.active
    dst_ws = dst_wb.create_sheet(title=sheet_name[:31])
    if tab_color:
        dst_ws.sheet_properties.tabColor = tab_color
    dst_ws.sheet_view.showGridLines = False

    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font      = cell.font.copy()
                dst_cell.fill      = cell.fill.copy()
                dst_cell.border    = cell.border.copy()
                dst_cell.alignment = cell.alignment.copy()
                dst_cell.number_format = cell.number_format

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height

    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes


# ═══════════════════════════════════════════════════════════════════════════════
#  KPI SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _build_kpi_summary(
    wb: openpyxl.Workbook,
    label: str,
    kpis: list[dict],
    tab_color: str = _C.NAVY,
) -> None:
    ws = wb.create_sheet(title="KPI Summary", index=0)
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines  = False
    border = _thin_border()

    ws.merge_cells("A1:E1")
    banner = ws["A1"]
    banner.value     = "MAKTech Financial Flow — Dashboard KPI Summary"
    banner.font      = Font(name="Calibri", bold=True, color=_C.WHITE, size=14)
    banner.fill      = _HEADER_FILL
    banner.alignment = _CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value     = f"Period: {label}   ·   Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    sub.font      = Font(name="Calibri", italic=True, color=_C.WHITE, size=10)
    sub.fill      = _fill(_C.TEAL)
    sub.alignment = _CENTER
    ws.row_dimensions[2].height = 20

    for col_i, h in enumerate(["#", "KPI", "Value ($)", "Formula / Note", "Status"], start=1):
        cell = ws.cell(row=3, column=col_i, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _fill(_C.TEAL)
        cell.alignment = _CENTER
        cell.border    = border
    ws.row_dimensions[3].height = 20

    for i, kpi in enumerate(kpis, start=1):
        row = 3 + i
        bg  = _fill(_C.ALT_ROW) if (i % 2 == 0) else _fill(_C.WHITE)

        c0 = ws.cell(row=row, column=1, value=i)
        c0.font = Font(name="Calibri", bold=True, color=_C.NAVY, size=10)
        c0.fill = bg; c0.alignment = _CENTER; c0.border = border

        c1 = ws.cell(row=row, column=2, value=kpi["label"])
        c1.font = _KPI_LBL_FONT; c1.fill = bg
        c1.alignment = _LEFT; c1.border = border

        c2 = ws.cell(row=row, column=3, value=kpi["value"])
        c2.font = _KPI_VAL_FONT; c2.fill = _KPI_CARD_FILL
        c2.alignment = _RIGHT; c2.border = border
        c2.number_format = "#,##0.00"

        c3 = ws.cell(row=row, column=4, value=kpi.get("note", ""))
        c3.font = _META_VAL_FONT; c3.fill = bg
        c3.alignment = _LEFT; c3.border = border

        status_val = kpi.get("status", "")
        c4 = ws.cell(row=row, column=5, value=status_val)
        if status_val.upper() in _STATUS_STYLES:
            sbg, sfg = _STATUS_STYLES[status_val.upper()]
            c4.fill = _fill(sbg)
            c4.font = Font(name="Calibri", bold=True, color=sfg, size=10)
        else:
            c4.font = _BODY_FONT; c4.fill = bg
        c4.alignment = _CENTER; c4.border = border
        ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 58
    ws.column_dimensions["E"].width = 16


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _f(v: Any) -> float:
    """Safe Decimal / None → float."""
    return float(v) if v is not None else 0.0


def _safe_label(label: str) -> str:
    return (
        label.lower()
             .replace(" ", "_")
             .replace("–", "-")
             .replace("/", "-")
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  ROW BUILDERS — one per model
# ═══════════════════════════════════════════════════════════════════════════════

# ── Fiverr Entries ────────────────────────────────────────────────────────────

def _fiverr_entry_rows(entries) -> list[dict]:
    return [
        {
            "date":              e.date,
            "profile":           e.profile.profileName if e.profile else "",
            "availableWithdraw": e.availableWithdraw,
            "notCleared":        e.notCleared,
            "activeOrders":      e.activeOrders,
            "activeOrderAmount": e.activeOrderAmount,   # NEW
            "submitted":         e.submitted,
            "withdrawn":         e.withdrawn,
            "sellerPlus":        "Yes" if e.sellerPlus else "No",
            "promotion":         e.promotion,
        }
        for e in entries
    ]


_FIVERR_ENTRY_COLS = [
    ("Date",                 "date"),
    ("Profile",              "profile"),
    ("Available Withdraw ($)","availableWithdraw"),
    ("Not Cleared ($)",       "notCleared"),
    ("Active Orders",         "activeOrders"),
    ("Active Order Amt ($)",  "activeOrderAmount"),   # NEW
    ("Submitted ($)",         "submitted"),
    ("Withdrawn ($)",         "withdrawn"),
    ("Seller Plus",           "sellerPlus"),
    ("Promotion ($)",         "promotion"),
]


# ── Fiverr Orders ─────────────────────────────────────────────────────────────

def _fiverr_order_rows(orders) -> list[dict]:
    return [
        {
            "date":        o.date,
            "profile":     o.profile.profileName if o.profile else "",
            "buyerName":   o.buyerName,
            "orderId":     o.orderId,
            "amount":      o.amount,
            "afterFiverr": o.afterFiverr,              # NEW
        }
        for o in orders
    ]


_FIVERR_ORDER_COLS = [
    ("Date",              "date"),
    ("Profile",           "profile"),
    ("Buyer",             "buyerName"),
    ("Order ID",          "orderId"),
    ("Amount ($)",        "amount"),
    ("After Fiverr ($)",  "afterFiverr"),              # NEW
]


# ── Upwork Entries ────────────────────────────────────────────────────────────

def _upwork_entry_rows(entries) -> list[dict]:
    return [
        {
            "date":              e.date,
            "profile":           e.profile.profileName if e.profile else "",
            "availableWithdraw": e.availableWithdraw,
            "pending":           e.pending,
            "inReview":          e.inReview,
            "workInProgress":    e.workInProgress,
            "withdrawn":         e.withdrawn,
            "connects":          e.connects,
            "upworkPlus":        "Yes" if e.upworkPlus else "No",
        }
        for e in entries
    ]


_UPWORK_ENTRY_COLS = [
    ("Date",                  "date"),
    ("Profile",               "profile"),
    ("Available Withdraw ($)", "availableWithdraw"),
    ("Pending ($)",            "pending"),
    ("In Review ($)",          "inReview"),
    ("Work In Progress ($)",   "workInProgress"),
    ("Withdrawn ($)",          "withdrawn"),
    ("Connects",               "connects"),
    ("Upwork Plus",            "upworkPlus"),
]


# ── Upwork Orders ─────────────────────────────────────────────────────────────

def _upwork_order_rows(orders) -> list[dict]:
    return [
        {
            "date":        o.date,
            "profile":     o.profile.profileName if o.profile else "",
            "clientName":  o.clientName,
            "orderId":     o.orderId,
            "amount":      o.amount,
            "afterUpwork": o.afterUpwork,              # NEW
        }
        for o in orders
    ]


_UPWORK_ORDER_COLS = [
    ("Date",              "date"),
    ("Profile",           "profile"),
    ("Client",            "clientName"),
    ("Order ID",          "orderId"),
    ("Amount ($)",        "amount"),
    ("After Upwork ($)",  "afterUpwork"),              # NEW
]


# ── Payoneer ──────────────────────────────────────────────────────────────────

def _payoneer_rows(txns) -> list[dict]:
    return [
        {
            "date":             t.date,
            "account":          t.account.accountName if t.account else "",
            "details":          t.details,
            "accountFrom":      t.accountFrom  or "",
            "accountTo":        t.accountTo    or "",
            "debit":            t.debit,
            "credit":           t.credit,
            "remainingBalance": t.remainingBalance,
        }
        for t in txns
    ]


_PAYONEER_COLS = [
    ("Date",                "date"),
    ("Account",             "account"),
    ("Details",             "details"),
    ("Account From",        "accountFrom"),
    ("Account To",          "accountTo"),
    ("Debit ($)",           "debit"),
    ("Credit ($)",          "credit"),
    ("Remaining Balance ($)","remainingBalance"),
]


# ── PMAK Ledger ───────────────────────────────────────────────────────────────

def _pmak_rows(txns) -> list[dict]:
    """PMAK ledger — standard double-entry rows. No buyer/seller (now in PmakInhouse)."""
    return [
        {
            "date":             t.date,
            "account":          t.account.accountName if t.account else "",
            "details":          t.details,
            "accountFrom":      t.accountFrom  or "",
            "accountTo":        t.accountTo    or "",
            "debit":            t.debit,
            "credit":           t.credit,
            "remainingBalance": t.remainingBalance,
        } for t in txns
    ]


_PMAK_COLS = [
    ("Date",                 "date"),
    ("Account",              "account"),
    ("Details",              "details"),
    ("Account From",         "accountFrom"),
    ("Account To",           "accountTo"),
    ("Debit ($)",            "debit"),
    ("Credit ($)",           "credit"),
    ("Remaining Balance ($)", "remainingBalance"),
]


# ── PMAK Inhouse ──────────────────────────────────────────────────────────────

def _pmak_inhouse_rows(deals) -> list[dict]:
    """PMAK Inhouse — buyer/seller deal tracking sheet."""
    return [
        {
            "date":        d.date,
            "account":     d.account.accountName if d.account else "",
            "details":     d.details or "",
            "buyerName":   d.buyerName,
            "sellerName":  d.sellerName,
            "orderAmount": d.orderAmount,
            "orderStatus": d.orderStatus if isinstance(d.orderStatus, str) else (d.orderStatus.value if d.orderStatus else ""),
        }
        for d in deals
    ]


_PMAK_INHOUSE_COLS = [
    ("Date",            "date"),
    ("PMAK Account",    "account"),
    ("Details",         "details"),
    ("Buyer Name",      "buyerName"),
    ("Seller Name",     "sellerName"),
    ("Order Amt ($)",   "orderAmount"),
    ("Order Status",    "orderStatus"),
]


# ── Outside Orders ────────────────────────────────────────────────────────────

def _outside_order_rows(orders) -> list[dict]:
    return [
        {
            "date":                 o.date,
            "clientId":             o.clientId,
            "clientName":           o.clientName,
            "clientLink":           o.clientLink           or "",
            "orderDetails":         o.orderDetails,
            "orderSheet":           o.orderSheet           or "",   # NEW
            "assignTeam":           o.assignTeam           or "",
            "orderStatus":          o.orderStatus if isinstance(o.orderStatus, str) else (o.orderStatus.value if o.orderStatus else ""),
            "orderAmount":          o.orderAmount,
            "receiveAmount":        o.receiveAmount,
            "dueAmount":            o.dueAmount,
            "paymentMethod":        o.paymentMethod        or "",
            "paymentMethodDetails": o.paymentMethodDetails or "",
        }
        for o in orders
    ]


_OUTSIDE_ORDER_COLS = [
    ("Date",              "date"),
    ("Client ID",         "clientId"),
    ("Client Name",       "clientName"),
    ("Client Link",       "clientLink"),
    ("Order Details",     "orderDetails"),
    ("Order Sheet",       "orderSheet"),           # NEW — documented order link
    ("Assigned Team",     "assignTeam"),
    ("Status",            "orderStatus"),
    ("Order Amt ($)",     "orderAmount"),           # USD label
    ("Received ($)",      "receiveAmount"),         # USD label
    ("Due ($)",           "dueAmount"),             # USD label
    ("Payment Method",    "paymentMethod"),
    ("Payment Details",   "paymentMethodDetails"),
]


# ── Dollar Exchange ───────────────────────────────────────────────────────────

def _dollar_exchange_rows(records, current_rate: float | None) -> list[dict]:
    rows = []
    for r in records:
        dollar_amt = _f(r.credit) if _f(r.credit) > 0 else _f(r.debit)
        live_bdt   = round(dollar_amt * current_rate, 2) if current_rate else None
        rows.append({
            "date":          r.date,
            "details":       r.details,
            "accountFrom":   r.accountFrom or "",
            "accountTo":     r.accountTo   or "",
            "debit":         r.debit,
            "credit":        r.credit,
            "rate":          r.rate,
            "totalBdt":      r.totalBdt,
            "totalBdtLive":  live_bdt,
            "paymentStatus": r.paymentStatus if isinstance(r.paymentStatus, str) else (r.paymentStatus.value if r.paymentStatus else ""),
        })
    return rows


def _dollar_exchange_cols(rate_label: str) -> list[tuple[str, str]]:
    return [
        ("Date",                           "date"),
        ("Details",                        "details"),
        ("Account From",                   "accountFrom"),
        ("Account To",                     "accountTo"),
        ("Debit ($)",                      "debit"),
        ("Credit ($)",                     "credit"),
        ("Rate at Entry (৳/$)",            "rate"),
        ("Total BDT (entry rate)",         "totalBdt"),
        (f"Total BDT (live @ {rate_label})", "totalBdtLive"),
        ("Payment Status",                 "paymentStatus"),
    ]


# ── HR Expense ────────────────────────────────────────────────────────────────

def _hr_expense_rows(expenses) -> list[dict]:
    return [
        {
            "date":             e.date,
            "details":          e.details,
            "accountFrom":      e.accountFrom or "",
            "accountTo":        e.accountTo   or "",
            "debit":            e.debit,
            "credit":           e.credit,
            "remainingBalance": e.remainingBalance,
            "remarks":          e.remarks     or "",   # NEW
        }
        for e in expenses
    ]


_HR_EXPENSE_COLS = [
    ("Date",                 "date"),
    ("Details",              "details"),
    ("Account From",         "accountFrom"),
    ("Account To",           "accountTo"),
    ("Debit ($)",            "debit"),
    ("Credit ($)",           "credit"),
    ("Remaining Balance ($)", "remainingBalance"),
    ("Remarks",              "remarks"),               # NEW — CEO judgement
]


# ── Inventory ─────────────────────────────────────────────────────────────────

def _inventory_rows(items) -> list[dict]:
    return [
        {
            "date":       i.date,
            "itemName":   i.itemName,
            "category":   i.category   or "",
            "quantity":   i.quantity,
            "unitPrice":  i.unitPrice,
            "totalPrice": i.totalPrice,
            "condition":  i.condition  or "",
            "assignedTo": i.assignedTo or "",
            "notes":      i.notes      or "",
        }
        for i in items
    ]


_INVENTORY_COLS = [
    ("Date",             "date"),
    ("Item Name",        "itemName"),
    ("Category",         "category"),
    ("Quantity",         "quantity"),
    ("Unit Price ($)",   "unitPrice"),
    ("Total Price ($)",  "totalPrice"),
    ("Condition",        "condition"),
    ("Assigned To",      "assignedTo"),
    ("Notes",            "notes"),
]


# ── Card Sharing ──────────────────────────────────────────────────────────────

def _card_sharing_rows(cards) -> list[dict]:
    """
    Card Sharing export — cardNo and cardCvc intentionally excluded.
    cardDetails (Cloudinary URLs) rendered as newline-joined string.
    accountId FK resolved to account name via the `account` relation.
    """
    return [
        {
            "date":               c.date,
            "serialNo":           c.serialNo,
            "details":            c.details or "",
            "payoneerAccount":    c.account.accountName if c.account else "",
            "cardExpire":         c.cardExpire,
            "cardVendor":         c.cardVendor,
            "cardLimit":          c.cardLimit,
            "cardPaymentReceive": c.cardPaymentReceive,
            "cardReceiveBank":   c.cardReceiveBank,
            "mailDetails":        c.mailDetails or "",
            "cardDetails":        "\n".join(c.cardDetails) if isinstance(c.cardDetails, list) else "",
        }
        for c in cards
    ]


_CARD_SHARING_COLS = [
    ("Date",                 "date"),
    ("Serial No",            "serialNo"),
    ("Details",              "details"),
    ("Payoneer Account",     "payoneerAccount"),
    ("Card Expiry",          "cardExpire"),
    ("Card Vendor",          "cardVendor"),
    ("Card Limit ($)",       "cardLimit"),
    ("Card Payment Received ($)", "cardPaymentReceive"),
    ("Card Receiver Bank",        "cardReceiveBank"),
    ("Mail Details",         "mailDetails"),
    ("Card Screenshots",     "cardDetails"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  PER-MODULE EXPORT FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

async def export_fiverr(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    """Two-sheet workbook: Fiverr Entries + Fiverr Orders."""
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    entries = await db.fiverrentry.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    orders = await db.fiverrorder.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _copy_sheet_into(
        _xl("Fiverr Entries", _FIVERR_ENTRY_COLS, _fiverr_entry_rows(entries),
            {"Module": "Fiverr Entries", "Period": label, "Records": str(len(entries))},
            _C.TABS[1]),
        wb, "Fiverr Entries", _C.TABS[1],
    )
    _copy_sheet_into(
        _xl("Fiverr Orders", _FIVERR_ORDER_COLS, _fiverr_order_rows(orders),
            {"Module": "Fiverr Orders", "Period": label, "Records": str(len(orders))},
            _C.TABS[2]),
        wb, "Fiverr Orders", _C.TABS[2],
    )
    return _wb_to_bytes(wb), f"fiverr_{_safe_label(label)}.xlsx"


async def export_upwork(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    """Two-sheet workbook: Upwork Entries + Upwork Orders."""
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    entries = await db.upworkentry.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    orders = await db.upworkorder.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _copy_sheet_into(
        _xl("Upwork Entries", _UPWORK_ENTRY_COLS, _upwork_entry_rows(entries),
            {"Module": "Upwork Entries", "Period": label, "Records": str(len(entries))},
            _C.TABS[3]),
        wb, "Upwork Entries", _C.TABS[3],
    )
    _copy_sheet_into(
        _xl("Upwork Orders", _UPWORK_ORDER_COLS, _upwork_order_rows(orders),
            {"Module": "Upwork Orders", "Period": label, "Records": str(len(orders))},
            _C.TABS[4]),
        wb, "Upwork Orders", _C.TABS[4],
    )
    return _wb_to_bytes(wb), f"upwork_{_safe_label(label)}.xlsx"


async def export_payoneer(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    txns = await db.payoneertransaction.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )
    return (
        _wb_to_bytes(_xl("Payoneer", _PAYONEER_COLS, _payoneer_rows(txns),
            {"Module": "Payoneer", "Period": label, "Records": str(len(txns))},
            _C.TABS[5])),
        f"payoneer_{_safe_label(label)}.xlsx",
    )


async def export_pmak(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    """Two-sheet workbook: PMAK Ledger + PMAK Inhouse."""
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    txns   = await db.pmaktransaction.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )
    inhouse = await db.pmakinhouse.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _copy_sheet_into(
        _xl("PMAK", _PMAK_COLS, _pmak_rows(txns),
            {"Module": "PMAK Ledger", "Period": label,
             "Note": "Status: PENDING | CLEARED | ON_HOLD | REJECTED"},
            _C.TABS[6]),
        wb, "PMAK", _C.TABS[6],
    )
    _copy_sheet_into(
        _xl("PMAK Inhouse", _PMAK_INHOUSE_COLS, _pmak_inhouse_rows(inhouse),
            {"Module": "PMAK Inhouse", "Period": label,
             "Note": "Buyer/Seller deal tracking"},
            _C.TABS[7]),
        wb, "PMAK Inhouse", _C.TABS[7],
    )
    return _wb_to_bytes(wb), f"pmak_{_safe_label(label)}.xlsx"


async def export_outside_orders(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    orders = await db.outsideorder.find_many(
        where=_date_where(df), order={"date": "asc"},
    )
    return (
        _wb_to_bytes(_xl("Outside Orders", _OUTSIDE_ORDER_COLS, _outside_order_rows(orders),
            {"Module": "Outside Orders", "Period": label,
             "Currency": "All amounts in USD ($)"},
            _C.TABS[8])),
        f"outside_orders_{_safe_label(label)}.xlsx",
    )


async def export_dollar_exchange(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    records = await db.dollarexchange.find_many(where=_date_where(df), order={"date": "asc"})
    rate, rate_date, rate_label = await _get_latest_exchange_rate(db)

    rows    = _dollar_exchange_rows(records, rate)
    columns = _dollar_exchange_cols(rate_label)
    meta = {
        "Module":    "Dollar Exchange",
        "Period":    label,
        "Records":   str(len(records)),
        "Live Rate": f"৳{rate:,.2f} / $1 USD (last entry: {rate_date})" if rate else "No rate data",
        "Note":      "Live BDT uses latest transaction rate. Entry BDT uses rate at time of transaction.",
    }
    return (
        _wb_to_bytes(_xl("Dollar Exchange", columns, rows, meta, _C.TABS[9])),
        f"dollar_exchange_{_safe_label(label)}.xlsx",
    )


async def export_card_sharing(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    """Card sharing — cardNo and cardCvc intentionally excluded for security."""
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    cards = await db.cardsharing.find_many(
        where=_date_where(df),
        include={"account": True},
        order={"date": "asc"},
    )
    meta = {
        "Module":   "Card Sharing",
        "Period":   label,
        "Records":  str(len(cards)),
        "SECURITY": "Card numbers and CVCs are intentionally excluded from this export.",
    }
    return (
        _wb_to_bytes(_xl("Card Sharing", _CARD_SHARING_COLS, _card_sharing_rows(cards),
            meta, _C.TABS[12])),
        f"card_sharing_{_safe_label(label)}.xlsx",
    )


async def export_hr_expense(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    expenses = await db.hrexpense.find_many(where=_date_where(df), order={"date": "asc"})
    return (
        _wb_to_bytes(_xl("HR Expense", _HR_EXPENSE_COLS, _hr_expense_rows(expenses),
            {"Module": "HR Expense", "Period": label, "Records": str(len(expenses))},
            _C.TABS[10])),
        f"hr_expense_{_safe_label(label)}.xlsx",
    )


async def export_inventory(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    items = await db.inventory.find_many(where=_date_where(df), order={"date": "asc"})
    return (
        _wb_to_bytes(_xl("Inventory", _INVENTORY_COLS, _inventory_rows(items),
            {"Module": "Inventory", "Period": label, "Records": str(len(items))},
            _C.TABS[11])),
        f"inventory_{_safe_label(label)}.xlsx",
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD EXPORT — 13-sheet multi-module workbook
# ═══════════════════════════════════════════════════════════════════════════════

async def export_dashboard(db: Prisma, params: ExportQueryParams) -> tuple[bytes, str]:
    """
    Full dashboard export — all 13 sheets with complete field coverage.

    Sheet order:
       1. KPI Summary
       2. Fiverr Entries      (+ activeOrderAmount)
       3. Fiverr Orders       (+ afterFiverr)
       4. Upwork Entries
       5. Upwork Orders       (+ afterUpwork)
       6. Payoneer
       7. PMAK                (ledger only)
       8. PMAK Inhouse        (buyer/seller — NEW)
       9. Outside Orders      (+ orderSheet; all $ labels)
      10. Dollar Exchange     (BDT column here only)
      11. HR Expense          (+ remarks)
      12. Inventory
      13. Card Sharing        (WAS MISSING — now included)
    """
    d_from, d_to, label = resolve_date_range(params)
    df = _prisma_date_filter(d_from, d_to)

    # ── Fetch period-filtered data ─────────────────────────────────────────────
    # _date_where(df) returns {} when df is None (period=ALL) so that
    # Prisma performs an unfiltered query and returns the full dataset.
    fiverr_entries  = await db.fiverrentry.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    fiverr_orders   = await db.fiverrorder.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    upwork_entries  = await db.upworkentry.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    upwork_orders   = await db.upworkorder.find_many(
        where=_date_where(df), include={"profile": True}, order={"date": "asc"},
    )
    payoneer_txns   = await db.payoneertransaction.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )
    pmak_txns       = await db.pmaktransaction.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )
    pmak_inhouse    = await db.pmakinhouse.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )
    outside_orders  = await db.outsideorder.find_many(
        where=_date_where(df), order={"date": "asc"},
    )
    dollar_records  = await db.dollarexchange.find_many(
        where=_date_where(df), order={"date": "asc"},
    )
    hr_expenses     = await db.hrexpense.find_many(
        where=_date_where(df), order={"date": "asc"},
    )
    inventory_items = await db.inventory.find_many(
        where=_date_where(df), order={"date": "asc"},
    )
    card_sharings   = await db.cardsharing.find_many(
        where=_date_where(df), include={"account": True}, order={"date": "asc"},
    )

    # ── All-time fetches for KPI balance calculations ──────────────────────────
    # Ledger modules always show CURRENT (all-time) balances on the KPI sheet.
    # When period=ALL the period-filtered data is already the full dataset,
    # so we reuse it directly instead of issuing duplicate queries.
    if params.period == ExportPeriod.ALL:
        payoneer_all  = payoneer_txns
        pmak_all      = pmak_txns
        hr_all        = hr_expenses
        inventory_all = inventory_items
    else:
        payoneer_all  = await db.payoneertransaction.find_many(
            include={"account": True}, order={"date": "asc"},
        )
        pmak_all      = await db.pmaktransaction.find_many(
            include={"account": True}, order={"date": "asc"},
        )
        hr_all        = await db.hrexpense.find_many(order={"date": "asc"})
        inventory_all = await db.inventory.find_many(order={"date": "asc"})

    current_rate, rate_date, rate_label = await _get_latest_exchange_rate(db)

    # ── KPI calculations ───────────────────────────────────────────────────────
    fiverr_avail         = sum(_f(e.availableWithdraw) for e in fiverr_entries)
    fiverr_not_cleared   = sum(_f(e.notCleared)        for e in fiverr_entries)
    fiverr_order_total   = sum(_f(o.amount)            for o in fiverr_orders)
    fiverr_after_total   = sum(_f(o.afterFiverr)       for o in fiverr_orders)

    upwork_avail         = sum(_f(e.availableWithdraw) for e in upwork_entries)
    upwork_pending       = sum(_f(e.pending)           for e in upwork_entries)
    upwork_order_total   = sum(_f(o.amount)            for o in upwork_orders)
    upwork_after_total   = sum(_f(o.afterUpwork)       for o in upwork_orders)

    # Latest running balance per account (all-time)
    payoneer_accounts: dict[str, float] = {}
    for t in payoneer_all:
        if t.account:
            payoneer_accounts[t.account.accountName] = _f(t.remainingBalance)
    payoneer_balance = sum(payoneer_accounts.values())

    pmak_accounts: dict[str, float] = {}
    for t in pmak_all:
        if t.account:
            pmak_accounts[t.account.accountName] = _f(t.remainingBalance)
    pmak_balance = sum(pmak_accounts.values())

    hr_balance = _f(hr_all[-1].remainingBalance) if hr_all else 0.0

    inv_period_count = len(inventory_items)
    inv_period_value = sum(_f(i.totalPrice) for i in inventory_items)
    inv_total_value  = sum(_f(i.totalPrice) for i in inventory_all)

    outside_order_total = sum(_f(o.orderAmount)  for o in outside_orders)
    outside_due_total   = sum(_f(o.dueAmount)    for o in outside_orders)

    dollar_total_bdt    = sum(_f(r.totalBdt) for r in dollar_records)
    dollar_due_bdt      = sum(_f(r.totalBdt) for r in dollar_records
                              if (r.paymentStatus if isinstance(r.paymentStatus, str) else r.paymentStatus.value) == "DUE")
    dollar_rcv_bdt      = sum(_f(r.totalBdt) for r in dollar_records
                              if (r.paymentStatus if isinstance(r.paymentStatus, str) else r.paymentStatus.value) == "RECEIVED")

    card_limit_total    = sum(_f(c.cardLimit)          for c in card_sharings)
    card_payment_total  = sum(_f(c.cardPaymentReceive) for c in card_sharings)
    card_rcv_back_banks = ", ".join({c.cardReceiveBank for c in card_sharings if c.cardReceiveBank})

    kpis = [
        {
            "label":  "Fiverr — Available Withdraw",
            "value":  fiverr_avail,
            "note":   f"{len(fiverr_entries)} entries · {len({e.profileId for e in fiverr_entries})} profiles",
            "status": "ACTIVE",
        },
        {
            "label":  "Fiverr — Not Cleared",
            "value":  fiverr_not_cleared,
            "note":   "Sum of notCleared across all Fiverr entries in period",
            "status": "PENDING",
        },
        {
            "label":  "Fiverr Orders — Gross / Net",
            "value":  fiverr_order_total,
            "note":   f"{len(fiverr_orders)} orders · After Fiverr fee: ${fiverr_after_total:,.2f}",
            "status": "ACTIVE",
        },
        {
            "label":  "Upwork — Available Withdraw",
            "value":  upwork_avail,
            "note":   f"{len(upwork_entries)} entries · {len({e.profileId for e in upwork_entries})} profiles",
            "status": "ACTIVE",
        },
        {
            "label":  "Upwork — Pending",
            "value":  upwork_pending,
            "note":   "Sum of pending across all Upwork entries in period",
            "status": "PENDING",
        },
        {
            "label":  "Upwork Orders — Gross / Net",
            "value":  upwork_order_total,
            "note":   f"{len(upwork_orders)} orders · After Upwork fee: ${upwork_after_total:,.2f}",
            "status": "ACTIVE",
        },
        {
            "label":  "Payoneer Balance (Current)",
            "value":  payoneer_balance,
            "note":   f"Latest all-time balance · {len(payoneer_txns)} txns in period",
            "status": "ACTIVE",
        },
        {
            "label":  "PMAK Balance (Current)",
            "value":  pmak_balance,
            "note":   f"Latest all-time balance · {len(pmak_txns)} txns · {len(pmak_inhouse)} inhouse deals in period",
            "status": "ACTIVE",
        },
        {
            "label":  "Dollar Exchange — Total BDT",
            "value":  dollar_total_bdt,
            "note":   f"RECEIVED ৳{dollar_rcv_bdt:,.2f}  |  DUE ৳{dollar_due_bdt:,.2f}  |  Rate: {rate_label}",
            "status": "RECEIVED" if dollar_due_bdt == 0 else "DUE",
        },
        {
            "label":  "Outside Orders — Total",
            "value":  outside_order_total,
            "note":   f"{len(outside_orders)} orders  |  Due: ${outside_due_total:,.2f}",
            "status": "PENDING" if outside_due_total > 0 else "COMPLETED",
        },
        {
            "label":  "HR Expense (Current Balance)",
            "value":  hr_balance,
            "note":   f"Latest all-time ledger balance · {len(hr_expenses)} entries in period",
            "status": "ACTIVE",
        },
        {
            "label":  "Inventory — Period Additions",
            "value":  inv_period_value,
            "note":   f"{inv_period_count} new items · Total asset value: ${inv_total_value:,.2f}",
            "status": "ACTIVE" if inv_period_count > 0 else "PENDING",
        },
        {
            "label":  "Card Sharing — Total Limit",
            "value":  card_limit_total,
            "note":   f"{len(card_sharings)} cards · Received: ${card_payment_total:,.2f} · Banks: {card_rcv_back_banks}",
            "status": "ACTIVE",
        },
    ]

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1 — KPI Summary
    _build_kpi_summary(wb, label, kpis, _C.TABS[0])

    # Sheet 2 — Fiverr Entries
    _copy_sheet_into(
        _xl("Fiverr Entries", _FIVERR_ENTRY_COLS, _fiverr_entry_rows(fiverr_entries),
            {"Module": "Fiverr Entries", "Period": label}),
        wb, "Fiverr Entries", _C.TABS[1],
    )

    # Sheet 3 — Fiverr Orders
    _copy_sheet_into(
        _xl("Fiverr Orders", _FIVERR_ORDER_COLS, _fiverr_order_rows(fiverr_orders),
            {"Module": "Fiverr Orders", "Period": label}),
        wb, "Fiverr Orders", _C.TABS[2],
    )

    # Sheet 4 — Upwork Entries
    _copy_sheet_into(
        _xl("Upwork Entries", _UPWORK_ENTRY_COLS, _upwork_entry_rows(upwork_entries),
            {"Module": "Upwork Entries", "Period": label}),
        wb, "Upwork Entries", _C.TABS[3],
    )

    # Sheet 5 — Upwork Orders
    _copy_sheet_into(
        _xl("Upwork Orders", _UPWORK_ORDER_COLS, _upwork_order_rows(upwork_orders),
            {"Module": "Upwork Orders", "Period": label}),
        wb, "Upwork Orders", _C.TABS[4],
    )

    # Sheet 6 — Payoneer
    _copy_sheet_into(
        _xl("Payoneer", _PAYONEER_COLS, _payoneer_rows(payoneer_txns),
            {"Module": "Payoneer", "Period": label}),
        wb, "Payoneer", _C.TABS[5],
    )

    # Sheet 7 — PMAK Ledger
    _copy_sheet_into(
        _xl("PMAK", _PMAK_COLS, _pmak_rows(pmak_txns),
            {"Module": "PMAK Ledger", "Period": label,
             "Note": "Status: PENDING | CLEARED | ON_HOLD | REJECTED"}),
        wb, "PMAK", _C.TABS[6],
    )

    # Sheet 8 — PMAK Inhouse (NEW)
    _copy_sheet_into(
        _xl("PMAK Inhouse", _PMAK_INHOUSE_COLS, _pmak_inhouse_rows(pmak_inhouse),
            {"Module": "PMAK Inhouse", "Period": label,
             "Note": "Buyer/Seller inhouse deal tracking"}),
        wb, "PMAK Inhouse", _C.TABS[7],
    )

    # Sheet 9 — Outside Orders
    _copy_sheet_into(
        _xl("Outside Orders", _OUTSIDE_ORDER_COLS, _outside_order_rows(outside_orders),
            {"Module": "Outside Orders", "Period": label,
             "Currency": "All amounts in USD ($)"}),
        wb, "Outside Orders", _C.TABS[8],
    )

    # Sheet 10 — Dollar Exchange (BDT column only here)
    _copy_sheet_into(
        _xl("Dollar Exchange", _dollar_exchange_cols(rate_label),
            _dollar_exchange_rows(dollar_records, current_rate),
            {"Module": "Dollar Exchange", "Period": label,
             "Live Rate": f"৳{current_rate:,.2f}/$1 (last entry: {rate_date})" if current_rate else "N/A"}),
        wb, "Dollar Exchange", _C.TABS[9],
    )

    # Sheet 11 — HR Expense
    _copy_sheet_into(
        _xl("HR Expense", _HR_EXPENSE_COLS, _hr_expense_rows(hr_expenses),
            {"Module": "HR Expense", "Period": label}),
        wb, "HR Expense", _C.TABS[10],
    )

    # Sheet 12 — Inventory
    _copy_sheet_into(
        _xl("Inventory", _INVENTORY_COLS, _inventory_rows(inventory_items),
            {"Module": "Inventory", "Period": label}),
        wb, "Inventory", _C.TABS[11],
    )

    # Sheet 13 — Card Sharing (WAS MISSING)
    _copy_sheet_into(
        _xl("Card Sharing", _CARD_SHARING_COLS, _card_sharing_rows(card_sharings),
            {"Module": "Card Sharing", "Period": label,
             "SECURITY": "Card numbers and CVCs excluded from all exports"}),
        wb, "Card Sharing", _C.TABS[12],
    )

    return _wb_to_bytes(wb), f"dashboard_{_safe_label(label)}.xlsx"