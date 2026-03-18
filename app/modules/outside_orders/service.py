"""
app/modules/outside_orders/service.py
========================================
v3 — date serialization fix

DateTime @db.Date fields in prisma-py v0.14.0 MUST be passed as
datetime.datetime objects (midnight UTC), identical to the pattern used
throughout the Fiverr module:

    datetime.combine(data.date, time.min)

Passing datetime.date directly or an ISO string both cause:
    TypeError / "Could not find field at createOneOutsideOrder.data.date"
"""
from __future__ import annotations

import io
from datetime import date as dt_date, datetime, time
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import HTTPException
from prisma import Prisma

from .schema import OutsideOrderCreate, OutsideOrderUpdate


# ─────────────────────────────────────────────────────────────────────────────
# Private helpers  (mirror the Fiverr service convention exactly)
# ─────────────────────────────────────────────────────────────────────────────

def _dt(d: dt_date) -> datetime:
    """
    Convert datetime.date → datetime.datetime at midnight.

    prisma-py v0.14.0 requires a full datetime object for every
    DateTime @db.Date field.  This is the same helper pattern used in
    the Fiverr service (datetime.combine(d, time.min)).
    """
    return datetime.combine(d, time.min)


def _enum_val(v) -> str:
    """Safely extract the string value from an Enum (or plain str)."""
    return v.value if hasattr(v, "value") else v


def _to_date(v) -> dt_date:
    """
    Normalise a value returned by the ORM back to a plain date.
    prisma-py may return DateTime @db.Date fields as datetime or date.
    """
    return v.date() if isinstance(v, datetime) else v


# ─────────────────────────────────────────────────────────────────────────────
# CRUD
# ─────────────────────────────────────────────────────────────────────────────

async def create_order(db: Prisma, data: OutsideOrderCreate):
    # clientId has no @unique constraint → use find_first, not find_unique
    existing = await db.outsideorder.find_first(where={"clientId": data.client_id})
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Client ID '{data.client_id}' already has an order (order ID: {existing.id})",
        )

    order_amount   = float(data.order_amount)
    receive_amount = float(data.receive_amount)
    due_amount     = order_amount - receive_amount

    return await db.outsideorder.create(
        data={
            "clientId":             data.client_id,
            "clientName":           data.client_name,
            "clientLink":           data.client_link,
            "orderDetails":         data.order_details,
            "orderSheet":           data.order_sheet,
            "assignTeam":           data.assign_team,
            "orderStatus":          _enum_val(data.order_status),
            "orderAmount":          order_amount,
            "receiveAmount":        receive_amount,
            "dueAmount":            due_amount,
            "paymentMethod":        data.payment_method,
            "paymentMethodDetails": data.payment_method_details,
            "date":                 _dt(data.date),   # ← datetime.combine(date, time.min)
        }
    )


async def list_orders(
    db: Prisma,
    date_filter: dict,
    status:      Optional[str] = None,
    client_name: Optional[str] = None,
    assign_team: Optional[str] = None,
):
    """
    List outside orders with optional filters.

    Filters (all combinable):
      date_filter  — Prisma-compatible dict from DateRangeFilter.to_prisma_filter()
      status       — exact match on orderStatus enum value (case-insensitive normalised)
      client_name  — case-insensitive substring search on clientName
      assign_team  — case-insensitive substring search on assignTeam
    """
    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if status:
        where["orderStatus"] = status.upper()

    if client_name:
        where["clientName"] = {"contains": client_name, "mode": "insensitive"}

    if assign_team:
        where["assignTeam"] = {"contains": assign_team, "mode": "insensitive"}

    return await db.outsideorder.find_many(where=where, order={"date": "desc"})


async def get_order(db: Prisma, order_id: str):
    order = await db.outsideorder.find_unique(where={"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


async def update_order(db: Prisma, order_id: str, data: OutsideOrderUpdate):
    order = await get_order(db, order_id)

    update_data: dict = {}

    if data.client_name is not None:
        update_data["clientName"] = data.client_name
    if data.client_link is not None:
        update_data["clientLink"] = data.client_link
    if data.order_details is not None:
        update_data["orderDetails"] = data.order_details
    if data.order_sheet is not None:
        update_data["orderSheet"] = data.order_sheet
    if data.assign_team is not None:
        update_data["assignTeam"] = data.assign_team
    if data.order_status is not None:
        update_data["orderStatus"] = _enum_val(data.order_status)
    if data.payment_method is not None:
        update_data["paymentMethod"] = data.payment_method
    if data.payment_method_details is not None:
        update_data["paymentMethodDetails"] = data.payment_method_details

    # Always recompute dueAmount from whichever amounts are changing
    order_amount   = float(data.order_amount)   if data.order_amount   is not None else float(order.orderAmount)
    receive_amount = float(data.receive_amount) if data.receive_amount is not None else float(order.receiveAmount)

    if data.order_amount is not None:
        update_data["orderAmount"]   = float(data.order_amount)
    if data.receive_amount is not None:
        update_data["receiveAmount"] = float(data.receive_amount)

    update_data["dueAmount"] = order_amount - receive_amount

    return await db.outsideorder.update(where={"id": order_id}, data=update_data)


async def delete_order(db: Prisma, order_id: str):
    await get_order(db, order_id)
    await db.outsideorder.delete(where={"id": order_id})


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HEADERS = [
    "Date", "Client ID", "Client Name", "Client Link",
    "Order Details", "Status", "Order Amount (USD)", "Received Amount (USD)",
    "Due Amount (USD)", "Payment Method", "Payment Details", "Assign Team",
]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_COL_WIDTHS = [12, 16, 22, 28, 40, 12, 20, 20, 18, 18, 28, 18]


def _fmt(v) -> str:
    """Safe string formatter for cell values."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, dt_date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


async def export_orders(
    db:          Prisma,
    date_filter: dict,
    status:      Optional[str] = None,
    client_name: Optional[str] = None,
    assign_team: Optional[str] = None,
    label:       str = "outside_orders",
) -> tuple[bytes, str]:
    """
    Build and return (xlsx_bytes, filename) for the filtered orders.
    Parameters mirror list_orders() so the same filters apply to export.
    """
    rows = await list_orders(db, date_filter, status, client_name, assign_team)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outside Orders"

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, order in enumerate(rows, start=2):
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else None
        values = [
            _fmt(order.date),         # normalise datetime → date string
            _fmt(order.clientId),
            _fmt(order.clientName),
            _fmt(order.clientLink),
            _fmt(order.orderDetails),
            _fmt(order.orderStatus),
            float(order.orderAmount),
            float(order.receiveAmount),
            float(order.dueAmount),
            _fmt(order.paymentMethod),
            _fmt(order.paymentMethodDetails),
            _fmt(order.assignTeam),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER if col_idx in (1, 6) else _LEFT
            if fill:
                cell.fill = fill

    # ── Summary row ───────────────────────────────────────────────────────────
    if rows:
        summary_row = len(rows) + 2
        ws.cell(row=summary_row, column=6,  value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=7,  value=sum(float(r.orderAmount)   for r in rows)).font = Font(bold=True)
        ws.cell(row=summary_row, column=8,  value=sum(float(r.receiveAmount) for r in rows)).font = Font(bold=True)
        ws.cell(row=summary_row, column=9,  value=sum(float(r.dueAmount)     for r in rows)).font = Font(bold=True)

    ws.freeze_panes = "A2"

    # ── Serialize ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{label}.xlsx"
    return buffer.read(), filename