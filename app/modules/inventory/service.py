"""
app/modules/inventory/service.py
════════════════════════════════════════════════════════════════════════════════
v3 — search/filter + bulk Excel import

  New:
    • list_inventory  — accepts optional `search` keyword; OR-matched across
                        itemName, category, condition, assignedTo, notes.
    • bulk_import_items — parses an uploaded .xlsx file and inserts each
                          valid row by calling the existing create_item().
                          Invalid rows are collected into an errors list —
                          good rows still import even if some rows fail.

  Unchanged:
    create_item, update_item, delete_item, get_item — byte-for-byte identical.
════════════════════════════════════════════════════════════════════════════════

DateTime @db.Date fields in prisma-py v0.14.0 MUST be passed as
datetime.datetime objects (midnight), using the same pattern as the
working Fiverr module:

    datetime.combine(data.date, time.min)

Decimal fields must be cast to float before being passed to the builder.

All field names verified against schema.prisma (model Inventory):
  date, itemName, category, quantity, unitPrice, totalPrice,
  condition, assignedTo, notes, createdAt, updatedAt
"""
import io
from datetime import date as dt_date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl
from fastapi import HTTPException, UploadFile
from prisma import Prisma

from .schema import (
    InventoryBulkImportResponse,
    InventoryCreate,
    InventoryImportError,
    InventoryUpdate,
)


# ── Date helper ───────────────────────────────────────────────────────────────

def _dt(d: dt_date) -> datetime:
    """
    Convert datetime.date → datetime.datetime at midnight.

    prisma-py v0.14.0 requires a full datetime object for every
    DateTime @db.Date field — the same pattern used across the Fiverr
    module: datetime.combine(d, time.min).
    """
    return datetime.combine(d, time.min)


# ── Excel column-header aliases ────────────────────────────────────────────────
# Keys   : normalised header strings (lowercase, stripped)
# Values : InventoryCreate field names
_HEADER_ALIASES: dict[str, str] = {
    "date":        "date",
    "itemname":    "itemName",
    "item_name":   "itemName",
    "item name":   "itemName",
    "item":        "itemName",
    "category":    "category",
    "cat":         "category",
    "quantity":    "quantity",
    "qty":         "quantity",
    "unitprice":   "unitPrice",
    "unit_price":  "unitPrice",
    "unit price":  "unitPrice",
    "price":       "unitPrice",
    "condition":   "condition",
    "assignedto":  "assignedTo",
    "assigned_to": "assignedTo",
    "assigned to": "assignedTo",
    "assignee":    "assignedTo",
    "notes":       "notes",
    "note":        "notes",
    "remarks":     "notes",
}


# ── Excel value coercers ───────────────────────────────────────────────────────

def _parse_date(raw) -> dt_date:
    """
    Coerce an Excel cell value to a Python date.

    openpyxl returns:
      • datetime / date objects for proper date-formatted cells
      • float (Excel serial) in rare edge cases — converted by openpyxl already
      • str for text-formatted date cells (ISO-8601 accepted)
    """
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, dt_date):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        # Accept YYYY-MM-DD or DD/MM/YYYY or MM/DD/YYYY heuristically
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Cannot parse date: {raw!r}")


def _parse_decimal(raw, default: Decimal = Decimal("0")) -> Decimal:
    if raw is None or str(raw).strip() == "":
        return default
    try:
        return Decimal(str(raw).strip().replace(",", ""))
    except InvalidOperation:
        raise ValueError(f"Invalid number: {raw!r}")


def _parse_int(raw, default: int = 1) -> int:
    if raw is None or str(raw).strip() == "":
        return default
    try:
        return int(str(raw).strip())
    except (ValueError, TypeError):
        raise ValueError(f"Invalid integer: {raw!r}")


def _str_or_none(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


# ── CRUD ──────────────────────────────────────────────────────────────────────

async def list_inventory(
    db:          Prisma,
    date_filter: dict,
    search:      str | None = None,
):
    """
    List inventory records with optional date filtering and keyword search.

    search — single case-insensitive keyword OR-matched across:
             itemName, category, condition, assignedTo, notes.
             Combinable freely with date_filter (AND between the two).
    """
    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if search:
        where["OR"] = [
            {"itemName":   {"contains": search, "mode": "insensitive"}},
            {"category":   {"contains": search, "mode": "insensitive"}},
            {"condition":  {"contains": search, "mode": "insensitive"}},
            {"assignedTo": {"contains": search, "mode": "insensitive"}},
            {"notes":      {"contains": search, "mode": "insensitive"}},
        ]

    return await db.inventory.find_many(where=where, order={"date": "desc"})


async def create_item(db: Prisma, data: InventoryCreate):
    return await db.inventory.create(
        data={
            "date":       _dt(data.date),            # ← datetime.combine(date, time.min)
            "itemName":   data.itemName,
            "category":   data.category,
            "quantity":   data.quantity,
            "unitPrice":  float(data.unitPrice),     # ← Decimal not serializable
            "totalPrice": float(data.totalPrice),    # ← Decimal not serializable
            "condition":  data.condition,
            "assignedTo": data.assignedTo,
            "notes":      data.notes,
        }
    )


async def update_item(db: Prisma, item_id: str, data: InventoryUpdate):
    existing = await db.inventory.find_unique(where={"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Inventory item not found")

    update_data = data.model_dump(exclude_none=True)

    # Cast any Decimal values to float before they reach the prisma-py builder
    for key in ("unitPrice", "totalPrice"):
        if key in update_data and isinstance(update_data[key], Decimal):
            update_data[key] = float(update_data[key])

    # Recompute totalPrice if quantity or unitPrice changed
    new_quantity   = update_data.get("quantity",  existing.quantity)
    new_unit_price = update_data.get("unitPrice", float(existing.unitPrice))
    update_data["totalPrice"] = float(new_unit_price) * int(new_quantity)

    return await db.inventory.update(where={"id": item_id}, data=update_data)


async def delete_item(db: Prisma, item_id: str):
    existing = await db.inventory.find_unique(where={"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    await db.inventory.delete(where={"id": item_id})


async def get_item(db: Prisma, item_id: str):
    item = await db.inventory.find_unique(where={"id": item_id})
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    return item


# ── Bulk Excel import ──────────────────────────────────────────────────────────

async def bulk_import_items(
    db:   Prisma,
    file: UploadFile,
) -> InventoryBulkImportResponse:
    """
    Parse an uploaded .xlsx file and insert each valid row as an inventory item.

    Excel format (row 1 = headers, row 2+ = data):
    ┌──────────┬──────────┬──────────┬──────────┬───────────┬───────────┬────────────┬───────┐
    │  date    │ itemName │ category │ quantity │ unitPrice │ condition │ assignedTo │ notes │
    └──────────┴──────────┴──────────┴──────────┴───────────┴───────────┴────────────┴───────┘

    • Required columns : date, itemName
    • Optional columns : category, quantity (default 1), unitPrice (default 0),
                         condition, assignedTo, notes
    • Header matching  : case-insensitive; common aliases accepted
                         (qty, unit price, item name, assignee, remarks, …)
    • Row isolation    : a bad row never blocks the rest — it goes to errors[]
    • totalPrice       : auto-computed (unitPrice × quantity) — never read from sheet

    Returns InventoryBulkImportResponse with:
      importedCount, skippedCount, records (success), errors (failures)
    """
    # ── File-type guard ────────────────────────────────────────────────────────
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls).",
        )

    # ── Read workbook ──────────────────────────────────────────────────────────
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400,
            detail="Could not read the Excel file. Ensure it is a valid .xlsx workbook.",
        )
    ws = wb.active

    # ── Parse and map headers (row 1) ──────────────────────────────────────────
    raw_headers = [
        str(cell.value).strip().lower() if cell.value is not None else ""
        for cell in ws[1]
    ]
    # col_map: field_name → column index (0-based)
    col_map: dict[str, int] = {}
    for idx, raw_h in enumerate(raw_headers):
        field = _HEADER_ALIASES.get(raw_h)
        if field and field not in col_map:          # first match wins
            col_map[field] = idx

    # Minimum required columns
    missing = [f for f in ("date", "itemName") if f not in col_map]
    if missing:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Required column(s) not found in Excel headers: {missing}. "
                f"Detected headers: {raw_headers}"
            ),
        )

    # ── Row iteration ──────────────────────────────────────────────────────────
    imported_records = []
    errors: list[InventoryImportError] = []

    def _cell(row_values: tuple, field: str):
        """Return the cell value for a mapped field, or None if column absent."""
        idx = col_map.get(field)
        return row_values[idx] if idx is not None and idx < len(row_values) else None

    for row_num, row_values in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip fully empty rows
        if all(v is None for v in row_values):
            continue

        try:
            # ── Required fields ────────────────────────────────────────────────
            raw_date      = _cell(row_values, "date")
            raw_item_name = _cell(row_values, "itemName")

            entry_date = _parse_date(raw_date)
            item_name  = _str_or_none(raw_item_name)
            if not item_name:
                raise ValueError("itemName is required and cannot be empty")

            # ── Optional fields with safe defaults ─────────────────────────────
            category   = _str_or_none(_cell(row_values, "category"))
            quantity   = _parse_int(_cell(row_values, "quantity"),     default=1)
            unit_price = _parse_decimal(_cell(row_values, "unitPrice"), default=Decimal("0"))
            condition  = _str_or_none(_cell(row_values, "condition"))
            assigned_to = _str_or_none(_cell(row_values, "assignedTo"))
            notes      = _str_or_none(_cell(row_values, "notes"))

            # ── Construct payload — model_validator auto-computes totalPrice ───
            payload = InventoryCreate(
                date=entry_date,
                itemName=item_name,
                category=category,
                quantity=quantity,
                unitPrice=unit_price,
                condition=condition,
                assignedTo=assigned_to,
                notes=notes,
            )

            # ── Insert via the existing create_item (signature untouched) ──────
            record = await create_item(db, payload)
            imported_records.append(record)

        except HTTPException as exc:
            errors.append(InventoryImportError(row=row_num, error=exc.detail))
        except Exception as exc:
            errors.append(InventoryImportError(row=row_num, error=str(exc)))

    return InventoryBulkImportResponse(
        importedCount=len(imported_records),
        skippedCount=len(errors),
        records=imported_records,
        errors=errors,
    )