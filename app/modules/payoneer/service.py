"""
app/modules/payoneer/service.py
════════════════════════════════════════════════════════════════════════════════
v7 — Enterprise Edition
════════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import io
import logging
import math
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Optional

from fastapi import HTTPException
from prisma import Prisma

from app.shared.filters import DateRangeFilter, to_dt_start, to_dt_end
from app.shared.pagination import PageParams
from .schema import (
    PayoneerAccountCreate,
    PayoneerAccountUpdate,
    PayoneerTransactionCreate,
    PayoneerTransactionUpdate,
)

logger = logging.getLogger(__name__)
_ZERO = Decimal("0")

# ─────────────────────────────────────────────────────────────────────────────
# § TS  Timestamp bootstrap
#
# PayoneerAccount     — no timestamps at all in Prisma schema
# PayoneerTransaction — has createdAt; missing updatedAt
#
# Each ALTER is a separate execute_raw call — PostgreSQL extended-query
# protocol forbids multiple commands in a single prepared statement.
# A module-level flag prevents re-running after the first request.
# ─────────────────────────────────────────────────────────────────────────────

_PAYONEER_TS_DDL: list[str] = [
    # payoneer_accounts — no timestamps in Prisma schema
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='payoneer_accounts' AND column_name='created_at'
      ) THEN
        ALTER TABLE payoneer_accounts ADD COLUMN created_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='payoneer_accounts' AND column_name='updated_at'
      ) THEN
        ALTER TABLE payoneer_accounts ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    # payoneer_transactions — Prisma has createdAt; missing updatedAt
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='payoneer_transactions' AND column_name='updated_at'
      ) THEN
        ALTER TABLE payoneer_transactions ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
]

_payoneer_ts_done = False


async def _ensure_payoneer_timestamps(db: Prisma) -> None:
    """
    Idempotent bootstrap — runs once per process lifetime.
    Adds createdAt / updatedAt to Payoneer tables that Prisma didn't generate
    them for.  One execute_raw per statement — never multiple commands per call.
    """
    global _payoneer_ts_done
    if _payoneer_ts_done:
        return
    for stmt in _PAYONEER_TS_DDL:
        await db.execute_raw(stmt)
    _payoneer_ts_done = True


async def _fetch_account_timestamps(db: Prisma, account_id: str) -> dict:
    """Read raw created_at / updated_at from payoneer_accounts. Falls back to None."""
    try:
        rows = await db.query_raw(
            "SELECT created_at, updated_at FROM payoneer_accounts WHERE id = $1",
            account_id,
        )
        if rows:
            return {
                "createdAt": rows[0].get("created_at"),
                "updatedAt": rows[0].get("updated_at"),
            }
    except Exception:
        pass
    return {"createdAt": None, "updatedAt": None}


async def _fetch_accounts_timestamps_batch(db: Prisma, account_ids: list) -> dict:
    """
    Batch-fetch created_at / updated_at for a list of account IDs.
    Returns {account_id: {createdAt, updatedAt}}.
    Single query — no N+1.
    """
    if not account_ids:
        return {}
    try:
        placeholders = ", ".join(f"${i+1}" for i in range(len(account_ids)))
        rows = await db.query_raw(
            f"SELECT id, created_at, updated_at FROM payoneer_accounts WHERE id IN ({placeholders})",
            *account_ids,
        )
        return {
            r["id"]: {"createdAt": r.get("created_at"), "updatedAt": r.get("updated_at")}
            for r in rows
        }
    except Exception:
        return {aid: {"createdAt": None, "updatedAt": None} for aid in account_ids}


async def _fetch_tx_timestamps_batch(db: Prisma, tx_ids: list) -> dict:
    """
    Batch-fetch createdAt / updatedAt for a list of transaction IDs.
    Returns {tx_id: {createdAt, updatedAt}}.  Single query — no N+1.
    """
    if not tx_ids:
        return {}
    try:
        placeholders = ", ".join(f"${i+1}" for i in range(len(tx_ids)))
        rows = await db.query_raw(
            f"SELECT id, created_at, updated_at FROM payoneer_transactions WHERE id IN ({placeholders})",
            *tx_ids,
        )
        return {
            r["id"]: {"createdAt": r.get("created_at"), "updatedAt": r.get("updated_at")}
            for r in rows
        }
    except Exception:
        return {tid: {"createdAt": None, "updatedAt": None} for tid in tx_ids}


async def _touch_tx_updated_at(db: Prisma, tx_id: str) -> None:
    """Bump updated_at on a payoneer_transactions row after any write."""
    try:
        await db.execute_raw(
            "UPDATE payoneer_transactions SET updated_at = now() WHERE id = $1",
            tx_id,
        )
    except Exception:
        pass


async def _touch_account_updated_at(db: Prisma, account_id: str) -> None:
    """Bump updated_at on a payoneer_accounts row after any write."""
    try:
        await db.execute_raw(
            "UPDATE payoneer_accounts SET updated_at = now() WHERE id = $1",
            account_id,
        )
    except Exception:
        pass




# ── Private helpers ───────────────────────────────────────────────────────────

def _d(v: Any) -> Decimal:
    return _ZERO if v is None else Decimal(str(v))


def _tx_to_dict(tx: Any, account_name: str, created_at: Any = None, updated_at: Any = None) -> dict:
    """
    ORM transaction → serialisable dict with accountName injected.

    ``created_at`` — read from the ORM object directly (Prisma has this column).
    ``updated_at``  — pass the value fetched from the raw payoneer_transactions.updated_at
    column (fetched separately via query_raw).  Defaults to None when the caller
    does not hold a DB handle (e.g. inside list builders that already have
    batch-fetched values).
    """
    return {
        "id":               tx.id,
        "accountId":        tx.accountId,
        "accountName":      account_name,
        "date":             tx.date.date() if isinstance(tx.date, datetime) else tx.date,
        "details":          tx.details,
        "accountFrom":      tx.accountFrom,
        "accountTo":        tx.accountTo,
        "debit":            _d(tx.debit),
        "credit":           _d(tx.credit),
        "remainingBalance": _d(tx.remainingBalance),
        "createdAt":        created_at if created_at is not None else getattr(tx, "createdAt", None),
        "updatedAt":        updated_at,
    }


def _pagination_meta(pagination: Optional[PageParams], total: int) -> dict:
    if pagination is None:
        return {"page": 1, "pageSize": total, "total": total, "totalPages": 1}
    return {
        "page":       pagination.page,
        "pageSize":   pagination.page_size,
        "total":      total,
        "totalPages": math.ceil(total / pagination.page_size) if total > 0 else 1,
    }


async def _resolve_account_by_name(db: Prisma, account_name: str):
    """
    Return the active PayoneerAccount whose name matches ``account_name``
    (case-insensitive).  Raises HTTP 404 if not found.
    """
    account = await db.payoneeraccount.find_first(
        where={
            "accountName": {"equals": account_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not account:
        raise HTTPException(
            status_code=404,
            detail=f"Payoneer account '{account_name}' not found.",
        )
    return account


async def _get_account_or_404(db: Prisma, account_id: str):
    account = await db.payoneeraccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Payoneer account not found.")
    return account


async def _get_transaction_or_404(db: Prisma, transaction_id: str):
    tx = await db.payoneertransaction.find_unique(where={"id": transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Payoneer transaction not found.")
    return tx


async def _latest_balance(db: Prisma, account_id: str) -> Decimal:
    """
    Fetch the current running balance for ``account_id`` by reading the
    ``remainingBalance`` of the most-recent transaction.

    Sort order: ``date DESC, id DESC``
    ────────────────────────────────────────────────────────────────────────────
    ``id`` is a CUID — lexicographically monotonic, so it reliably breaks ties
    when two or more transactions share the same calendar date (e.g. bulk
    imports, same-day opening balance + first transaction).  Without the
    secondary sort, ``find_first`` returns an arbitrary row among same-day ties,
    which makes ``currentBalance`` and ``totalBalance`` non-deterministic.

    Returns ``Decimal("0")`` when no transactions exist yet.
    """
    latest_tx = await db.payoneertransaction.find_first(
        where={"accountId": account_id},
        order=[{"date": "desc"}, {"id": "desc"}],
    )
    return _d(latest_tx.remainingBalance) if latest_tx else _ZERO


# ── Account CRUD ──────────────────────────────────────────────────────────────

async def create_account(db: Prisma, data: PayoneerAccountCreate) -> dict:
    """
    Create a Payoneer account.  If ``initial_balance`` is provided, an
    opening credit transaction is inserted immediately.
    """
    existing = await db.payoneeraccount.find_unique(
        where={"accountName": data.accountName}
    )
    if existing:
        raise HTTPException(status_code=409, detail="Account name already exists.")

    account = await db.payoneeraccount.create(
        data={"accountName": data.accountName}
    )

    opening_tx: Optional[dict] = None
    if data.initial_balance is not None and data.initial_balance > 0:
        today = date.today()
        tx = await db.payoneertransaction.create(
            data={
                "accountId":        account.id,
                "date":             datetime.combine(today, time.min),
                "details":          data.opening_note,
                "accountFrom":      None,
                "accountTo":        data.accountName,
                "debit":            Decimal("0"),
                "credit":           data.initial_balance,
                "remainingBalance": data.initial_balance,
            }
        )
        opening_tx = _tx_to_dict(tx, account.accountName)

    await _ensure_payoneer_timestamps(db)
    await _touch_account_updated_at(db, account.id)
    ts = await _fetch_account_timestamps(db, account.id)
    return {
        "id":                account.id,
        "accountName":       account.accountName,
        "isActive":          account.isActive,
        "createdAt":         ts["createdAt"],
        "updatedAt":         ts["updatedAt"],
        "currentBalance":    float(data.initial_balance or _ZERO),
        "openingTransaction": opening_tx,
    }


async def update_account(
    db: Prisma,
    account_id: str,
    data: PayoneerAccountUpdate,
) -> dict:
    """
    PATCH /accounts/{id} — partial account update (v4).

    Handles two independent concerns in one call:

    1. Account metadata  — rename (with uniqueness check) and/or isActive toggle.
    2. Balance adjustment — if ``initial_balance`` is supplied, a new credit
       transaction is appended to the ledger with the supplied (or default)
       ``opening_note`` as its details text.

    ``description`` is stored in the adjustment transaction's ``details`` field
    when ``initial_balance`` is also provided; otherwise it is returned in the
    response as an acknowledged note (the schema has no dedicated DB column for
    account-level description — the transaction ledger is the source of truth).

    Returns the updated account dict plus the adjustment transaction (if any).
    """
    account = await _get_account_or_404(db, account_id)

    # ── 1. Account metadata patch ─────────────────────────────────────────────
    account_patch: dict = {}

    if data.accountName is not None and data.accountName != account.accountName:
        conflict = await db.payoneeraccount.find_first(
            where={
                "accountName": {"equals": data.accountName, "mode": "insensitive"},
                "id":          {"not": account_id},
            }
        )
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Account name '{data.accountName}' is already taken.",
            )
        account_patch["accountName"] = data.accountName

    if data.isActive is not None:
        account_patch["isActive"] = data.isActive

    if account_patch:
        account = await db.payoneeraccount.update(
            where={"id": account_id},
            data=account_patch,
        )

    # ── 2. Balance-adjustment transaction (v4) ────────────────────────────────
    adjustment_tx: Optional[dict] = None

    if data.initial_balance is not None and data.initial_balance > 0:
        # Determine the running balance to use as the new remainingBalance.
        # We take the latest transaction's remainingBalance and add the credit.
        current_balance = await _latest_balance(db, account_id)
        new_balance     = current_balance + data.initial_balance

        # Build the details string — prefer explicit description over opening_note
        details_text = (
            data.description
            or data.opening_note
            or "Balance adjustment"
        )

        tx = await db.payoneertransaction.create(
            data={
                "accountId":        account_id,
                "date":             datetime.combine(date.today(), time.min),
                "details":          details_text,
                "accountFrom":      None,
                "accountTo":        account.accountName,
                "debit":            _ZERO,
                "credit":           data.initial_balance,
                "remainingBalance": new_balance,
            }
        )
        adjustment_tx = _tx_to_dict(tx, account.accountName)

    await _ensure_payoneer_timestamps(db)
    await _touch_account_updated_at(db, account.id)
    ts = await _fetch_account_timestamps(db, account.id)
    return {
        "id":                    account.id,
        "accountName":           account.accountName,
        "isActive":              account.isActive,
        "createdAt":             ts["createdAt"],
        "updatedAt":             ts["updatedAt"],
        # Echo back description so the caller can confirm it was received,
        # even when no transaction was created.
        "description":           data.description,
        "adjustmentTransaction": adjustment_tx,
    }


async def list_accounts(
    db: Prisma,
    filters: DateRangeFilter,
    name: Optional[str] = None,
    pagination: Optional[PageParams] = None,
) -> dict:
    """
    Combined totals + paginated per-account breakdown  (v8).

    ``name``   — case-insensitive partial search on accountName.

    Totals guarantee (v6 fix)
    ─────────────────────────
    totalBalance, totalCredit, totalDebit, totalTransactions are computed
    across **all** matching accounts before pagination is applied, so page 2+
    always returns the same correct aggregate as page 1.

    ``_latest_balance`` results are cached per account so the same DB query
    is never issued twice in a single request.
    """
    await _ensure_payoneer_timestamps(db)
    date_f      = filters.to_prisma_filter()
    date_f_where = {"date": date_f} if date_f else {}

    where: dict = {"isActive": True}
    if name:
        where["accountName"] = {"contains": name, "mode": "insensitive"}

    total_accounts = await db.payoneeraccount.count(where=where)

    # ── Fetch ALL matching accounts with their period-filtered transactions ────
    # Must happen before pagination so totals cover the full matching set.
    all_accounts = await db.payoneeraccount.find_many(
        where=where,
        include={
            "transactions": {
                "where":    date_f_where,
                "order_by": {"date": "desc"},
            },
        },
    )

    # ── Pre-fetch latest all-time balance per account — cached to avoid N×2 ──
    account_balances: dict[str, Decimal] = {}
    for acc in all_accounts:
        account_balances[acc.id] = await _latest_balance(db, acc.id)

    # ── Cross-account totals (full matching set, not paginated) ───────────────
    t_balance = t_credit = t_debit = _ZERO
    t_txcount = 0

    for acc in all_accounts:
        bal           = account_balances[acc.id]
        t_balance    += bal
        period_credit = sum((_d(t.credit) for t in acc.transactions), _ZERO)
        period_debit  = sum((_d(t.debit)  for t in acc.transactions), _ZERO)
        t_credit     += period_credit
        t_debit      += period_debit
        t_txcount    += len(acc.transactions)

    # ── Paginate in-process (totals are already captured above) ──────────────
    if pagination:
        page_slice = all_accounts[pagination.skip : pagination.skip + pagination.take]
    else:
        page_slice = all_accounts

    # Batch-fetch account timestamps for the current page — single query, no N+1
    page_ids  = [acc.id for acc in page_slice]
    ts_map    = await _fetch_accounts_timestamps_batch(db, page_ids)

    # Batch-fetch transaction timestamps for recent-5 rows across the page
    all_recent_ids = [
        t.id
        for acc in page_slice
        for t in acc.transactions[:5]
    ]
    tx_ts_map = await _fetch_tx_timestamps_batch(db, all_recent_ids)

    summaries: list[dict] = []
    for acc in page_slice:
        bal           = account_balances[acc.id]
        period_credit = sum((_d(t.credit) for t in acc.transactions), _ZERO)
        period_debit  = sum((_d(t.debit)  for t in acc.transactions), _ZERO)
        acc_ts        = ts_map.get(acc.id, {"createdAt": None, "updatedAt": None})

        summaries.append({
            "id":                acc.id,
            "accountName":       acc.accountName,
            "isActive":          acc.isActive,
            "createdAt":         acc_ts["createdAt"],
            "updatedAt":         acc_ts["updatedAt"],
            "currentBalance":    float(bal),
            "periodCredit":      float(period_credit),
            "periodDebit":       float(period_debit),
            "transactionCount":  len(acc.transactions),
            "recentTransactions": [
                _tx_to_dict(
                    t, acc.accountName,
                    created_at=tx_ts_map.get(t.id, {}).get("createdAt"),
                    updated_at=tx_ts_map.get(t.id, {}).get("updatedAt"),
                )
                for t in acc.transactions[:5]
            ],
        })

    return {
        "filter": filters.meta(),
        "totals": {
            "totalBalance":       float(t_balance),
            "totalCredit":        float(t_credit),
            "totalDebit":         float(t_debit),
            "totalTransactions":  t_txcount,
            "activeAccountCount": total_accounts,
        },
        "pagination": _pagination_meta(pagination, total_accounts),
        "accounts":   summaries,
    }


async def deactivate_account(db: Prisma, account_id: str) -> None:
    acc = await db.payoneeraccount.find_unique(where={"id": account_id})
    if not acc:
        raise HTTPException(status_code=404, detail="Payoneer account not found.")
    await db.payoneeraccount.update(
        where={"id": account_id}, data={"isActive": False}
    )


# ── Transaction CRUD ──────────────────────────────────────────────────────────

async def add_transaction(db: Prisma, data: PayoneerTransactionCreate) -> dict:
    """
    POST /transactions — add a ledger transaction (v6).

    Resolves account by ``data.account_name`` — staff never handle UUIDs.

    remainingBalance computation  (system-owned, always)
    ─────────────────────────────────────────────────────
    The service is the exclusive source of truth for the running balance.
    ``data.remaining_balance`` is accepted at the schema level for API
    compatibility but is **unconditionally ignored here**.

    The balance is ALWAYS computed as:
        new_balance = latest_balance + credit - debit

    This guarantees that:
    • ``currentBalance`` per account is always mathematically consistent.
    • ``totalBalance`` in GET /accounts is always correct.
    • A caller supplying ``0.00`` (or any other value) cannot accidentally
      corrupt the ledger.

    The combined totals in GET /accounts are computed live from the DB on every
    request — totalBalance, totalCredit, totalDebit, totalTransactions are
    updated immediately after this call completes with no extra step required.
    """
    account = await _resolve_account_by_name(db, data.account_name)

    # ── Always auto-compute remainingBalance — caller-supplied value ignored ──
    current_balance   = await _latest_balance(db, account.id)
    remaining_balance = current_balance + _d(data.credit) - _d(data.debit)

    # ── Persist ───────────────────────────────────────────────────────────────
    await _ensure_payoneer_timestamps(db)
    tx = await db.payoneertransaction.create(
        data={
            "accountId":        account.id,
            "date":             datetime.combine(data.date, time.min),
            "details":          data.details,
            "accountFrom":      data.accountFrom,
            "accountTo":        data.accountTo,
            "debit":            data.debit,
            "credit":           data.credit,
            "remainingBalance": remaining_balance,
        }
    )
    await _touch_tx_updated_at(db, tx.id)
    tx_ts = await _fetch_tx_timestamps_batch(db, [tx.id])
    ts    = tx_ts.get(tx.id, {"createdAt": None, "updatedAt": None})
    return _tx_to_dict(tx, account.accountName, created_at=ts["createdAt"], updated_at=ts["updatedAt"])


async def update_transaction(
    db: Prisma,
    transaction_id: str,
    data: PayoneerTransactionUpdate,
) -> dict:
    """
    PATCH /transactions/{id} — partial transaction update.

    • All fields are optional; only supplied ones are written.
    • remainingBalance is updated only when explicitly provided — the system
      does NOT auto-recompute it (ledger integrity is the caller's responsibility,
      mirroring the POST contract).
    • Returns the updated full transaction dict (with accountName).
    """
    tx      = await _get_transaction_or_404(db, transaction_id)
    account = await _get_account_or_404(db, tx.accountId)

    patch: dict = {}

    if data.date is not None:
        patch["date"] = datetime.combine(data.date, time.min)

    if data.details is not None:
        patch["details"] = data.details

    # accountFrom / accountTo: distinguish "not supplied" from "explicitly cleared"
    # Pydantic v2 passes None for both cases with Optional fields, so we rely on
    # model_fields_set to detect which keys the caller actually sent.
    sent = data.model_fields_set

    if "accountFrom" in sent:
        patch["accountFrom"] = data.accountFrom   # may be None → explicit clear

    if "accountTo" in sent:
        patch["accountTo"] = data.accountTo        # may be None → explicit clear

    if data.debit is not None:
        patch["debit"] = data.debit

    if data.credit is not None:
        patch["credit"] = data.credit

    if data.remaining_balance is not None:
        patch["remainingBalance"] = data.remaining_balance

    if not patch:
        return _tx_to_dict(tx, account.accountName)

    updated = await db.payoneertransaction.update(
        where={"id": transaction_id},
        data=patch,
    )
    return _tx_to_dict(updated, account.accountName)


async def get_account_transactions(
    db: Prisma,
    account_id: str,
    date_filter: dict,
    pagination: Optional[PageParams] = None,
    search: Optional[str] = None,
) -> dict:
    """
    Paginated transaction list for one account — every row includes accountName.

    ``search`` — case-insensitive keyword search on the ``details`` column.
    When supplied, only transactions whose details contain the keyword are returned.
    """
    await _ensure_payoneer_timestamps(db)
    account = await db.payoneeraccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Payoneer account not found.")

    where: dict = {"accountId": account_id}
    if date_filter:
        where["date"] = date_filter
    if search:
        where["details"] = {"contains": search, "mode": "insensitive"}

    total   = await db.payoneertransaction.count(where=where)
    find_kw = dict(where=where, order={"date": "desc"})
    if pagination:
        find_kw["skip"] = pagination.skip
        find_kw["take"] = pagination.take

    transactions = await db.payoneertransaction.find_many(**find_kw)

    # Current balance from latest tx across all time
    current_balance = await _latest_balance(db, account_id)

    period_credit = sum((_d(t.credit) for t in transactions), _ZERO)
    period_debit  = sum((_d(t.debit)  for t in transactions), _ZERO)

    # Batch-fetch timestamps for the current page — single query, no N+1
    tx_ids    = [t.id for t in transactions]
    tx_ts_map = await _fetch_tx_timestamps_batch(db, tx_ids)

    # Fetch account timestamps
    acc_ts = await _fetch_account_timestamps(db, account_id)

    return {
        "account": {
            "id":          account.id,
            "accountName": account.accountName,
            "isActive":    account.isActive,
            "createdAt":   acc_ts["createdAt"],
            "updatedAt":   acc_ts["updatedAt"],
        },
        "currentBalance": float(current_balance),
        "periodCredit":   float(period_credit),
        "periodDebit":    float(period_debit),
        "pagination":     _pagination_meta(pagination, total),
        "transactions":   [
            _tx_to_dict(
                t, account.accountName,
                created_at=tx_ts_map.get(t.id, {}).get("createdAt"),
                updated_at=tx_ts_map.get(t.id, {}).get("updatedAt"),
            )
            for t in transactions
        ],
    }


async def get_account_detail(
    db: Prisma,
    account_id: str,
    filters: DateRangeFilter,
    pagination: Optional[PageParams] = None,
    search: Optional[str] = None,
) -> dict:
    """
    GET /payoneer/accounts/{account_id}
    ─────────────────────────────────────
    Returns a full detail view for a single Payoneer account:

    - Account metadata (id, accountName, isActive, createdAt, updatedAt)
    - currentBalance — latest remainingBalance across all time
    - periodCredit / periodDebit — sums within the selected filter window
    - Paginated transactions in the window (newest first)

    ``search`` — case-insensitive keyword search on the transaction ``details`` column.
    Each transaction row includes accountName, createdAt, and updatedAt.
    """
    await _ensure_payoneer_timestamps(db)

    # ── Resolve account ───────────────────────────────────────────────────────
    account = await db.payoneeraccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Payoneer account not found.")

    date_f = filters.to_prisma_filter()
    acc_ts = await _fetch_account_timestamps(db, account_id)

    # ── Current balance — latest transaction across all time ──────────────────
    current_balance = await _latest_balance(db, account_id)

    # ── Period-scoped WHERE clause ────────────────────────────────────────────
    period_where: dict = {"accountId": account_id}
    if date_f:
        period_where["date"] = date_f
    if search:
        period_where["details"] = {"contains": search, "mode": "insensitive"}

    # ── Period aggregates ─────────────────────────────────────────────────────
    period_txs    = await db.payoneertransaction.find_many(where=period_where)
    period_credit = sum(float(t.credit) for t in period_txs)
    period_debit  = sum(float(t.debit)  for t in period_txs)
    total_count   = len(period_txs)

    # ── Paginated transaction list ────────────────────────────────────────────
    find_kw: dict = dict(where=period_where, order={"date": "desc"})
    if pagination:
        find_kw["skip"] = pagination.skip
        find_kw["take"] = pagination.take

    transactions = await db.payoneertransaction.find_many(**find_kw)

    # Batch-fetch tx timestamps for this page — single query, no N+1
    tx_ids    = [t.id for t in transactions]
    tx_ts_map = await _fetch_tx_timestamps_batch(db, tx_ids)

    tx_rows = [
        {
            "id":               t.id,
            "accountId":        t.accountId,
            "accountName":      account.accountName,
            "date":             t.date.date() if isinstance(t.date, datetime) else t.date,
            "details":          t.details,
            "accountFrom":      t.accountFrom,
            "accountTo":        t.accountTo,
            "debit":            float(t.debit),
            "credit":           float(t.credit),
            "remainingBalance": float(t.remainingBalance),
            "createdAt":        tx_ts_map.get(t.id, {}).get("createdAt"),
            "updatedAt":        tx_ts_map.get(t.id, {}).get("updatedAt"),
        }
        for t in transactions
    ]

    return {
        "filter": filters.meta(),
        "account": {
            "id":          account.id,
            "accountName": account.accountName,
            "isActive":    account.isActive,
            "createdAt":   acc_ts["createdAt"],
            "updatedAt":   acc_ts["updatedAt"],
        },
        "currentBalance": float(current_balance),
        "periodCredit":   period_credit,
        "periodDebit":    period_debit,
        "pagination":     _pagination_meta(pagination, total_count),
        "transactions":   tx_rows,
    }


async def delete_transaction(db: Prisma, transaction_id: str) -> None:
    tx = await db.payoneertransaction.find_unique(where={"id": transaction_id})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found.")
    await db.payoneertransaction.delete(where={"id": transaction_id})


# ── Profile-level Excel export ────────────────────────────────────────────────

async def export_account_excel(
    db: Prisma,
    account_id: str,
    filters: DateRangeFilter,
) -> tuple[bytes, str]:
    """Single-account Excel export — all transactions for the period."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot generate Excel export.",
        )

    detail   = await get_account_transactions(db, account_id, filters.to_prisma_filter())
    acc_name = detail["account"]["accountName"]
    start, end = filters.window()

    wb          = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")

    def _header(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def _autofit(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    ws = wb.active
    ws.title = "Transactions"
    _header(ws, [
        "Date", "Details", "Account From", "Account To",
        "Debit ($)", "Credit ($)", "Balance ($)",
    ])
    for ri, t in enumerate(detail["transactions"], 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([
            str(t["date"]), t["details"], t["accountFrom"] or "",
            t["accountTo"] or "", float(t["debit"]),
            float(t["credit"]), float(t["remainingBalance"]),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if fill:
                cell.fill = fill
    _autofit(ws)

    buf      = io.BytesIO()
    wb.save(buf)
    tag      = f"{start}_{end}" if start else "all"
    filename = f"payoneer_{acc_name.replace(' ', '_')}_{tag}.xlsx"
    return buf.getvalue(), filename