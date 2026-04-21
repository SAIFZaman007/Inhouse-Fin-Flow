"""
app/modules/fiverr/service.py
════════════════════════════════════════════════════════════════════════════════
v3 — Enterprise Edition
════════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import io
import logging
import math
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Optional

from fastapi import HTTPException
from prisma import Prisma

from app.core import trash_store
from app.shared.filters import DateRangeFilter
from app.shared.pagination import PageParams
from .schema import (
    FiverrOrderCreate,
    FiverrOrderUpdate,
    FiverrProfileCreate,
    FiverrProfileUpdate,
    FiverrSnapshotCreate,
    FiverrSnapshotUpdate,
)

logger = logging.getLogger(__name__)

_FIVERR_FEE = Decimal("0.20")
_AFTER_RATE = Decimal("1") - _FIVERR_FEE   # 0.80
_ZERO       = Decimal("0")

# Snapshot field names present on FiverrProfileUpdate (for presence check)
_SNAPSHOT_FIELDS = frozenset({
    "available_withdraw",
    "not_cleared",
    "active_orders",
    "active_order_amount",
    "withdrawn",
    "seller_plus",
    "promotion",
})

# ─────────────────────────────────────────────────────────────────────────────
# § TS  Timestamp bootstrap — adds createdAt / updatedAt to Fiverr tables
#       that don't have them in schema.prisma (cannot touch the schema).
#       Each ALTER is a separate execute_raw call — PostgreSQL extended-query
#       protocol forbids multiple commands per prepared statement.
# ─────────────────────────────────────────────────────────────────────────────

_FIVERR_TS_DDL: list[str] = [
    # fiverr_profiles: no timestamps at all in Prisma schema
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='fiverr_profiles' AND column_name='created_at'
      ) THEN
        ALTER TABLE fiverr_profiles ADD COLUMN created_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='fiverr_profiles' AND column_name='updated_at'
      ) THEN
        ALTER TABLE fiverr_profiles ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    # fiverr_entries: Prisma has createdAt, missing updatedAt
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='fiverr_entries' AND column_name='updated_at'
      ) THEN
        ALTER TABLE fiverr_entries ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    # fiverr_orders: Prisma has createdAt, missing updatedAt
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='fiverr_orders' AND column_name='updated_at'
      ) THEN
        ALTER TABLE fiverr_orders ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
]

_fiverr_ts_done = False


async def _ensure_fiverr_timestamps(db: Prisma) -> None:
    """
    Idempotent bootstrap — runs once per process lifetime.
    Adds createdAt / updatedAt columns to fiverr tables where Prisma
    didn't generate them.  Each DDL statement is a separate execute_raw
    call so PostgreSQL's extended-query protocol is never violated.
    """
    global _fiverr_ts_done
    if _fiverr_ts_done:
        return
    for stmt in _FIVERR_TS_DDL:
        await db.execute_raw(stmt)
    _fiverr_ts_done = True


async def _fetch_profile_timestamps(db: Prisma, profile_id: str) -> dict:
    """
    Read the raw created_at / updated_at columns from fiverr_profiles.
    Falls back to None gracefully before the bootstrap has run.
    """
    try:
        rows = await db.query_raw(
            "SELECT created_at, updated_at FROM fiverr_profiles WHERE id = $1",
            profile_id,
        )
        if rows:
            return {
                "createdAt": rows[0].get("created_at"),
                "updatedAt": rows[0].get("updated_at"),
            }
    except Exception:
        pass
    return {"createdAt": None, "updatedAt": None}


async def _fetch_entry_updated_at(db: Prisma, entry_id: str):
    """Read the raw updated_at column from fiverr_entries."""
    try:
        rows = await db.query_raw(
            "SELECT updated_at FROM fiverr_entries WHERE id = $1",
            entry_id,
        )
        if rows:
            return rows[0].get("updated_at")
    except Exception:
        pass
    return None


async def _fetch_order_updated_at(db: Prisma, order_id: str):
    """Read the raw updated_at column from fiverr_orders."""
    try:
        rows = await db.query_raw(
            "SELECT updated_at FROM fiverr_orders WHERE id = $1",
            order_id,
        )
        if rows:
            return rows[0].get("updated_at")
    except Exception:
        pass
    return None


async def _touch_entry_updated_at(db: Prisma, entry_id: str) -> None:
    """Bump updated_at on a FiverrEntry row after any write."""
    try:
        await db.execute_raw(
            "UPDATE fiverr_entries SET updated_at = now() WHERE id = $1",
            entry_id,
        )
    except Exception:
        pass


async def _touch_order_updated_at(db: Prisma, order_id: str) -> None:
    """Bump updated_at on a FiverrOrder row after any write."""
    try:
        await db.execute_raw(
            "UPDATE fiverr_orders SET updated_at = now() WHERE id = $1",
            order_id,
        )
    except Exception:
        pass




# ── Private helpers ───────────────────────────────────────────────────────────

def _d(v: Any) -> Decimal:
    return _ZERO if v is None else Decimal(str(v))


def _after_fee(amount: Any) -> Decimal:
    return (_d(amount) * _AFTER_RATE).quantize(Decimal("0.01"))


def _entry_to_dict(entry: Any, profile_name: str, updated_at: Any = None) -> dict:
    """
    Serialise a FiverrEntry ORM object to a plain dict.

    ``updated_at`` — pass the value fetched from the raw fiverr_entries.updated_at
    column.  Defaults to None when the caller does not have a DB handle (e.g.
    inside list builders that don't need per-row round-trips).
    """
    return {
        "id":                 entry.id,
        "profileId":          entry.profileId,
        "profileName":        profile_name,
        "date":               entry.date.date() if isinstance(entry.date, datetime) else entry.date,
        "availableWithdraw":  float(_d(entry.availableWithdraw)),
        "notCleared":         float(_d(entry.notCleared)),
        "activeOrders":       entry.activeOrders,
        "activeOrderAmount":  float(_d(entry.activeOrderAmount)),
        "withdrawn":          float(_d(entry.withdrawn)),
        "sellerPlus":         entry.sellerPlus,
        "promotion":          float(_d(entry.promotion)),
        "createdAt":          entry.createdAt,
        "updatedAt":          updated_at,
    }


def _order_to_dict(order: Any, updated_at: Any = None) -> dict:
    """
    Serialise a FiverrOrder ORM object to a plain dict.

    ``updated_at`` — pass the value fetched from the raw fiverr_orders.updated_at
    column.  Defaults to None for list-context calls.
    """
    return {
        "id":          order.id,
        "profileId":   order.profileId,
        "date":        order.date.date() if isinstance(order.date, datetime) else order.date,
        "buyerName":   order.buyerName,
        "orderId":     order.orderId,
        "amount":      float(_d(order.amount)),
        "afterFiverr": float(_d(order.afterFiverr)),
        "createdAt":   order.createdAt,
        "updatedAt":   updated_at,
    }


def _pagination_meta(pagination: Optional[PageParams], total: int) -> dict:
    if pagination is None:
        return {"page": 1, "pageSize": total, "total": total, "totalPages": 1}
    return {
        "page":       pagination.page,
        "pageSize":   pagination.page_size,
        "total":      total,
        "totalPages": math.ceil(total / pagination.page_size) if total > 0 else 1,
    }


async def _get_profile_or_404(db: Prisma, profile_id: str):
    profile = await db.fiverrprofile.find_unique(where={"id": profile_id})
    if not profile:
        raise HTTPException(status_code=404, detail="Fiverr profile not found.")
    return profile


async def _resolve_profile_by_name(db: Prisma, profile_name: str):
    profile = await db.fiverrprofile.find_first(
        where={
            "profileName": {"equals": profile_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not profile:
        raise HTTPException(
            status_code=404,
            detail=f"Fiverr profile '{profile_name}' not found.",
        )
    return profile


# ── Profile CRUD ──────────────────────────────────────────────────────────────

async def create_profile(db: Prisma, data: FiverrProfileCreate) -> dict:
    if not data.profileName:
        raise HTTPException(status_code=422, detail="profileName is required.")

    existing = await db.fiverrprofile.find_unique(
        where={"profileName": data.profileName}
    )
    if existing:
        raise HTTPException(status_code=409, detail="Profile name already exists.")

    profile = await db.fiverrprofile.create(
        data={"profileName": data.profileName}
    )

    snapshot: Optional[dict] = None
    has_snapshot_fields = any([
        data.available_withdraw is not None,
        data.not_cleared is not None,
        data.active_orders is not None,
        data.active_order_amount is not None,
        data.withdrawn is not None,
        data.promotion is not None,
    ])

    if has_snapshot_fields:
        snap_date = data.snapshot_date or date.today()
        snap_dt   = datetime.combine(snap_date, time.min)
        entry = await db.fiverrentry.create(
            data={
                "profileId":         profile.id,
                "date":              snap_dt,
                "availableWithdraw": data.available_withdraw or _ZERO,
                "notCleared":        data.not_cleared        or _ZERO,
                "activeOrders":      data.active_orders      or 0,
                "activeOrderAmount": data.active_order_amount or _ZERO,
                "submitted":         _ZERO,
                "withdrawn":         data.withdrawn          or _ZERO,
                "sellerPlus":        data.seller_plus,
                "promotion":         data.promotion          or _ZERO,
            }
        )
        snapshot = _entry_to_dict(entry, profile.profileName)

    await _ensure_fiverr_timestamps(db)
    ts = await _fetch_profile_timestamps(db, profile.id)
    return {
        "id":              profile.id,
        "profileName":     profile.profileName,
        "isActive":        profile.isActive,
        "createdAt":       ts["createdAt"],
        "updatedAt":       ts["updatedAt"],
        "initialSnapshot": snapshot,
    }


async def update_profile(
    db: Prisma,
    profile_id: str,
    data: FiverrProfileUpdate,
) -> dict:
    """
    PATCH /profiles/{id} — partial profile update.

    1. Profile metadata  — rename (with uniqueness check) and/or isActive toggle.
    2. Snapshot upsert   — if any snapshot field is supplied, upsert the
                           FiverrEntry for ``snapshot_date`` (defaults to today).
    """
    profile = await _get_profile_or_404(db, profile_id)

    profile_patch: dict = {}

    if data.profileName is not None and data.profileName != profile.profileName:
        conflict = await db.fiverrprofile.find_first(
            where={
                "profileName": {"equals": data.profileName, "mode": "insensitive"},
                "id":          {"not": profile_id},
            }
        )
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Profile name '{data.profileName}' is already taken.",
            )
        profile_patch["profileName"] = data.profileName

    if data.isActive is not None:
        profile_patch["isActive"] = data.isActive

    if profile_patch:
        profile = await db.fiverrprofile.update(
            where={"id": profile_id},
            data=profile_patch,
        )

    sent            = data.model_fields_set
    snapshot_fields = sent & _SNAPSHOT_FIELDS
    upserted_snapshot: Optional[dict] = None

    if snapshot_fields:
        snap_date = data.snapshot_date or date.today()
        snap_dt   = datetime.combine(snap_date, time.min)

        existing_entry = await db.fiverrentry.find_unique(
            where={"profileId_date": {"profileId": profile_id, "date": snap_dt}}
        )

        def _pick(field: str, existing_attr: str, default: Any) -> Any:
            if field in sent:
                return getattr(data, field)
            if existing_entry is not None:
                return getattr(existing_entry, existing_attr)
            return default

        aw            = _pick("available_withdraw",  "availableWithdraw",  _ZERO)
        not_cleared   = _pick("not_cleared",         "notCleared",         _ZERO)
        active_orders = _pick("active_orders",       "activeOrders",       0)
        aoa           = _pick("active_order_amount", "activeOrderAmount",  _ZERO)
        withdrawn     = _pick("withdrawn",            "withdrawn",          _ZERO)
        seller_plus   = _pick("seller_plus",          "sellerPlus",         False)
        promotion     = _pick("promotion",            "promotion",          _ZERO)

        entry_data = {
            "availableWithdraw":  aw,
            "notCleared":         not_cleared,
            "activeOrders":       active_orders,
            "activeOrderAmount":  aoa,
            "submitted":          _ZERO,
            "withdrawn":          withdrawn,
            "sellerPlus":         seller_plus,
            "promotion":          promotion,
        }

        entry = await db.fiverrentry.upsert(
            where={"profileId_date": {"profileId": profile_id, "date": snap_dt}},
            data={
                "create": {"profileId": profile_id, "date": snap_dt, **entry_data},
                "update": entry_data,
            },
        )
        upserted_snapshot = _entry_to_dict(entry, profile.profileName)

    await _ensure_fiverr_timestamps(db)
    # Touch updated_at for the profile record
    try:
        await db.execute_raw(
            "UPDATE fiverr_profiles SET updated_at = now() WHERE id = $1",
            profile_id,
        )
    except Exception:
        pass
    ts = await _fetch_profile_timestamps(db, profile.id)
    return {
        "id":               profile.id,
        "profileName":      profile.profileName,
        "isActive":         profile.isActive,
        "createdAt":        ts["createdAt"],
        "updatedAt":        ts["updatedAt"],
        "snapshotUpserted": upserted_snapshot,
    }


async def soft_delete_profile(db: Prisma, profile_id: str) -> dict:
    """
    Full soft-delete for a Fiverr profile (v3).

    1. Fetch profile + all its entries + all its orders.
    2. Write the profile record to trash_store (type="profile").
    3. Write each entry to trash_store (type="snapshot").
    4. Write each order to trash_store (type="order").
    5. Set isActive=False on the profile.
    6. Return confirmation with counts.

    Dynamic totalActiveOrders is reduced accordingly because is_deleted()
    guards filter out these entries in the list/detail builders.
    """
    profile = await db.fiverrprofile.find_unique(
        where={"id": profile_id},
        include={"entries": True, "orders": True},
    )
    if not profile:
        raise HTTPException(status_code=404, detail="Fiverr profile not found.")

    # Capture profile for trash
    profile_snap = {
        "id":          profile.id,
        "profileName": profile.profileName,
        "isActive":    profile.isActive,
    }
    await trash_store.add(
        record_id=profile.id,
        module="fiverr",
        record_type="profile",
        snapshot=profile_snap,
    )

    # Archive each snapshot
    for entry in profile.entries:
        entry_snap = _entry_to_dict(entry, profile.profileName)
        await trash_store.add(
            record_id=entry.id,
            module="fiverr",
            record_type="snapshot",
            snapshot=entry_snap,
        )

    # Archive each order
    for order in profile.orders:
        order_snap = _order_to_dict(order)
        await trash_store.add(
            record_id=order.id,
            module="fiverr",
            record_type="order",
            snapshot=order_snap,
        )

    # Soft-delete at DB level
    await db.fiverrprofile.update(
        where={"id": profile_id}, data={"isActive": False}
    )

    return {
        "success":           True,
        "message":           "Fiverr profile has been soft-deleted.",
        "profileId":         profile_id,
        "profileName":       profile.profileName,
        "snapshotsArchived": len(profile.entries),
        "ordersArchived":    len(profile.orders),
    }


# ── Snapshot CRUD ─────────────────────────────────────────────────────────────

async def create_snapshot(db: Prisma, data: FiverrSnapshotCreate) -> dict:
    """
    POST /snapshots — additive daily snapshot.

    Accumulation behaviour:
    • First submission for (profile_name, date) → plain INSERT.
    • Subsequent submission for the same pair   → ADD incoming values to stored.
      seller_plus uses OR semantics (sticky True).

    ``submitted`` is always written as Decimal("0") (service-internal).
    """
    if not data.profile_name:
        raise HTTPException(status_code=422, detail="profile_name is required.")
    if not data.date:
        raise HTTPException(status_code=422, detail="date is required.")

    profile = await _resolve_profile_by_name(db, data.profile_name)
    snap_dt = datetime.combine(data.date, time.min)

    existing = await db.fiverrentry.find_unique(
        where={"profileId_date": {"profileId": profile.id, "date": snap_dt}}
    )

    # Guard: if this entry is in trash, reject
    if existing and trash_store.is_deleted(existing.id):
        raise HTTPException(
            status_code=409,
            detail="A snapshot for this date was soft-deleted. Restore it first via POST /restore-trash.",
        )

    aw_in  = data.available_withdraw  or _ZERO
    nc_in  = data.not_cleared         or _ZERO
    ao_in  = data.active_orders       or 0
    aoa_in = data.active_order_amount or _ZERO
    wd_in  = data.withdrawn           or _ZERO
    sp_in  = data.seller_plus         or False
    pr_in  = data.promotion           or _ZERO

    if existing is None:
        entry = await db.fiverrentry.create(
            data={
                "profileId":         profile.id,
                "date":              snap_dt,
                "availableWithdraw": aw_in,
                "notCleared":        nc_in,
                "activeOrders":      ao_in,
                "activeOrderAmount": aoa_in,
                "submitted":         _ZERO,
                "withdrawn":         wd_in,
                "sellerPlus":        sp_in,
                "promotion":         pr_in,
            }
        )
    else:
        entry = await db.fiverrentry.update(
            where={"profileId_date": {"profileId": profile.id, "date": snap_dt}},
            data={
                "availableWithdraw": _d(existing.availableWithdraw) + aw_in,
                "notCleared":        _d(existing.notCleared)        + nc_in,
                "activeOrders":      existing.activeOrders           + ao_in,
                "activeOrderAmount": _d(existing.activeOrderAmount)  + aoa_in,
                "submitted":         _ZERO,
                "withdrawn":         _d(existing.withdrawn)          + wd_in,
                "sellerPlus":        existing.sellerPlus              or sp_in,
                "promotion":         _d(existing.promotion)          + pr_in,
            },
        )

    # Aggregate totals for the response
    all_orders  = await db.fiverrorder.find_many(where={"profileId": profile.id})
    live_orders = [o for o in all_orders if not trash_store.is_deleted(o.id)]
    total_revenue       = sum((_d(o.afterFiverr) for o in live_orders), _ZERO)
    total_order_amount  = sum((_d(o.amount)      for o in live_orders), _ZERO)

    await _ensure_fiverr_timestamps(db)
    await _touch_entry_updated_at(db, entry.id)
    entry_updated_at = await _fetch_entry_updated_at(db, entry.id)
    snapshot_dict = _entry_to_dict(entry, profile.profileName, updated_at=entry_updated_at)
    return {
        **snapshot_dict,
        "syncedTotals": {
            "orderRevenueAllTime":  float(total_revenue),
            "orderAmountAllTime":   float(total_order_amount),
            "latestSnapshot": {
                "availableWithdraw":  float(_d(entry.availableWithdraw)),
                "notCleared":         float(_d(entry.notCleared)),
                "activeOrders":       entry.activeOrders,
                "activeOrderAmount":  float(_d(entry.activeOrderAmount)),
                "withdrawn":          float(_d(entry.withdrawn)),
                "promotion":          float(_d(entry.promotion)),
            },
        },
    }


async def update_snapshot(
    db: Prisma,
    snapshot_id: str,
    data: FiverrSnapshotUpdate,
) -> dict:
    """
    PATCH /snapshots/{id} — partial update with SET semantics (v3).

    Only supplied fields are overwritten — not accumulated.
    Correction-mode, not accumulation-mode.
    """
    entry = await db.fiverrentry.find_unique(where={"id": snapshot_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Fiverr snapshot not found.")

    if trash_store.is_deleted(snapshot_id):
        raise HTTPException(
            status_code=410,
            detail="This snapshot has been soft-deleted. Restore it first via POST /restore-trash.",
        )

    patch: dict = {}
    sent = data.model_fields_set

    if "available_withdraw" in sent and data.available_withdraw is not None:
        patch["availableWithdraw"] = data.available_withdraw
    if "not_cleared" in sent and data.not_cleared is not None:
        patch["notCleared"] = data.not_cleared
    if "active_orders" in sent and data.active_orders is not None:
        patch["activeOrders"] = data.active_orders
    if "active_order_amount" in sent and data.active_order_amount is not None:
        patch["activeOrderAmount"] = data.active_order_amount
    if "withdrawn" in sent and data.withdrawn is not None:
        patch["withdrawn"] = data.withdrawn
    if "seller_plus" in sent and data.seller_plus is not None:
        patch["sellerPlus"] = data.seller_plus
    if "promotion" in sent and data.promotion is not None:
        patch["promotion"] = data.promotion

    await _ensure_fiverr_timestamps(db)
    if not patch:
        profile        = await _get_profile_or_404(db, entry.profileId)
        existing_upd   = await _fetch_entry_updated_at(db, entry.id)
        return _entry_to_dict(entry, profile.profileName, updated_at=existing_upd)

    updated = await db.fiverrentry.update(
        where={"id": snapshot_id},
        data=patch,
    )
    await _touch_entry_updated_at(db, updated.id)
    upd_ts  = await _fetch_entry_updated_at(db, updated.id)
    profile = await _get_profile_or_404(db, updated.profileId)
    return _entry_to_dict(updated, profile.profileName, updated_at=upd_ts)


async def soft_delete_snapshot(db: Prisma, snapshot_id: str) -> dict:
    """
    DELETE /snapshots/{id} — soft-delete a Fiverr snapshot (v3).

    1. Fetch snapshot + owning profile.
    2. Write full snapshot dict to trash_store (module="fiverr", type="snapshot").
    3. DB row remains — excluded from all live calculations via is_deleted() guards.
    4. Return confirmation + trash item.
    """
    entry = await db.fiverrentry.find_unique(where={"id": snapshot_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Fiverr snapshot not found.")

    if trash_store.is_deleted(snapshot_id):
        raise HTTPException(status_code=409, detail="Snapshot is already in trash.")

    profile   = await _get_profile_or_404(db, entry.profileId)
    snap_dict = _entry_to_dict(entry, profile.profileName)

    trash_item = await trash_store.add(
        record_id=snapshot_id,
        module="fiverr",
        record_type="snapshot",
        snapshot=snap_dict,
    )

    return {
        "success":   True,
        "message":   "Fiverr snapshot has been soft-deleted.",
        "trashItem": trash_item,
    }


async def get_profile_snapshots(
    db: Prisma,
    profile_id: str,
    date_filter: dict,
    pagination: Optional[PageParams] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    where: dict = {"profileId": profile_id}
    if date_filter:
        where["date"] = date_filter

    all_entries  = await db.fiverrentry.find_many(where=where, order={"date": "desc"})
    live_entries = [e for e in all_entries if not trash_store.is_deleted(e.id)]
    total        = len(live_entries)

    if pagination:
        live_entries = live_entries[pagination.skip: pagination.skip + pagination.take]

    await _ensure_fiverr_timestamps(db)
    # Batch-fetch updated_at for the visible page (single query, no N+1)
    upd_map: dict = {}
    if live_entries:
        try:
            ids_sql   = ", ".join(f"${i+1}" for i in range(len(live_entries)))
            upd_rows  = await db.query_raw(
                f"SELECT id, updated_at FROM fiverr_entries WHERE id IN ({ids_sql})",
                *[e.id for e in live_entries],
            )
            upd_map = {r["id"]: r.get("updated_at") for r in upd_rows}
        except Exception:
            pass

    return {
        "profileId":   profile_id,
        "profileName": profile.profileName,
        "pagination":  _pagination_meta(pagination, total),
        "snapshots":   [
            _entry_to_dict(e, profile.profileName, updated_at=upd_map.get(e.id))
            for e in live_entries
        ],
    }


# ── Order CRUD ────────────────────────────────────────────────────────────────

async def add_order(db: Prisma, data: FiverrOrderCreate) -> dict:
    """
    POST /orders — log a new Fiverr order and sync the daily snapshot.

    1. Resolve active profile by name (case-insensitive).
    2. Guard against duplicate orderId.
    3. Persist FiverrOrder with server-computed afterFiverr (amount × 0.80).
    4. Additively update the FiverrEntry for (profile, date):
         activeOrders      += 1
         activeOrderAmount += order.amount
       Snapshot is upserted if it doesn't exist for that date.
    5. Return order dict + syncedTotals.

    totalActiveOrders is dynamic — computed from live entries in list builders.
    """
    if not data.profile_name:
        raise HTTPException(status_code=422, detail="profile_name is required.")
    if not data.date:
        raise HTTPException(status_code=422, detail="date is required.")
    if not data.buyer_name:
        raise HTTPException(status_code=422, detail="buyer_name is required.")
    if not data.order_id:
        raise HTTPException(status_code=422, detail="order_id is required.")
    if data.amount is None:
        raise HTTPException(status_code=422, detail="amount is required.")

    profile = await _resolve_profile_by_name(db, data.profile_name)

    existing_order = await db.fiverrorder.find_unique(where={"orderId": data.order_id})
    if existing_order:
        raise HTTPException(status_code=409, detail="Order ID already exists.")

    snap_dt = datetime.combine(data.date, time.min)

    order = await db.fiverrorder.create(
        data={
            "profileId":   profile.id,
            "date":        snap_dt,
            "buyerName":   data.buyer_name,
            "orderId":     data.order_id,
            "amount":      data.amount,
            "afterFiverr": _after_fee(data.amount),
        }
    )

    # ── Sync snapshot for this date ──────────────────────────────────────────
    existing_entry = await db.fiverrentry.find_unique(
        where={"profileId_date": {"profileId": profile.id, "date": snap_dt}}
    )

    if existing_entry is None:
        await db.fiverrentry.create(
            data={
                "profileId":         profile.id,
                "date":              snap_dt,
                "availableWithdraw": _ZERO,
                "notCleared":        _ZERO,
                "activeOrders":      1,
                "activeOrderAmount": _d(data.amount),
                "submitted":         _ZERO,
                "withdrawn":         _ZERO,
                "sellerPlus":        False,
                "promotion":         _ZERO,
            }
        )
    else:
        if not trash_store.is_deleted(existing_entry.id):
            await db.fiverrentry.update(
                where={"profileId_date": {"profileId": profile.id, "date": snap_dt}},
                data={
                    "activeOrders":      existing_entry.activeOrders + 1,
                    "activeOrderAmount": _d(existing_entry.activeOrderAmount) + _d(data.amount),
                },
            )

    # ── Aggregate live totals ─────────────────────────────────────────────────
    all_orders  = await db.fiverrorder.find_many(where={"profileId": profile.id})
    live_orders = [o for o in all_orders if not trash_store.is_deleted(o.id)]

    all_entries  = await db.fiverrentry.find_many(where={"profileId": profile.id})
    live_entries = [e for e in all_entries if not trash_store.is_deleted(e.id)]

    total_revenue      = sum((_d(o.afterFiverr) for o in live_orders), _ZERO)
    total_order_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)
    total_active_orders = sum((e.activeOrders   for e in live_entries), 0)

    await _ensure_fiverr_timestamps(db)
    await _touch_order_updated_at(db, order.id)
    order_upd_at = await _fetch_order_updated_at(db, order.id)
    return {
        **_order_to_dict(order, updated_at=order_upd_at),
        "syncedTotals": {
            "orderCount":         len(live_orders),
            "totalActiveOrders":  total_active_orders,   # dynamic
            "revenueAllTime":     float(total_revenue),
            "orderAmountAllTime": float(total_order_amount),
        },
    }


async def update_order(
    db: Prisma,
    order_id: str,
    data: FiverrOrderUpdate,
) -> dict:
    order = await db.fiverrorder.find_unique(where={"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Fiverr order not found.")

    if trash_store.is_deleted(order_id):
        raise HTTPException(
            status_code=410,
            detail="This order has been soft-deleted. Restore it via POST /restore-trash.",
        )

    patch: dict = {}
    sent = data.model_fields_set

    if "date" in sent and data.date is not None:
        patch["date"] = datetime.combine(data.date, time.min)

    if data.buyer_name is not None:
        patch["buyerName"] = data.buyer_name

    if data.order_id is not None and data.order_id != order.orderId:
        conflict = await db.fiverrorder.find_unique(where={"orderId": data.order_id})
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Order ID '{data.order_id}' is already taken.",
            )
        patch["orderId"] = data.order_id

    if data.amount is not None:
        patch["amount"]      = data.amount
        patch["afterFiverr"] = _after_fee(data.amount)

    await _ensure_fiverr_timestamps(db)
    if not patch:
        existing_upd = await _fetch_order_updated_at(db, order.id)
        return _order_to_dict(order, updated_at=existing_upd)

    updated = await db.fiverrorder.update(where={"id": order_id}, data=patch)
    await _touch_order_updated_at(db, updated.id)
    upd_ts = await _fetch_order_updated_at(db, updated.id)
    return _order_to_dict(updated, updated_at=upd_ts)


async def get_profile_orders(
    db: Prisma,
    profile_id: str,
    date_filter: dict,
    pagination: Optional[PageParams] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    where: dict = {"profileId": profile_id}
    if date_filter:
        where["date"] = date_filter

    all_orders  = await db.fiverrorder.find_many(where=where, order={"date": "desc"})
    live_orders = [o for o in all_orders if not trash_store.is_deleted(o.id)]
    total       = len(live_orders)

    if pagination:
        live_orders = live_orders[pagination.skip: pagination.skip + pagination.take]

    await _ensure_fiverr_timestamps(db)
    # Batch-fetch updated_at for the visible page (single query, no N+1)
    ord_upd_map: dict = {}
    if live_orders:
        try:
            ids_sql     = ", ".join(f"${i+1}" for i in range(len(live_orders)))
            ord_upd_rows = await db.query_raw(
                f"SELECT id, updated_at FROM fiverr_orders WHERE id IN ({ids_sql})",
                *[o.id for o in live_orders],
            )
            ord_upd_map = {r["id"]: r.get("updated_at") for r in ord_upd_rows}
        except Exception:
            pass

    return {
        "profileId":   profile_id,
        "profileName": profile.profileName,
        "orderCount":  total,
        "pagination":  _pagination_meta(pagination, total),
        "orders":      [
            _order_to_dict(o, updated_at=ord_upd_map.get(o.id))
            for o in live_orders
        ],
    }


# ── Trash / Restore ───────────────────────────────────────────────────────────

async def get_trash(record_type: Optional[str] = None) -> dict:
    """
    GET /trash — return all Fiverr soft-deleted records.

    Optionally filter by ``record_type`` ("profile" | "snapshot" | "order").
    Results are sorted newest-deleted-first.
    """
    items = await trash_store.get_all(module="fiverr", record_type=record_type)
    return {
        "total": len(items),
        "items": items,
    }


async def restore_trash(db: Prisma, ids: list[str]) -> dict:
    """
    POST /restore-trash — restore one or more soft-deleted Fiverr records (v3).

    For each ID:
    • type="profile"  → set isActive=True in DB + remove from trash.
    • type="snapshot" → DB row still exists; remove from trash (re-appears in live calculations).
    • type="order"    → DB row still exists; remove from trash.
    """
    restored: list[str] = []
    failed:   list[str] = []

    for record_id in ids:
        try:
            item = await trash_store.get_by_id(record_id)
            if not item or item.get("module") != "fiverr":
                failed.append(record_id)
                continue

            record_type = item.get("type")

            if record_type == "profile":
                profile = await db.fiverrprofile.find_unique(where={"id": record_id})
                if profile:
                    await db.fiverrprofile.update(
                        where={"id": record_id},
                        data={"isActive": True},
                    )

            elif record_type in ("snapshot", "order"):
                # DB row was never deleted — just remove from trash registry
                pass

            removed = await trash_store.remove(record_id)
            if removed:
                restored.append(record_id)
            else:
                failed.append(record_id)

        except Exception as exc:
            logger.error("fiverr restore_trash: failed to restore %s — %s", record_id, exc)
            failed.append(record_id)

    total = len(restored)
    return {
        "restored": restored,
        "failed":   failed,
        "message":  f"{total} record(s) restored successfully." if total else "No records were restored.",
    }


# ── List / detail ─────────────────────────────────────────────────────────────

async def list_profiles_summary(
    db: Prisma,
    filters: DateRangeFilter,
    name: Optional[str] = None,
    pagination: Optional[PageParams] = None,
) -> dict:
    """
    GET /profiles — combined totals + paginated per-profile breakdown (v3).

    totalActiveOrders is DYNAMIC:
    • Computed as Σ entry.activeOrders across all live (non-trashed) snapshots.
    • Increases by +1 when an order is posted (via snapshot sync).
    • Decreases by -1 when an order or snapshot is soft-deleted.
    • Decreases by the profile's total activeOrders when a profile is soft-deleted.

    Soft-deleted snapshots and orders are excluded from all calculations.
    """
    await _ensure_fiverr_timestamps(db)
    date_f = filters.to_prisma_filter()

    where: dict = {"isActive": True}
    if name:
        where["profileName"] = {"contains": name, "mode": "insensitive"}

    total_profiles = await db.fiverrprofile.count(where=where)

    find_kw: dict = dict(
        where=where,
        include={
            "entries": {
                "where":    {"date": date_f} if date_f else {},
                "order_by": {"date": "desc"},
            },
            "orders": {
                "where":    {"date": date_f} if date_f else {},
                "order_by": {"date": "desc"},
            },
        },
    )
    if pagination:
        find_kw["skip"] = pagination.skip
        find_kw["take"] = pagination.take

    profiles = await db.fiverrprofile.find_many(**find_kw)

    t_avail        = t_not_cleared = t_aoa = t_withdrawn = t_promotion = _ZERO
    t_active_orders = 0
    t_revenue      = _ZERO    # Σ afterFiverr
    t_order_amount = _ZERO    # Σ order.amount

    summaries = []
    for p in profiles:
        # Exclude trashed snapshots and orders
        live_entries = [e for e in p.entries if not trash_store.is_deleted(e.id)]
        live_orders  = [o for o in p.orders  if not trash_store.is_deleted(o.id)]

        latest = live_entries[0] if live_entries else None

        aw      = _d(latest.availableWithdraw) if latest else _ZERO
        nc      = _d(latest.notCleared)        if latest else _ZERO
        ao      = latest.activeOrders           if latest else 0
        aoa     = _d(latest.activeOrderAmount)  if latest else _ZERO
        wdrawn  = _d(latest.withdrawn)          if latest else _ZERO
        promo   = _d(latest.promotion)          if latest else _ZERO

        period_revenue      = sum((_d(o.afterFiverr) for o in live_orders), _ZERO)
        period_order_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)

        # totalActiveOrders — sum of activeOrders across ALL live entries (not just latest)
        profile_active_orders = sum((e.activeOrders for e in live_entries), 0)

        t_avail         += aw
        t_not_cleared   += nc
        t_active_orders += profile_active_orders   # dynamic
        t_aoa           += aoa
        t_withdrawn     += wdrawn
        t_promotion     += promo
        t_revenue       += period_revenue
        t_order_amount  += period_order_amount

        period_totals = {
            "availableWithdraw":  float(aw),
            "notCleared":         float(nc),
            "activeOrders":       profile_active_orders,
            "activeOrderAmount":  float(aoa),
            "withdrawn":          float(wdrawn),
            "sellerPlus":         latest.sellerPlus if latest else False,
            "promotion":          float(promo),
            "revenueInPeriod":    float(period_revenue),
            "orderAmountInPeriod": float(period_order_amount),
        }

        p_ts = await _fetch_profile_timestamps(db, p.id)
        summaries.append({
            "id":              p.id,
            "profileName":     p.profileName,
            "isActive":        p.isActive,
            "createdAt":       p_ts["createdAt"],
            "updatedAt":       p_ts["updatedAt"],
            "latestSnapshot":  _entry_to_dict(latest, p.profileName) if latest else None,
            "periodTotals":    period_totals,
            "snapshotCount":   len(live_entries),
            "orderCount":      len(live_orders),
            "revenueInPeriod": float(period_revenue),
            "orders":          [_order_to_dict(o) for o in live_orders],
        })

    return {
        "filter": filters.meta(),
        "totals": {
            "totalAvailableWithdraw":  float(t_avail),
            "totalNotCleared":         float(t_not_cleared),
            "totalActiveOrders":       t_active_orders,     # DYNAMIC
            "totalActiveOrderAmount":  float(t_aoa),
            "totalWithdrawn":          float(t_withdrawn),
            "totalPromotion":          float(t_promotion),
            "totalRevenueInPeriod":    float(t_revenue),
            "totalOrderAmount":        float(t_order_amount),
            "activeProfileCount":      total_profiles,
        },
        "pagination": _pagination_meta(pagination, total_profiles),
        "profiles":   summaries,
    }


async def get_profile_detail(
    db: Prisma,
    profile_id: str,
    filters: DateRangeFilter,
    pagination: Optional[PageParams] = None,
    name: Optional[str] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    if name and name.lower() not in profile.profileName.lower():
        raise HTTPException(
            status_code=404,
            detail=f"Fiverr profile not found matching name '{name}'.",
        )

    date_f = filters.to_prisma_filter()

    snap_where:  dict = {"profileId": profile_id}
    order_where: dict = {"profileId": profile_id}
    if date_f:
        snap_where["date"]  = date_f
        order_where["date"] = date_f

    all_entries = await db.fiverrentry.find_many(where=snap_where, order={"date": "desc"})
    all_orders  = await db.fiverrorder.find_many(where=order_where, order={"date": "desc"})

    live_entries = [e for e in all_entries if not trash_store.is_deleted(e.id)]
    live_orders  = [o for o in all_orders  if not trash_store.is_deleted(o.id)]

    snap_total  = len(live_entries)
    order_total = len(live_orders)

    if pagination:
        live_entries = live_entries[pagination.skip: pagination.skip + pagination.take]
        live_orders  = live_orders[pagination.skip:  pagination.skip + pagination.take]

    latest         = live_entries[0] if live_entries else None
    aw             = _d(latest.availableWithdraw) if latest else _ZERO
    order_revenue  = sum((_d(o.afterFiverr) for o in live_orders), _ZERO)
    order_amount   = sum((_d(o.amount)      for o in live_orders), _ZERO)
    active_orders  = sum((e.activeOrders    for e in (live_entries if live_entries else [])), 0)

    return {
        "filter": filters.meta(),
        "profile": {
            "id":          profile.id,
            "profileName": profile.profileName,
            "isActive":    profile.isActive,
        },
        "periodTotals": {
            "availableWithdraw":  float(aw),
            "notCleared":         float(_d(latest.notCleared)       if latest else _ZERO),
            "activeOrders":       active_orders,
            "activeOrderAmount":  float(_d(latest.activeOrderAmount) if latest else _ZERO),
            "withdrawn":          float(_d(latest.withdrawn)         if latest else _ZERO),
            "sellerPlus":         latest.sellerPlus                  if latest else False,
            "promotion":          float(_d(latest.promotion)         if latest else _ZERO),
            "revenueInPeriod":    float(order_revenue),
            "orderAmountInPeriod": float(order_amount),
            "snapshotCount":      snap_total,
            "orderCount":         order_total,
        },
        "pagination": _pagination_meta(pagination, max(snap_total, order_total)),
        "snapshots":  [_entry_to_dict(e, profile.profileName) for e in live_entries],
        "orders":     [_order_to_dict(o) for o in live_orders],
    }


# ── Excel export ──────────────────────────────────────────────────────────────

async def export_profile_excel(
    db: Prisma,
    profile_id: str,
    filters: DateRangeFilter,
) -> tuple[bytes, str]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot generate Excel export.",
        )

    detail   = await get_profile_detail(db, profile_id, filters)
    pname    = detail["profile"]["profileName"]
    start, end = filters.window()

    wb          = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")

    def _header(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def _autofit(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    ws1 = wb.active
    ws1.title = "Snapshots"
    _header(ws1, [
        "Date", "Available Withdraw ($)", "Not Cleared ($)",
        "Active Orders", "Active Order Amount ($)", "Withdrawn ($)",
        "Seller Plus", "Promotion ($)",
    ])
    for ri, s in enumerate(detail["snapshots"], 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([
            str(s["date"]),
            s["availableWithdraw"], s["notCleared"],
            s["activeOrders"], s["activeOrderAmount"],
            s["withdrawn"], s["sellerPlus"], s["promotion"],
        ], 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            if fill:
                cell.fill = fill
    _autofit(ws1)

    ws2 = wb.create_sheet("Orders")
    _header(ws2, ["Date", "Buyer", "Order ID", "Amount ($)", "After Fiverr ($)"])
    for ri, o in enumerate(detail["orders"], 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([
            str(o["date"]), o["buyerName"], o["orderId"],
            o["amount"], o["afterFiverr"],
        ], 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            if fill:
                cell.fill = fill
    _autofit(ws2)

    buf      = io.BytesIO()
    wb.save(buf)
    tag      = f"{start}_{end}" if start else "all"
    filename = f"fiverr_{pname.replace(' ', '_')}_{tag}.xlsx"
    return buf.getvalue(), filename