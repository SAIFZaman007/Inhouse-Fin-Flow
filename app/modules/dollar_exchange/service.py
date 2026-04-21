"""
app/modules/dollar_exchange/service.py
========================================
v3 — date serialization fix

DateTime @db.Date fields in prisma-py v0.14.0 MUST be passed as
datetime.datetime objects (midnight), using the same pattern as the
working Fiverr module:

    datetime.combine(data.date, time.min)

SCHEMA FACTS (schema.prisma — ground truth):
  Field name   : paymentStatus            ← camelCase, matches schema exactly
  Enum type    : PaymentStatus            ← defined in schema.prisma
  Enum values  : RECEIVED | DUE           ← use these exact strings everywhere

PRISMA-CLIENT-PY RULES:
  create/update data dict  → "paymentStatus": "RECEIVED"
  where/filter dict        → "paymentStatus": "RECEIVED"
  Python attribute access  → r.paymentStatus

NOTE ON .aggregate():
  prisma-client-py does NOT expose a model-level .aggregate() method.
  All aggregation is done via db.query_raw() with raw PostgreSQL SQL.
"""
import io
from datetime import date as dt_date, datetime, time
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import HTTPException
from prisma import Prisma

from .schema import DollarExchangeCreate, DollarExchangeUpdate


# ─────────────────────────────────────────────────────────────────────────────
# Private helper  (mirrors the Fiverr service convention exactly)
# ─────────────────────────────────────────────────────────────────────────────

def _dt(d: dt_date) -> datetime:
    """
    Convert datetime.date → datetime.datetime at midnight.

    prisma-py v0.14.0 requires a full datetime object for every
    DateTime @db.Date field — the same pattern used across the Fiverr
    module: datetime.combine(d, time.min).
    """
    return datetime.combine(d, time.min)



# ─────────────────────────────────────────────────────────────────────────────
# § TS  Timestamp bootstrap
#
# DollarExchange has createdAt natively in schema.prisma.
# updatedAt is absent — added via idempotent ALTER TABLE.
# One execute_raw per statement: PostgreSQL extended-query protocol forbids
# multiple commands per prepared statement.
# ─────────────────────────────────────────────────────────────────────────────

_DE_TS_DDL: list[str] = [
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='dollar_exchanges' AND column_name='updated_at'
      ) THEN
        ALTER TABLE dollar_exchanges ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
]

_de_ts_done = False


async def _ensure_de_timestamps(db: Prisma) -> None:
    """
    Idempotent bootstrap — runs once per process lifetime.
    Adds updated_at to dollar_exchanges where Prisma didn't generate it.
    """
    global _de_ts_done
    if _de_ts_done:
        return
    for stmt in _DE_TS_DDL:
        await db.execute_raw(stmt)
    _de_ts_done = True


async def _fetch_updated_at(db: Prisma, exchange_id: str):
    """Read raw updated_at from dollar_exchanges. Returns None on any error."""
    try:
        rows = await db.query_raw(
            "SELECT updated_at FROM dollar_exchanges WHERE id = $1",
            exchange_id,
        )
        if rows:
            return rows[0].get("updated_at")
    except Exception:
        pass
    return None


async def _fetch_updated_at_batch(db: Prisma, ids: list) -> dict:
    """
    Batch-fetch updated_at for a list of IDs.
    Returns {id: updated_at}.  Single query — no N+1.
    """
    if not ids:
        return {}
    try:
        placeholders = ", ".join(f"${i+1}" for i in range(len(ids)))
        rows = await db.query_raw(
            f"SELECT id, updated_at FROM dollar_exchanges WHERE id IN ({placeholders})",
            *ids,
        )
        return {r["id"]: r.get("updated_at") for r in rows}
    except Exception:
        return {}


async def _touch_updated_at(db: Prisma, exchange_id: str) -> None:
    """Bump updated_at on a dollar_exchanges row after any write."""
    try:
        await db.execute_raw(
            "UPDATE dollar_exchanges SET updated_at = now() WHERE id = $1",
            exchange_id,
        )
    except Exception:
        pass


# ── Serialiser ────────────────────────────────────────────────────────────────

def _serialize(record, updated_at=None) -> dict:
    """
    ORM DollarExchange → serialisable dict.

    ``createdAt`` — read directly from the ORM object (native Prisma column).
    ``updatedAt``  — pass the value fetched from the raw updated_at column.
                     Defaults to None when called without a DB handle (e.g. in
                     the Excel export path where we don't need timestamps).
    """
    return {
        "id":            record.id,
        "date":          record.date.date() if isinstance(record.date, datetime) else record.date,
        "details":       record.details,
        "accountFrom":   record.accountFrom,
        "accountTo":     record.accountTo,
        "debit":         record.debit,
        "credit":        record.credit,
        "rate":          record.rate,
        "totalBdt":      record.totalBdt,
        "paymentStatus": record.paymentStatus if isinstance(record.paymentStatus, str) else record.paymentStatus.value,
        "createdAt":     record.createdAt,
        "updatedAt":     updated_at,
    }

# ─────────────────────────────────────────────────────────────────────────────
# CRUD
# ─────────────────────────────────────────────────────────────────────────────

async def create_exchange(db: Prisma, data: DollarExchangeCreate):
    await _ensure_de_timestamps(db)
    record = await db.dollarexchange.create(
        data={
            "date":          _dt(data.date),   # ← datetime.combine(date, time.min)
            "details":       data.details,
            "accountFrom":   data.accountFrom,
            "accountTo":     data.accountTo,
            "debit":         float(data.debit),
            "credit":        float(data.credit),
            "rate":          float(data.rate),
            "totalBdt":      float(data.total_bdt),
            "paymentStatus": data.payment_status.value,
        }
    )
    await _touch_updated_at(db, record.id)
    upd_at = await _fetch_updated_at(db, record.id)
    return _serialize(record, updated_at=upd_at)


async def list_exchanges(
    db:             Prisma,
    date_filter:    dict,
    payment_status: Optional[str] = None,
    account_from:   Optional[str] = None,
    account_name:   Optional[str] = None,
    search:         Optional[str] = None,
):
    """
    List dollar exchange records with optional filters.

    Filters (all combinable):
      date_filter    — Prisma-compatible dict from DateRangeFilter.to_prisma_filter()
      payment_status — "RECEIVED" | "DUE" (also accepts legacy "RCV" alias)
      account_from   — case-insensitive substring search on accountFrom (legacy param)
      account_name   — case-insensitive OR search on accountFrom AND accountTo
                       simultaneously (matches either side of the exchange)
      search         — case-insensitive keyword search on the details column
    """
    await _ensure_de_timestamps(db)
    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if payment_status:
        status = (
            "RECEIVED"
            if payment_status.upper() in ("RECEIVED", "RCV")
            else payment_status.upper()
        )
        where["paymentStatus"] = status

    # Legacy param — exact accountFrom search (kept for backward compatibility)
    if account_from:
        where["accountFrom"] = {"contains": account_from, "mode": "insensitive"}

    # New param — OR search across both accountFrom and accountTo
    if account_name:
        where["OR"] = [
            {"accountFrom": {"contains": account_name, "mode": "insensitive"}},
            {"accountTo":   {"contains": account_name, "mode": "insensitive"}},
        ]

    # Details keyword search
    if search:
        where["details"] = {"contains": search, "mode": "insensitive"}

    records = await db.dollarexchange.find_many(where=where, order={"date": "desc"})

    # Batch-fetch updated_at for all returned rows — single query, no N+1
    ids    = [r.id for r in records]
    upd_map = await _fetch_updated_at_batch(db, ids)

    return [_serialize(r, updated_at=upd_map.get(r.id)) for r in records]


async def get_exchange(db: Prisma, exchange_id: str):
    await _ensure_de_timestamps(db)
    record = await db.dollarexchange.find_unique(where={"id": exchange_id})
    if not record:
        raise HTTPException(status_code=404, detail="Exchange record not found")
    upd_at = await _fetch_updated_at(db, record.id)
    return _serialize(record, updated_at=upd_at)


async def update_exchange(db: Prisma, exchange_id: str, data: DollarExchangeUpdate):
    await _ensure_de_timestamps(db)
    # Fetch raw ORM object for field access during patch computation
    existing_orm = await db.dollarexchange.find_unique(where={"id": exchange_id})
    if not existing_orm:
        raise HTTPException(status_code=404, detail="Exchange record not found")

    update_data: dict = {}

    if data.payment_status is not None:
        update_data["paymentStatus"] = data.payment_status.value

    if data.details is not None:
        update_data["details"] = data.details

    if data.accountFrom is not None:
        update_data["accountFrom"] = data.accountFrom

    if data.accountTo is not None:
        update_data["accountTo"] = data.accountTo

    if data.rate is not None:
        new_rate = float(data.rate)
        exchange_amount = float(
            existing_orm.credit
            if (existing_orm.credit and existing_orm.credit > 0)
            else existing_orm.debit or Decimal("0")
        )
        update_data["rate"]     = new_rate
        update_data["totalBdt"] = exchange_amount * new_rate

    # Idempotent — nothing changed
    if not update_data:
        upd_at = await _fetch_updated_at(db, existing_orm.id)
        return _serialize(existing_orm, updated_at=upd_at)

    updated = await db.dollarexchange.update(
        where={"id": exchange_id},
        data=update_data,
    )
    await _touch_updated_at(db, updated.id)
    upd_at = await _fetch_updated_at(db, updated.id)
    return _serialize(updated, updated_at=upd_at)


async def delete_exchange(db: Prisma, exchange_id: str) -> None:
    record = await db.dollarexchange.find_unique(where={"id": exchange_id})
    if not record:
        raise HTTPException(status_code=404, detail="Exchange record not found")
    await db.dollarexchange.delete(where={"id": exchange_id})


async def get_total_bdt(
    db:             Prisma,
    date_filter:    Optional[dict] = None,
    payment_status: Optional[str]  = None,
    account_name:   Optional[str]  = None,
    search:         Optional[str]  = None,
) -> dict:
    """
    Return total BDT split by payment status.

    Supports all the same filters as list_exchanges so the totals widget
    always reflects the currently active filter state in the UI.

    Uses query_raw — prisma-client-py does not expose .aggregate().
    """
    await _ensure_de_timestamps(db)

    # Build a matching Prisma WHERE clause — reuse list_exchanges logic to
    # get the filtered set, then aggregate from Python (simpler than raw SQL
    # with dynamic filters).
    records = await list_exchanges(
        db,
        date_filter=date_filter or {},
        payment_status=payment_status,
        account_name=account_name,
        search=search,
    )

    total    = sum(float(r["totalBdt"]) for r in records)
    received = sum(float(r["totalBdt"]) for r in records if r.get("paymentStatus") == "RECEIVED")
    due      = sum(float(r["totalBdt"]) for r in records if r.get("paymentStatus") == "DUE")

    return {
        "total":    round(total,    2),
        "received": round(received, 2),
        "due":      round(due,      2),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HEADERS = [
    "Date", "Details", "Account From", "Account To",
    "Debit (USD)", "Credit (USD)", "Rate", "Total BDT", "Payment Status",
]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
_DUE_FILL     = PatternFill("solid", fgColor="FCE4D6")   # light red for DUE rows
_CENTER       = Alignment(horizontal="center", vertical="center")
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_COL_WIDTHS = [12, 36, 22, 22, 14, 14, 10, 18, 16]


def _fmt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):          # ORM may return DateTime @db.Date as datetime
        return v.date().isoformat()
    if isinstance(v, dt_date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


async def export_exchanges(
    db:             Prisma,
    date_filter:    dict,
    payment_status: Optional[str] = None,
    account_from:   Optional[str] = None,
    account_name:   Optional[str] = None,
    search:         Optional[str] = None,
    label:          str = "dollar_exchange",
) -> tuple[bytes, str]:
    """
    Build and return (xlsx_bytes, filename) for the filtered exchange records.
    DUE rows are highlighted in light red for quick identification.
    Supports all the same filters as list_exchanges.
    """
    rows = await list_exchanges(
        db, date_filter,
        payment_status=payment_status,
        account_from=account_from,
        account_name=account_name,
        search=search,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dollar Exchange"

    # ── Header ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, exc in enumerate(rows, start=2):
        # rows are now dicts from _serialize() — access via []
        is_due   = str(exc.get("paymentStatus", "")).upper() == "DUE"
        row_fill = _DUE_FILL if is_due else (_ALT_ROW_FILL if row_idx % 2 == 0 else None)

        values = [
            _fmt(exc.get("date")),
            _fmt(exc.get("details")),
            _fmt(exc.get("accountFrom")),
            _fmt(exc.get("accountTo")),
            float(exc.get("debit",    0) or 0),
            float(exc.get("credit",   0) or 0),
            float(exc.get("rate",     0) or 0),
            float(exc.get("totalBdt", 0) or 0),
            str(exc.get("paymentStatus", "")),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER if col_idx in (1, 9) else _LEFT
            if row_fill:
                cell.fill = row_fill

    # ── Summary ───────────────────────────────────────────────────────────────
    if rows:
        s = len(rows) + 2
        ws.cell(row=s, column=7,  value="TOTAL").font = Font(bold=True)
        ws.cell(row=s, column=5,  value=sum(float(r.get("debit",    0) or 0) for r in rows)).font = Font(bold=True)
        ws.cell(row=s, column=6,  value=sum(float(r.get("credit",   0) or 0) for r in rows)).font = Font(bold=True)
        ws.cell(row=s, column=8,  value=sum(float(r.get("totalBdt", 0) or 0) for r in rows)).font = Font(bold=True)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{label}.xlsx"
    return buffer.read(), filename