"""
app/modules/card_sharing/service.py
================================================================================
v8 — Multi-column OR keyword search
================================================================================
"""
import io
import uuid
from datetime import date as dt_date, datetime, time
from decimal import Decimal
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, UploadFile
from prisma import Json, Prisma

from app.core.cloudinary_service import upload_card_screenshot
from app.core.security import decrypt_value, encrypt_value

from .schema import CardSharingCreate, CardSharingUpdate


# ── Constants / defaults ──────────────────────────────────────────────────────

_ZERO        = Decimal("0")
_EMPTY_STR   = ""
_EMPTY_JSON  = Json([])


# ── Date helper ───────────────────────────────────────────────────────────────

def _to_datetime(d: dt_date) -> datetime:
    """Convert a bare ``date`` to midnight ``datetime``. Prisma rejects bare dates."""
    if isinstance(d, datetime):
        return d
    return datetime.combine(d, time.min)


def _now_datetime() -> datetime:
    """Return the current server timestamp as a ``datetime`` object."""
    return datetime.now()


# ── Account resolver ──────────────────────────────────────────────────────────

async def _resolve_or_create_account(db: Prisma, card_name: str) -> str:
    """
    Return the ``PayoneerAccount.id`` whose ``accountName`` matches
    ``card_name`` (case-insensitive exact match).

    If no such account exists, one is created on-the-fly so that
    ``card_name`` behaves like a standalone free-form label from the
    caller's perspective — no manual account setup is required.

    Returns the account ``id`` (string) to be stored as ``accountId`` on
    the card row.
    """
    account = await db.payoneeraccount.find_first(
        where={"accountName": {"equals": card_name, "mode": "insensitive"}}
    )
    if account:
        return account.id

    # Auto-create a PayoneerAccount to back the free-form label.
    new_account = await db.payoneeraccount.create(
        data={"accountName": card_name}
    )
    return new_account.id


# ── Serialiser ────────────────────────────────────────────────────────────────

def _serialize_card(card, include_sensitive: bool = False) -> dict:
    """
    Map a Prisma ``CardSharing`` model instance to a response dict.

    REQUIRES the card to have been fetched with ``include={"account": True}``
    so that ``card.account.accountName`` is available.

    ``cardName`` in the response = ``card.account.accountName``.
    This is completely transparent to the router and schema — they always
    work with the clean ``cardName`` key.

    ``cardNo`` and ``cardCvc`` are ALWAYS decrypted before returning.
    ``include_sensitive`` is accepted for backward-compatible call-sites
    but intentionally ignored — encryption is purely at-rest.
    """
    card_details = card.cardDetails
    if not isinstance(card_details, list):
        card_details = []

    # Read the card's display name through the account relation.
    card_name = card.account.accountName if card.account else ""

    return {
        "id":                  card.id,
        "serialNo":            card.serialNo,
        "date":                card.date,
        "details":             card.details,
        "cardName":            card_name,
        "cardNo":              decrypt_value(card.cardNo),
        "cardExpire":          card.cardExpire,
        "cardCvc":             decrypt_value(card.cardCvc),
        "cardDetails":         card_details,
        "cardVendor":          card.cardVendor,
        "cardLimit":           card.cardLimit,
        "cardPaymentReceive":  card.cardPaymentReceive,
        "cardReceiveBank":     card.cardReceiveBank,
        "mailDetails":         card.mailDetails,
        "createdAt":           card.createdAt,
        "updatedAt":           card.updatedAt,
    }


# ── Where-clause builder ──────────────────────────────────────────────────────

def _build_where(
    date_filter:  Optional[dict],
    serial_no:    Optional[str],
    account_name: Optional[str],
    search:       Optional[str] = None,
) -> dict:
    """
    Build the Prisma ``where`` clause from all active filter parameters.

    ``serial_no``    — case-insensitive substring match on ``serialNo``.
    ``account_name`` — case-insensitive substring match on card label
                       (``PayoneerAccount.accountName`` via relation).
    ``search``       — case-insensitive OR match across FIVE columns:
                         • details          (direct column, String?)
                         • cardName         (via account.accountName relation)
                         • cardReceiveBank  (direct column, String)
                         • mailDetails      (direct column, String?)
                         • cardVendor       (direct column, String)
                       A single keyword matches any of these columns.
                       This is implemented as a single Prisma OR list —
                       one DB query, no post-processing.
    """
    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if serial_no:
        where["serialNo"] = {"contains": serial_no, "mode": "insensitive"}

    if account_name:
        # Filter via the relation — CardSharing has no direct name column
        where["account"] = {
            "accountName": {"contains": account_name, "mode": "insensitive"}
        }

    if search:
        # OR search across all 5 searchable text columns in a single query.
        # ``cardName`` is resolved through the PayoneerAccount relation.
        # All four direct columns (details, cardReceiveBank, mailDetails,
        # cardVendor) use standard Prisma contains filters.
        where["OR"] = [
            {"details":         {"contains": search, "mode": "insensitive"}},
            {"cardVendor":      {"contains": search, "mode": "insensitive"}},
            {"cardReceiveBank": {"contains": search, "mode": "insensitive"}},
            {"mailDetails":     {"contains": search, "mode": "insensitive"}},
            {"account": {"accountName": {"contains": search, "mode": "insensitive"}}},
        ]

    return where


# ── CRUD ──────────────────────────────────────────────────────────────────────

async def create_card(db: Prisma, data: CardSharingCreate) -> dict:
    """
    Create a new card record.

    ``card_name`` is the only required field — all others are optional and
    receive safe server-side defaults when omitted:

    - ``date``        → current server timestamp (``datetime.now()``)
    - ``serial_no``   → auto-generated UUID v4
    - ``card_no``     → ``""`` (encrypted at rest)
    - ``card_expire`` → ``""``
    - ``card_cvc``    → ``""`` (encrypted at rest)
    - ``card_vendor`` → ``""``
    - ``card_limit``  → ``Decimal("0")``
    - ``card_payment_received`` → ``Decimal("0")``
    - ``card_receiver_bank``    → ``""``
    - ``details``     → ``None``
    - ``mail_details``→ ``None``
    - ``card_details``→ ``[]`` (screenshots added via dedicated endpoint)

    Raises HTTP 409 if ``serial_no`` already exists.
    """
    serial_no = data.serial_no or str(uuid.uuid4())

    existing = await db.cardsharing.find_unique(where={"serialNo": serial_no})
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Serial number '{serial_no}' already exists.",
        )

    account_id = await _resolve_or_create_account(db, data.card_name)

    card_no     = data.card_no     or _EMPTY_STR
    card_expire = data.card_expire or _EMPTY_STR
    card_cvc    = data.card_cvc    or _EMPTY_STR
    card_vendor = data.card_vendor or _EMPTY_STR
    card_limit  = data.card_limit  if data.card_limit  is not None else _ZERO
    card_payment_received = (
        data.card_payment_received
        if data.card_payment_received is not None
        else _ZERO
    )
    card_receiver_bank = data.card_receiver_bank or _EMPTY_STR

    card = await db.cardsharing.create(
        data={
            "serialNo":           serial_no,
            "date":               _now_datetime(),
            "details":            data.details,
            "accountId":          account_id,
            "cardNo":             encrypt_value(card_no),
            "cardExpire":         card_expire,
            "cardCvc":            encrypt_value(card_cvc),
            "cardDetails":        _EMPTY_JSON,
            "cardVendor":         card_vendor,
            "cardLimit":          card_limit,
            "cardPaymentReceive": card_payment_received,
            "cardReceiveBank":    card_receiver_bank,
            "mailDetails":        data.mail_details,
        },
        include={"account": True},
    )
    return _serialize_card(card)


async def list_cards(
    db:                Prisma,
    include_sensitive: bool = False,
    serial_no:         Optional[str] = None,
    account_name:      Optional[str] = None,
    date_filter:       Optional[dict] = None,
    search:            Optional[str] = None,
) -> list[dict]:
    """
    Return all card records, with optional filters.

    ``serial_no``    — case-insensitive substring match on serial number.
    ``account_name`` — case-insensitive substring match on card label
                       (``PayoneerAccount.accountName`` via relation).
    ``search``       — case-insensitive OR keyword search across FIVE
                       columns simultaneously:
                         details | cardName | cardReceiveBank |
                         mailDetails | cardVendor
                       A single keyword matches any of these columns.
    ``date_filter``  — Prisma-compatible date range dict.

    All filters are combinable.
    All returned records include decrypted ``cardNo`` and ``cardCvc``.
    All returned records include ``cardName`` sourced from the account relation.
    """
    where = _build_where(date_filter, serial_no, account_name, search)

    cards = await db.cardsharing.find_many(
        where=where,
        order={"date": "desc"},
        include={"account": True},
    )
    return [_serialize_card(c) for c in cards]


async def get_card(
    db: Prisma,
    card_id: str,
    include_sensitive: bool = False,
) -> dict:
    """
    Fetch a single card by ID.

    Raises HTTP 404 if not found.
    Returns decrypted ``cardNo`` and ``cardCvc`` unconditionally.
    """
    card = await db.cardsharing.find_unique(
        where={"id": card_id},
        include={"account": True},
    )
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")
    return _serialize_card(card)


async def update_card(db: Prisma, card_id: str, data: CardSharingUpdate) -> dict:
    """
    Partially update a card record.

    Only fields present in ``data`` (non-None) are written.
    When ``card_name`` is supplied, a matching ``PayoneerAccount`` is
    resolved (or created) and ``accountId`` is updated on the card.
    ``cardNo`` and ``cardCvc`` are re-encrypted whenever they are supplied.
    """
    card = await db.cardsharing.find_unique(
        where={"id": card_id},
        include={"account": True},
    )
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")

    patch: dict = {}

    if data.card_name is not None:
        patch["accountId"] = await _resolve_or_create_account(db, data.card_name)

    if data.serial_no is not None:
        existing = await db.cardsharing.find_unique(where={"serialNo": data.serial_no})
        if existing and existing.id != card_id:
            raise HTTPException(
                status_code=409,
                detail=f"Serial number '{data.serial_no}' already exists.",
            )
        patch["serialNo"] = data.serial_no

    if data.card_no is not None:
        patch["cardNo"] = encrypt_value(data.card_no)

    if data.card_expire is not None:
        patch["cardExpire"] = data.card_expire

    if data.card_cvc is not None:
        patch["cardCvc"] = encrypt_value(data.card_cvc)

    if data.card_vendor is not None:
        patch["cardVendor"] = data.card_vendor

    if data.card_limit is not None:
        patch["cardLimit"] = data.card_limit

    if data.card_payment_received is not None:
        patch["cardPaymentReceive"] = data.card_payment_received

    if data.card_receiver_bank is not None:
        patch["cardReceiveBank"] = data.card_receiver_bank

    if data.details is not None:
        patch["details"] = data.details

    if data.mail_details is not None:
        patch["mailDetails"] = data.mail_details

    if not patch:
        return _serialize_card(card)

    updated = await db.cardsharing.update(
        where={"id": card_id},
        data=patch,
        include={"account": True},
    )
    return _serialize_card(updated)


async def delete_card(db: Prisma, card_id: str) -> None:
    """Hard-delete a card record and all associated data."""
    card = await db.cardsharing.find_unique(where={"id": card_id})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")
    await db.cardsharing.delete(where={"id": card_id})


# ── Screenshot management ─────────────────────────────────────────────────────

async def add_screenshot(db: Prisma, card_id: str, file: UploadFile) -> dict:
    """
    Upload a single screenshot to Cloudinary and append the secure URL to
    the card's ``cardDetails`` list.

    Raises HTTP 404 if the card does not exist.
    """
    card = await db.cardsharing.find_unique(where={"id": card_id})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")

    result     = await upload_card_screenshot(file, card.serialNo)
    secure_url = result["secure_url"]
    public_id  = result["public_id"]

    existing: list = card.cardDetails if isinstance(card.cardDetails, list) else []
    new_list       = existing + [secure_url]

    await db.cardsharing.update(
        where={"id": card_id},
        data={"cardDetails": Json(new_list)},
    )
    return {
        "url":              secure_url,
        "publicId":         public_id,
        "totalScreenshots": len(new_list),
    }


async def add_screenshots_bulk(
    db: Prisma,
    card_id: str,
    files: List[UploadFile],
) -> dict:
    """
    Upload multiple screenshots to Cloudinary in sequence and append all
    secure URLs to the card's ``cardDetails`` list in a single DB write.

    Raises HTTP 404 if the card does not exist.
    """
    card = await db.cardsharing.find_unique(where={"id": card_id})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")

    existing: list       = card.cardDetails if isinstance(card.cardDetails, list) else []
    uploaded_urls: list[str] = []

    for file in files:
        result = await upload_card_screenshot(file, card.serialNo)
        uploaded_urls.append(result["secure_url"])

    new_list = existing + uploaded_urls

    await db.cardsharing.update(
        where={"id": card_id},
        data={"cardDetails": Json(new_list)},
    )
    return {
        "uploadedCount":    len(uploaded_urls),
        "totalScreenshots": len(new_list),
        "urls":             uploaded_urls,
    }


async def remove_screenshot(
    db: Prisma,
    card_id: str,
    screenshot_url: str,
) -> dict:
    """
    Remove a Cloudinary URL from the card's ``cardDetails`` list.

    Raises HTTP 404 if the card or the URL is not found.
    """
    card = await db.cardsharing.find_unique(where={"id": card_id})
    if not card:
        raise HTTPException(status_code=404, detail="Card not found.")

    existing: list = card.cardDetails if isinstance(card.cardDetails, list) else []
    if screenshot_url not in existing:
        raise HTTPException(status_code=404, detail="Screenshot URL not found.")

    new_list = [u for u in existing if u != screenshot_url]

    await db.cardsharing.update(
        where={"id": card_id},
        data={"cardDetails": Json(new_list)},
    )
    return {"totalScreenshots": len(new_list)}


# ── Excel export ──────────────────────────────────────────────────────────────
#
# SECURITY: cardNo and cardCvc are intentionally EXCLUDED from exports.
# Sensitive PAN / CVC data must never appear in downloadable files.
# ─────────────────────────────────────────────────────────────────────────────

_HEADERS: list[str] = [
    "Date", "Serial No", "Card Name", "Card Vendor",
    "Card Expire", "Card Limit", "Payment Received", "Receiver Bank",
    "Screenshots", "Details", "Mail Details",
]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")
_CENTER       = Alignment(horizontal="center", vertical="center")
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_COL_WIDTHS   = [12, 18, 26, 16, 14, 16, 18, 22, 12, 36, 36]


def _fmt(v) -> str:
    """Safely convert any DB value to a plain string for Excel output."""
    if v is None:
        return ""
    if isinstance(v, dt_date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


async def export_cards(
    db:           Prisma,
    date_filter:  Optional[dict] = None,
    serial_no:    Optional[str]  = None,
    account_name: Optional[str]  = None,
    search:       Optional[str]  = None,
    label:        str            = "card_sharing",
) -> tuple[bytes, str]:
    """
    Build and return ``(xlsx_bytes, filename)``.

    Accepts the same ``search`` param as ``list_cards`` — the five-column
    OR search is applied consistently via ``_build_where``.

    cardNo and cardCvc are intentionally excluded from the workbook —
    sensitive PAN / CVC data must never appear in downloadable files.
    """
    where = _build_where(date_filter, serial_no, account_name, search)

    cards_raw = await db.cardsharing.find_many(
        where=where,
        order={"date": "desc"},
        include={"account": True},
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Card Sharing"

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, card in enumerate(cards_raw, start=2):
        fill         = _ALT_ROW_FILL if row_idx % 2 == 0 else None
        details_list = card.cardDetails if isinstance(card.cardDetails, list) else []
        card_name    = card.account.accountName if card.account else ""

        values = [
            _fmt(card.date),
            _fmt(card.serialNo),
            _fmt(card_name),
            _fmt(card.cardVendor),
            _fmt(card.cardExpire),
            float(card.cardLimit),
            float(card.cardPaymentReceive),
            _fmt(card.cardReceiveBank),
            len(details_list),
            _fmt(card.details),
            _fmt(card.mailDetails),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CENTER if col_idx in (1, 8, 9) else _LEFT
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.read(), f"{label}.xlsx"