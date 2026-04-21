"""
app/modules/pmak/service.py
════════════════════════════════════════════════════════════════════════════════
v7.0 — Enterprise Edition
════════════════════════════════════════════════════════════════════════════════
"""
import io
import uuid
from datetime import date as dt_date, datetime, time, timezone
from decimal import Decimal
from typing import List, Optional

import openpyxl
from fastapi import HTTPException
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from prisma import Prisma

from app.shared.filters import DateRangeFilter
from app.shared.pagination import PageParams

from .schema import (
    PmakAccountCreate,
    PmakInhouseCreate,
    PmakInhouseFullUpdate,
    PmakInhouseStatusUpdate,
    PmakToolCreate,
    PmakToolUpdate,
    PmakTransactionCreate,
    PmakTransactionStatusUpdate,
)


# ═════════════════════════════════════════════════════════════════════════════
# § 0  Utilities
# ═════════════════════════════════════════════════════════════════════════════

def _to_dt(d: dt_date) -> datetime:
    """date → datetime midnight.  Prisma-client-py rejects bare date objects."""
    if isinstance(d, datetime):
        return d
    return datetime.combine(d, time.min)


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _empty_inhouse_by_status() -> dict:
    return {
        "PENDING":     {"count": 0, "totalAmount": 0.0},
        "IN_PROGRESS": {"count": 0, "totalAmount": 0.0},
        "COMPLETED":   {"count": 0, "totalAmount": 0.0},
        "CANCELLED":   {"count": 0, "totalAmount": 0.0},
    }


def _build_inhouse_by_status(deals: list) -> dict:
    result = _empty_inhouse_by_status()
    for d in deals:
        key = d.orderStatus if isinstance(d.orderStatus, str) else d.orderStatus.value
        if key in result:
            result[key]["count"] += 1
            result[key]["totalAmount"] += float(d.orderAmount)
    return result


def _serialize_inhouse_with_account(d) -> dict:
    """Serialise a PmakInhouse row fetched with include={"account": True}."""
    return {
        "id":          d.id,
        "accountId":   d.accountId,
        "accountName": d.account.accountName if d.account else "",
        "date":        d.date.date() if hasattr(d.date, "date") else d.date,
        "details":     d.details,
        "buyerName":   d.buyerName,
        "sellerName":  d.sellerName,
        "orderAmount": d.orderAmount,
        "orderStatus": d.orderStatus if isinstance(d.orderStatus, str) else d.orderStatus.value,
        "createdAt":   d.createdAt,
        "updatedAt":   d.updatedAt,
    }


def _serialize_txn(txn, account_name: str) -> dict:
    return {
        "id":               txn.id,
        "accountId":        txn.accountId,
        "accountName":      account_name,
        "date":             txn.date.date() if hasattr(txn.date, "date") else txn.date,
        "details":          txn.details,
        "accountFrom":      txn.accountFrom,
        "accountTo":        txn.accountTo,
        "debit":            txn.debit,
        "credit":           txn.credit,
        "remainingBalance": txn.remainingBalance,
        "status":           txn.status if isinstance(txn.status, str) else txn.status.value,
        "createdAt":        txn.createdAt,
    }


def _serialize_inhouse(deal, account_name: str) -> dict:
    return {
        "id":          deal.id,
        "accountId":   deal.accountId,
        "accountName": account_name,
        "date":        deal.date.date() if hasattr(deal.date, "date") else deal.date,
        "details":     deal.details,
        "buyerName":   deal.buyerName,
        "sellerName":  deal.sellerName,
        "orderAmount": deal.orderAmount,
        "orderStatus": deal.orderStatus if isinstance(deal.orderStatus, str) else deal.orderStatus.value,
        "createdAt":   deal.createdAt,
        "updatedAt":   deal.updatedAt,
    }


# ═════════════════════════════════════════════════════════════════════════════
# § 1  Raw-SQL bootstrap for pmak_tools and pmak_accounts timestamps
# ═════════════════════════════════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

# pmak_accounts timestamp columns — two separate ALTER TABLE statements
_ACCOUNT_TIMESTAMPS_DDL: list[str] = [
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='pmak_accounts' AND column_name='created_at'
      ) THEN
        ALTER TABLE pmak_accounts ADD COLUMN created_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
    """
    DO $$ BEGIN
      IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name='pmak_accounts' AND column_name='updated_at'
      ) THEN
        ALTER TABLE pmak_accounts ADD COLUMN updated_at TIMESTAMPTZ DEFAULT now();
      END IF;
    END $$
    """,
]

# pmak_tools table + two indexes — three separate statements
_TOOLS_TABLE_DDL: list[str] = [
    """
    CREATE TABLE IF NOT EXISTS pmak_tools (
        id           TEXT          PRIMARY KEY,
        account_id   TEXT          NOT NULL REFERENCES pmak_accounts(id) ON DELETE CASCADE,
        date         DATE          NOT NULL,
        details      TEXT,
        debit        NUMERIC(12,2) NOT NULL DEFAULT 0,
        credit       NUMERIC(12,2) NOT NULL DEFAULT 0,
        total        NUMERIC(12,2) NOT NULL DEFAULT 0,
        created_at   TIMESTAMPTZ   NOT NULL DEFAULT now(),
        updated_at   TIMESTAMPTZ   NOT NULL DEFAULT now()
    )
    """,
    "CREATE INDEX IF NOT EXISTS idx_pmak_tools_account_id ON pmak_tools(account_id)",
    "CREATE INDEX IF NOT EXISTS idx_pmak_tools_date        ON pmak_tools(date)",
]

_bootstrap_done = False


async def _ensure_schema(db: Prisma) -> None:
    """
    Idempotent schema bootstrap — runs once per process lifetime.

    Each statement is issued in its own execute_raw call so that PostgreSQL's
    extended-query protocol never sees more than one command per round-trip.
    The module-level flag prevents redundant DDL on subsequent requests.
    """
    global _bootstrap_done
    if _bootstrap_done:
        return
    for stmt in _ACCOUNT_TIMESTAMPS_DDL:
        await db.execute_raw(stmt)
    for stmt in _TOOLS_TABLE_DDL:
        await db.execute_raw(stmt)
    _bootstrap_done = True


# ═════════════════════════════════════════════════════════════════════════════
# § 2  Account timestamps helper
# ═════════════════════════════════════════════════════════════════════════════

async def _fetch_account_timestamps(db: Prisma, account_id: str) -> dict:
    """
    Return {'createdAt': datetime|None, 'updatedAt': datetime|None} for one
    account.  Falls back to None if the columns are not yet present (pre-v7
    deployment guard).
    """
    try:
        rows = await db.query_raw(
            "SELECT created_at, updated_at FROM pmak_accounts WHERE id = $1",
            account_id,
        )
        if rows:
            row = rows[0]
            return {
                "createdAt": row.get("created_at"),
                "updatedAt": row.get("updated_at"),
            }
    except Exception:
        pass
    return {"createdAt": None, "updatedAt": None}


async def _fetch_all_account_timestamps(db: Prisma, account_ids: list) -> dict:
    """
    Batch-fetch timestamps for a list of account IDs.
    Returns {account_id: {'createdAt': ..., 'updatedAt': ...}}.
    """
    if not account_ids:
        return {}
    try:
        placeholders = ", ".join(f"${i+1}" for i in range(len(account_ids)))
        rows = await db.query_raw(
            f"SELECT id, created_at, updated_at FROM pmak_accounts WHERE id IN ({placeholders})",
            *account_ids,
        )
        return {
            row["id"]: {
                "createdAt": row.get("created_at"),
                "updatedAt": row.get("updated_at"),
            }
            for row in rows
        }
    except Exception:
        return {aid: {"createdAt": None, "updatedAt": None} for aid in account_ids}


# ═════════════════════════════════════════════════════════════════════════════
# § 3  Tools raw-query helpers
# ═════════════════════════════════════════════════════════════════════════════

def _serialize_tool_row(row: dict, account_name: str) -> dict:
    """
    Normalise a raw-SQL dict row from pmak_tools into the PmakToolResponse
    shape.  Handles both string and datetime values for date columns.
    """
    raw_date = row.get("date")
    if isinstance(raw_date, datetime):
        tool_date = raw_date.date()
    elif isinstance(raw_date, dt_date):
        tool_date = raw_date
    elif isinstance(raw_date, str):
        tool_date = dt_date.fromisoformat(raw_date[:10])
    else:
        tool_date = dt_date.today()

    return {
        "id":          row["id"],
        "accountId":   row["account_id"],
        "accountName": account_name,
        "date":        tool_date,
        "details":     row.get("details"),
        "debit":       Decimal(str(row["debit"])),
        "credit":      Decimal(str(row["credit"])),
        "total":       Decimal(str(row["total"])),
        "createdAt":   row.get("created_at"),
        "updatedAt":   row.get("updated_at"),
    }


async def _resolve_active_account(db: Prisma, account_name: str):
    """Case-insensitive lookup of an active PmakAccount by name."""
    account = await db.pmakaccount.find_first(
        where={
            "accountName": {"equals": account_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not account:
        raise HTTPException(
            status_code=404,
            detail=f"Active PMAK account '{account_name}' not found",
        )
    return account


# ═════════════════════════════════════════════════════════════════════════════
# § 4  Accounts
# ═════════════════════════════════════════════════════════════════════════════

async def create_account(db: Prisma, data: PmakAccountCreate):
    await _ensure_schema(db)
    existing = await db.pmakaccount.find_first(
        where={"accountName": {"equals": data.accountName, "mode": "insensitive"}}
    )
    if existing:
        raise HTTPException(status_code=409, detail="Account name already exists")

    account = await db.pmakaccount.create(data={"accountName": data.accountName})
    ts = await _fetch_account_timestamps(db, account.id)
    return {
        "id":                account.id,
        "accountName":       account.accountName,
        "isActive":          account.isActive,
        "createdAt":         ts["createdAt"],
        "updatedAt":         ts["updatedAt"],
        "currentBalance":    0.0,
        "totalTransactions": 0,
        "totalInhouse":      0,
        "inhouseByStatus":   _empty_inhouse_by_status(),
    }


async def deactivate_account(db: Prisma, account_id: str):
    await _ensure_schema(db)
    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")
    await db.pmakaccount.update(
        where={"id": account_id},
        data={"isActive": False},
    )
    # Touch updated_at on the raw column
    try:
        await db.execute_raw(
            "UPDATE pmak_accounts SET updated_at = now() WHERE id = $1",
            account_id,
        )
    except Exception:
        pass


async def update_account(db: Prisma, account_id: str, data: PmakAccountCreate):
    """
    PATCH /accounts/{account_id} — rename or toggle isActive on a PMAK account.
    Raises 404 if the account does not exist.
    Raises 409 if the new name collides with an existing account.
    """
    await _ensure_schema(db)
    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")

    if data.accountName and data.accountName != account.accountName:
        conflict = await db.pmakaccount.find_first(
            where={
                "accountName": {"equals": data.accountName, "mode": "insensitive"},
                "id":          {"not": account_id},
            }
        )
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Account name '{data.accountName}' is already taken.",
            )

    updated = await db.pmakaccount.update(
        where={"id": account_id},
        data={"accountName": data.accountName},
    )
    try:
        await db.execute_raw(
            "UPDATE pmak_accounts SET updated_at = now() WHERE id = $1",
            account_id,
        )
    except Exception:
        pass

    ts = await _fetch_account_timestamps(db, account_id)
    return {
        "id":          updated.id,
        "accountName": updated.accountName,
        "isActive":    updated.isActive,
        "createdAt":   ts["createdAt"],
        "updatedAt":   ts["updatedAt"],
    }


async def list_accounts(
    db:         Prisma,
    filters:    DateRangeFilter,
    name:       Optional[str] = None,
    pagination: PageParams = None,
):
    """Combined totals + paginated per-account breakdown."""
    await _ensure_schema(db)
    date_filter = filters.to_prisma_filter()

    where: dict = {"isActive": True}
    if name:
        where["accountName"] = {"contains": name, "mode": "insensitive"}

    total_accounts = await db.pmakaccount.count(where=where)

    skip = pagination.skip if pagination else 0
    take = pagination.take if pagination else 50

    txn_include:     dict = {}
    inhouse_include: dict = {}
    if date_filter:
        txn_include     = {"where": {"date": date_filter}}
        inhouse_include = {"where": {"date": date_filter}}

    accounts = await db.pmakaccount.find_many(
        where=where,
        skip=skip,
        take=take,
        order={"accountName": "asc"},
        include={
            "transactions": txn_include if txn_include else True,
            "inhouseDeals": inhouse_include if inhouse_include else True,
        },
    )

    account_ids = [a.id for a in accounts]

    # Latest stored balance per account (single query, no N+1)
    latest_balance_map: dict = {}
    if account_ids:
        all_latest = await db.pmaktransaction.find_many(
            where={"accountId": {"in": account_ids}},
            order={"createdAt": "desc"},
        )
        for txn in all_latest:
            if txn.accountId not in latest_balance_map:
                latest_balance_map[txn.accountId] = float(txn.remainingBalance)

    # All-time inhouse order totals per account (single query, no N+1)
    all_time_inhouse_map: dict = {}
    if account_ids:
        all_inhouse = await db.pmakinhouse.find_many(
            where={"accountId": {"in": account_ids}},
        )
        for deal in all_inhouse:
            all_time_inhouse_map[deal.accountId] = (
                all_time_inhouse_map.get(deal.accountId, 0.0) + float(deal.orderAmount)
            )

    # Batch-fetch account timestamps (v7)
    ts_map = await _fetch_all_account_timestamps(db, account_ids)

    combined_balance      = 0.0
    combined_credit       = 0.0
    combined_debit        = 0.0
    combined_transactions = 0
    combined_inhouse      = 0
    combined_inhouse_amt  = 0.0
    combined_by_status    = _empty_inhouse_by_status()
    account_summaries     = []

    for acct in accounts:
        txns  = sorted(acct.transactions or [], key=lambda t: t.createdAt, reverse=True)
        deals = sorted(acct.inhouseDeals or [], key=lambda d: d.createdAt, reverse=True)

        latest_stored_balance = latest_balance_map.get(acct.id, 0.0)
        period_credit         = sum(float(t.credit) for t in txns)
        period_debit          = sum(float(t.debit)  for t in txns)
        current_balance       = round(latest_stored_balance, 2)

        inhouse_by_status        = _build_inhouse_by_status(deals)
        inhouse_total_amt        = sum(float(d.orderAmount) for d in deals)
        total_inhouse_order_amt  = round(all_time_inhouse_map.get(acct.id, 0.0), 2)

        combined_balance      += current_balance
        combined_credit       += period_credit
        combined_debit        += period_debit
        combined_transactions += len(txns)
        combined_inhouse      += len(deals)
        combined_inhouse_amt  += inhouse_total_amt

        for status_key, v in inhouse_by_status.items():
            combined_by_status[status_key]["count"]       += v["count"]
            combined_by_status[status_key]["totalAmount"] += v["totalAmount"]

        acct_ts = ts_map.get(acct.id, {"createdAt": None, "updatedAt": None})

        account_summaries.append({
            "id":                      acct.id,
            "accountName":             acct.accountName,
            "isActive":                acct.isActive,
            "createdAt":               acct_ts["createdAt"],
            "updatedAt":               acct_ts["updatedAt"],
            "currentBalance":          current_balance,
            "periodCredit":            period_credit,
            "periodDebit":             period_debit,
            "transactionCount":        len(txns),
            "inhouseCount":            len(deals),
            "totalInhouseOrderAmount": total_inhouse_order_amt,
            "inhouseByStatus":         inhouse_by_status,
            "recentTransactions": [_serialize_txn(t, acct.accountName) for t in txns[:5]],
            "recentInhouse":      [_serialize_inhouse(d, acct.accountName) for d in deals[:5]],
        })

    total_pages = max(1, -(-total_accounts // take))

    return {
        "filter": filters.meta(),
        "totals": {
            "totalBalance":       round(combined_balance, 2),
            "totalCredit":        combined_credit,
            "totalDebit":         combined_debit,
            "totalTransactions":  combined_transactions,
            "totalInhouse":       combined_inhouse,
            "totalInhouseAmount": combined_inhouse_amt,
            "inhouseByStatus":    combined_by_status,
            "activeAccountCount": total_accounts,
        },
        "pagination": {
            "page":       pagination.page if pagination else 1,
            "pageSize":   take,
            "total":      total_accounts,
            "totalPages": total_pages,
        },
        "accounts": account_summaries,
    }


# ═════════════════════════════════════════════════════════════════════════════
# § 5  Ledger Transactions
# ═════════════════════════════════════════════════════════════════════════════

async def add_transaction(db: Prisma, data: PmakTransactionCreate):
    """
    POST /transactions — Add a ledger entry.

    remainingBalance is auto-computed server-side:
      latest_balance − debit + credit

    The caller may pass remaining_balance explicitly to override (non-zero, non-None).
    """
    await _ensure_schema(db)
    account = await db.pmakaccount.find_first(
        where={
            "accountName": {"equals": data.account_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not account:
        raise HTTPException(
            status_code=404,
            detail=f"Active PMAK account '{data.account_name}' not found",
        )

    latest_txn = await db.pmaktransaction.find_first(
        where={"accountId": account.id},
        order={"createdAt": "desc"},
    )
    latest_balance = float(latest_txn.remainingBalance) if latest_txn else 0.0

    caller_override = (
        data.remaining_balance is not None
        and data.remaining_balance != Decimal("0")
    )

    if caller_override:
        computed_balance = Decimal(str(data.remaining_balance))
    else:
        computed_balance = Decimal(str(round(
            latest_balance - float(data.debit) + float(data.credit), 2
        )))

    txn = await db.pmaktransaction.create(
        data={
            "date":             _to_dt(data.date),
            "details":          data.details,
            "accountFrom":      data.accountFrom,
            "accountTo":        data.accountTo,
            "debit":            data.debit,
            "credit":           data.credit,
            "remainingBalance": computed_balance,
            "status":           data.status.value,
            "account":          {"connect": {"id": account.id}},
        },
    )

    persisted_txn = await db.pmaktransaction.find_unique(where={"id": txn.id})
    response = _serialize_txn(persisted_txn, account.accountName)
    response["remainingBalance"] = computed_balance
    return response


async def get_account_transactions(
    db:          Prisma,
    account_id:  str,
    date_filter: dict,
    pagination:  PageParams = None,
    search:      Optional[str] = None,
):
    """
    GET /accounts/{id}/transactions

    v7: Added ``search`` — case-insensitive substring match on details OR
    a case-insensitive match on accountName (the account name is fixed per
    endpoint so the useful search here is details-keyword).
    """
    await _ensure_schema(db)
    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")

    where: dict = {"accountId": account_id}
    if date_filter:
        where["date"] = date_filter
    if search:
        where["details"] = {"contains": search, "mode": "insensitive"}

    skip = pagination.skip if pagination else 0
    take = pagination.take if pagination else 50

    total = await db.pmaktransaction.count(where=where)
    txns  = await db.pmaktransaction.find_many(
        where=where,
        order={"date": "desc"},
        skip=skip,
        take=take,
    )

    latest = await db.pmaktransaction.find_first(
        where={"accountId": account_id},
        order={"createdAt": "desc"},
    )
    latest_stored   = float(latest.remainingBalance) if latest else 0.0
    period_credit   = sum(float(t.credit) for t in txns)
    period_debit    = sum(float(t.debit)  for t in txns)
    current_balance = round(latest_stored, 2)

    ts = await _fetch_account_timestamps(db, account_id)

    return {
        "account": {
            "id":          account.id,
            "accountName": account.accountName,
            "isActive":    account.isActive,
            "createdAt":   ts["createdAt"],
            "updatedAt":   ts["updatedAt"],
        },
        "currentBalance": current_balance,
        "periodCredit":   period_credit,
        "periodDebit":    period_debit,
        "pagination": {
            "page":       pagination.page if pagination else 1,
            "pageSize":   take,
            "total":      total,
            "totalPages": max(1, -(-total // take)),
        },
        "transactions": [_serialize_txn(t, account.accountName) for t in txns],
    }


async def update_transaction_status(
    db:             Prisma,
    transaction_id: str,
    data:           PmakTransactionStatusUpdate,
):
    txn = await db.pmaktransaction.find_unique(where={"id": transaction_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    update_data: dict = {}
    if data.status is not None:
        update_data["status"] = data.status.value
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")

    updated = await db.pmaktransaction.update(
        where={"id": transaction_id},
        data=update_data,
    )
    account = await db.pmakaccount.find_unique(where={"id": updated.accountId})
    return _serialize_txn(updated, account.accountName if account else "")


async def delete_transaction(db: Prisma, transaction_id: str):
    txn = await db.pmaktransaction.find_unique(where={"id": transaction_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    await db.pmaktransaction.delete(where={"id": transaction_id})


async def update_transaction(
    db:             Prisma,
    transaction_id: str,
    data:           PmakTransactionCreate,
):
    """
    PATCH /transactions/{transaction_id} — partial update of any field.

    remainingBalance auto-recomputation when debit/credit changes.
    """
    txn = await db.pmaktransaction.find_unique(where={"id": transaction_id})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")

    patch: dict = {}

    if data.date is not None:
        patch["date"] = _to_dt(data.date)
    if data.details is not None:
        patch["details"] = data.details
    if data.accountFrom is not None:
        patch["accountFrom"] = data.accountFrom
    if data.accountTo is not None:
        patch["accountTo"] = data.accountTo
    if data.debit is not None:
        patch["debit"] = data.debit
    if data.credit is not None:
        patch["credit"] = data.credit

    computed_balance: Optional[Decimal] = None

    caller_override = (
        data.remaining_balance is not None
        and data.remaining_balance != Decimal("0")
    )

    if caller_override:
        computed_balance          = Decimal(str(data.remaining_balance))
        patch["remainingBalance"] = computed_balance

    elif "debit" in patch or "credit" in patch:
        old_remaining  = float(txn.remainingBalance)
        old_debit      = float(txn.debit)
        old_credit     = float(txn.credit)
        balance_before = old_remaining + old_debit - old_credit

        new_debit  = float(patch.get("debit",  txn.debit))
        new_credit = float(patch.get("credit", txn.credit))

        computed_balance          = Decimal(str(round(
            balance_before - new_debit + new_credit, 2
        )))
        patch["remainingBalance"] = computed_balance

    if data.status is not None:
        patch["status"] = data.status.value

    if not patch:
        account = await db.pmakaccount.find_unique(where={"id": txn.accountId})
        return _serialize_txn(txn, account.accountName if account else "")

    updated = await db.pmaktransaction.update(
        where={"id": transaction_id},
        data=patch,
    )
    account = await db.pmakaccount.find_unique(where={"id": updated.accountId})
    response = _serialize_txn(updated, account.accountName if account else "")

    if computed_balance is not None:
        response["remainingBalance"] = computed_balance

    return response


# ═════════════════════════════════════════════════════════════════════════════
# § 6  Inhouse Deals
# ═════════════════════════════════════════════════════════════════════════════

async def create_inhouse_deal(db: Prisma, data: PmakInhouseCreate):
    await _ensure_schema(db)
    account = await db.pmakaccount.find_first(
        where={
            "accountName": {"equals": data.account_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not account:
        raise HTTPException(
            status_code=404,
            detail=f"Active PMAK account '{data.account_name}' not found",
        )

    deal = await db.pmakinhouse.create(
        data={
            "date":        _to_dt(data.date),
            "details":     data.details,
            "buyerName":   data.buyer_name,
            "sellerName":  data.seller_name,
            "orderAmount": data.order_amount,
            "orderStatus": data.order_status.value,
            "account":     {"connect": {"id": account.id}},
        },
    )
    return _serialize_inhouse(deal, account.accountName)


async def get_account_inhouse_deals(
    db:          Prisma,
    account_id:  str,
    date_filter: dict,
    pagination:  PageParams = None,
):
    await _ensure_schema(db)
    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")

    where: dict = {"accountId": account_id}
    if date_filter:
        where["date"] = date_filter

    skip = pagination.skip if pagination else 0
    take = pagination.take if pagination else 50

    total = await db.pmakinhouse.count(where=where)
    deals = await db.pmakinhouse.find_many(
        where=where,
        order={"date": "desc"},
        skip=skip,
        take=take,
    )

    inhouse_by_status = _build_inhouse_by_status(deals)
    total_amount      = sum(float(d.orderAmount) for d in deals)

    ts = await _fetch_account_timestamps(db, account_id)

    return {
        "account": {
            "id":          account.id,
            "accountName": account.accountName,
            "isActive":    account.isActive,
            "createdAt":   ts["createdAt"],
            "updatedAt":   ts["updatedAt"],
        },
        "inhouseByStatus": inhouse_by_status,
        "totalAmount":     total_amount,
        "pagination": {
            "page":       pagination.page if pagination else 1,
            "pageSize":   take,
            "total":      total,
            "totalPages": max(1, -(-total // take)),
        },
        "deals": [_serialize_inhouse(d, account.accountName) for d in deals],
    }


async def list_all_inhouse_deals(
    db:           Prisma,
    filters:      DateRangeFilter,
    pagination:   PageParams,
    account_name: Optional[str] = None,
    buyer_name:   Optional[str] = None,
    seller_name:  Optional[str] = None,
    order_status: Optional[str] = None,
    search:       Optional[str] = None,
):
    """
    Cross-account flat deal list with combined totals.

    Filters (all combinable):
      account_name  — case-insensitive substring on PmakAccount.accountName
      buyer_name    — case-insensitive substring on PmakInhouse.buyerName
      seller_name   — case-insensitive substring on PmakInhouse.sellerName
      order_status  — exact enum: PENDING | IN_PROGRESS | COMPLETED | CANCELLED
      search        — case-insensitive keyword search on PmakInhouse.details
      period / from / to / year / month — standard DateRangeFilter
    """
    await _ensure_schema(db)
    date_filter = filters.to_prisma_filter()

    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if buyer_name:
        where["buyerName"] = {"contains": buyer_name, "mode": "insensitive"}

    if seller_name:
        where["sellerName"] = {"contains": seller_name, "mode": "insensitive"}

    if search:
        where["details"] = {"contains": search, "mode": "insensitive"}

    if order_status:
        valid     = {"PENDING", "IN_PROGRESS", "COMPLETED", "CANCELLED"}
        normalised = order_status.upper()
        if normalised not in valid:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid order_status '{order_status}'. "
                       f"Valid values: {', '.join(sorted(valid))}",
            )
        where["orderStatus"] = normalised

    account_filter: dict = {"isActive": True}
    if account_name:
        account_filter["accountName"] = {"contains": account_name, "mode": "insensitive"}
    where["account"] = {"is": account_filter}

    total = await db.pmakinhouse.count(where=where)

    deals = await db.pmakinhouse.find_many(
        where=where,
        order={"date": "desc"},
        skip=pagination.skip,
        take=pagination.take,
        include={"account": True},
    )

    all_for_totals = await db.pmakinhouse.find_many(where=where)

    total_amount       = sum(float(d.orderAmount) for d in all_for_totals)
    combined_by_status = _empty_inhouse_by_status()
    for d in all_for_totals:
        key = d.orderStatus if isinstance(d.orderStatus, str) else d.orderStatus.value
        if key in combined_by_status:
            combined_by_status[key]["count"]       += 1
            combined_by_status[key]["totalAmount"] += float(d.orderAmount)

    return {
        "filter": filters.meta(),
        "totals": {
            "totalDeals":  total,
            "totalAmount": round(total_amount, 2),
            "byStatus":    combined_by_status,
        },
        "pagination": {
            "page":       pagination.page,
            "pageSize":   pagination.take,
            "total":      total,
            "totalPages": max(1, -(-total // pagination.take)),
        },
        "deals": [_serialize_inhouse_with_account(d) for d in deals],
    }


async def update_inhouse_deal(
    db:      Prisma,
    deal_id: str,
    data:    PmakInhouseFullUpdate,
):
    """
    PATCH /inhouse/{deal_id} — Full optional-field update.
    If account_name is supplied the deal is re-linked to the new account.
    Sending an empty body {} is idempotent.
    """
    deal = await db.pmakinhouse.find_unique(
        where={"id": deal_id},
        include={"account": True},
    )
    if not deal:
        raise HTTPException(status_code=404, detail="Inhouse deal not found")

    update_data: dict = {}

    new_account = None
    if data.account_name is not None:
        new_account = await db.pmakaccount.find_first(
            where={
                "accountName": {"equals": data.account_name, "mode": "insensitive"},
                "isActive":    True,
            }
        )
        if not new_account:
            raise HTTPException(
                status_code=404,
                detail=f"Active PMAK account '{data.account_name}' not found",
            )
        update_data["account"] = {"connect": {"id": new_account.id}}

    if data.date is not None:
        update_data["date"] = _to_dt(data.date)
    if data.details is not None:
        update_data["details"] = data.details
    if data.buyer_name is not None:
        update_data["buyerName"] = data.buyer_name
    if data.seller_name is not None:
        update_data["sellerName"] = data.seller_name
    if data.order_amount is not None:
        update_data["orderAmount"] = data.order_amount
    if data.order_status is not None:
        update_data["orderStatus"] = data.order_status.value

    if not update_data:
        account_name = deal.account.accountName if deal.account else ""
        return _serialize_inhouse(deal, account_name)

    updated = await db.pmakinhouse.update(
        where={"id": deal_id},
        data=update_data,
    )

    if new_account:
        resolved_account_name = new_account.accountName
    elif deal.account:
        resolved_account_name = deal.account.accountName
    else:
        acct = await db.pmakaccount.find_unique(where={"id": updated.accountId})
        resolved_account_name = acct.accountName if acct else ""

    return _serialize_inhouse(updated, resolved_account_name)


async def delete_inhouse_deal(db: Prisma, deal_id: str):
    deal = await db.pmakinhouse.find_unique(where={"id": deal_id})
    if not deal:
        raise HTTPException(status_code=404, detail="Inhouse deal not found")
    await db.pmakinhouse.delete(where={"id": deal_id})


# ═════════════════════════════════════════════════════════════════════════════
# § 7  Tools CRUD  (pmak_tools — raw SQL, schema.prisma untouched)
# ═════════════════════════════════════════════════════════════════════════════

async def add_tool(db: Prisma, data: PmakToolCreate):
    """
    POST /tools — Add a PMAK Tools ledger entry.

    total = latest_total_for_account − debit + credit
    (same running-balance semantics as Transactions.remainingBalance).
    """
    await _ensure_schema(db)

    account = await _resolve_active_account(db, data.account_name)

    # Resolve latest stored total for this account
    rows = await db.query_raw(
        """
        SELECT total FROM pmak_tools
         WHERE account_id = $1
         ORDER BY created_at DESC
         LIMIT 1
        """,
        account.id,
    )
    latest_total = float(rows[0]["total"]) if rows else 0.0

    debit  = float(data.debit  or 0)
    credit = float(data.credit or 0)

    caller_override = (
        data.total is not None
        and data.total != Decimal("0")
    )
    if caller_override:
        computed_total = float(data.total)
    else:
        computed_total = round(latest_total - debit + credit, 2)

    # Prisma execute_raw encodes params via its JSON builder.
    # DATE columns require a "YYYY-MM-DD" string param with a ::date cast in SQL.
    # TIMESTAMPTZ columns require an ISO-8601 string with a ::timestamptz cast.
    # This is the exact same pattern list_all_tools / get_account_tools already
    # use successfully for their WHERE date filters.
    tool_date_str = (data.date or dt_date.today()).isoformat()   # "YYYY-MM-DD"
    tool_id       = str(uuid.uuid4())
    now_str       = _now_utc().isoformat()                       # ISO-8601 with tz

    await db.execute_raw(
        """
        INSERT INTO pmak_tools
            (id, account_id, date, details, debit, credit, total, created_at, updated_at)
        VALUES ($1, $2, $3::date, $4, $5, $6, $7, $8::timestamptz, $9::timestamptz)
        """,
        tool_id,
        account.id,
        tool_date_str,   # "YYYY-MM-DD" string  →  ::date cast  → PostgreSQL DATE
        data.details,
        debit,
        credit,
        computed_total,
        now_str,         # ISO-8601 string  →  ::timestamptz cast  → PostgreSQL TIMESTAMPTZ
        now_str,
    )

    rows = await db.query_raw(
        "SELECT * FROM pmak_tools WHERE id = $1",
        tool_id,
    )
    return _serialize_tool_row(rows[0], account.accountName)


async def list_all_tools(
    db:           Prisma,
    filters:      DateRangeFilter,
    pagination:   PageParams,
    account_name: Optional[str] = None,
    search:       Optional[str] = None,
):
    """
    GET /tools — Cross-account flat tools list with per-account totals.

    Response shape:
      filter     — period metadata
      accounts   — per-account aggregates: accountId, accountName,
                   totalDebit, totalCredit, latestTotal, entryCount
      pagination — page, pageSize, total, totalPages
      tools      — paginated flat rows (each includes accountName)

    Filters (all combinable):
      account_name — case-insensitive substring on account name
      search       — case-insensitive substring on the details field
      period / from / to / year / month — standard DateRangeFilter
    """
    await _ensure_schema(db)

    # ── Build shared WHERE clause ─────────────────────────────────────────────
    conditions: List[str] = []
    params:     List     = []
    p = 1  # PostgreSQL placeholder counter

    date_meta = filters.to_prisma_filter()
    if date_meta:
        if "gte" in date_meta and "lte" in date_meta:
            conditions.append(f"t.date >= ${p}::date AND t.date <= ${p+1}::date")
            params.extend([date_meta["gte"].isoformat()[:10], date_meta["lte"].isoformat()[:10]])
            p += 2
        elif "gte" in date_meta:
            conditions.append(f"t.date >= ${p}::date")
            params.append(date_meta["gte"].isoformat()[:10])
            p += 1
        elif "lte" in date_meta:
            conditions.append(f"t.date <= ${p}::date")
            params.append(date_meta["lte"].isoformat()[:10])
            p += 1

    if account_name:
        conditions.append(f'LOWER(a."accountName") LIKE ${p}')
        params.append(f"%{account_name.lower()}%")
        p += 1

    if search:
        conditions.append(f"LOWER(t.details) LIKE ${p}")
        params.append(f"%{search.lower()}%")
        p += 1

    where_sql = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    # ── Total row count (for pagination) ──────────────────────────────────────
    count_rows = await db.query_raw(
        f"""
        SELECT COUNT(*) AS cnt
          FROM pmak_tools t
          JOIN pmak_accounts a ON a.id = t.account_id
         {where_sql}
        """,
        *params,
    )
    total = int(count_rows[0]["cnt"]) if count_rows else 0

    # ── Per-account aggregates ─────────────────────────────────────────────────
    # Group by account so the UI can show a "totals per account" breakdown
    # regardless of which page of entries is currently visible.
    #
    # latestTotal: correlated sub-select fetches the `total` column from the
    # chronologically last row for each account (ORDER BY created_at DESC LIMIT 1).
    # MAX(t.total) is WRONG — it returns the highest value ever, not the current
    # running balance (which decreases on debit entries).
    per_account_rows = await db.query_raw(
        f"""
        SELECT
            t.account_id                        AS account_id,
            a."accountName"                     AS account_name,
            COUNT(*)                            AS entry_count,
            COALESCE(SUM(t.debit),  0)          AS total_debit,
            COALESCE(SUM(t.credit), 0)          AS total_credit,
            COALESCE(
                (
                  SELECT t2.total
                    FROM pmak_tools t2
                   WHERE t2.account_id = t.account_id
                   ORDER BY t2.created_at DESC
                   LIMIT 1
                ), 0
            )                                   AS latest_total
          FROM pmak_tools t
          JOIN pmak_accounts a ON a.id = t.account_id
         {where_sql}
         GROUP BY t.account_id, a."accountName"
         ORDER BY a."accountName" ASC
        """,
        *params,
    )
    accounts_summary = [
        {
            "accountId":   r["account_id"],
            "accountName": r["account_name"],
            "entryCount":  int(r["entry_count"]),
            "totalDebit":  float(r["total_debit"]),
            "totalCredit": float(r["total_credit"]),
            "latestTotal": float(r["latest_total"]),
        }
        for r in per_account_rows
    ]

    # ── Paginated flat tool rows ───────────────────────────────────────────────
    skip = pagination.skip
    take = pagination.take
    rows = await db.query_raw(
        f"""
        SELECT t.*, a."accountName" AS account_name
          FROM pmak_tools t
          JOIN pmak_accounts a ON a.id = t.account_id
         {where_sql}
         ORDER BY t.date DESC, t.created_at DESC
         LIMIT ${p} OFFSET ${p+1}
        """,
        *params,
        take,
        skip,
    )

    tools = []
    for row in rows:
        account_name_val = row.pop("account_name", "")
        tools.append(_serialize_tool_row(row, account_name_val))

    return {
        "filter":   filters.meta(),
        "accounts": accounts_summary,
        "pagination": {
            "page":       pagination.page,
            "pageSize":   take,
            "total":      total,
            "totalPages": max(1, -(-total // take)),
        },
        "tools": tools,
    }


async def get_account_tools(
    db:          Prisma,
    account_id:  str,
    date_filter: dict,
    pagination:  PageParams = None,
    search:      Optional[str] = None,
):
    """
    GET /accounts/{id}/tools — Period-aware tools list for one account.

    search — case-insensitive substring on details.
    """
    await _ensure_schema(db)

    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")

    conditions: List[str] = ["account_id = $1"]
    params:     List     = [account_id]
    p = 2

    if date_filter:
        if "gte" in date_filter and "lte" in date_filter:
            conditions.append(f"date >= ${p}::date AND date <= ${p+1}::date")
            params.extend([date_filter["gte"].isoformat()[:10], date_filter["lte"].isoformat()[:10]])
            p += 2
        elif "gte" in date_filter:
            conditions.append(f"date >= ${p}::date")
            params.append(date_filter["gte"].isoformat()[:10])
            p += 1
        elif "lte" in date_filter:
            conditions.append(f"date <= ${p}::date")
            params.append(date_filter["lte"].isoformat()[:10])
            p += 1

    if search:
        conditions.append(f"LOWER(details) LIKE ${p}")
        params.append(f"%{search.lower()}%")
        p += 1

    where_sql = "WHERE " + " AND ".join(conditions)

    skip = pagination.skip if pagination else 0
    take = pagination.take if pagination else 50

    count_rows = await db.query_raw(
        f"SELECT COUNT(*) AS cnt FROM pmak_tools {where_sql}",
        *params,
    )
    total = int(count_rows[0]["cnt"]) if count_rows else 0

    agg_rows = await db.query_raw(
        f"""
        SELECT
            COALESCE(SUM(debit),  0)                          AS total_debit,
            COALESCE(SUM(credit), 0)                          AS total_credit,
            COALESCE(
                (
                  SELECT t2.total
                    FROM pmak_tools t2
                   WHERE t2.account_id = $1
                   ORDER BY t2.created_at DESC
                   LIMIT 1
                ), 0
            )                                                 AS latest_total
          FROM pmak_tools {where_sql}
        """,
        *params,
    )
    agg = agg_rows[0] if agg_rows else {}

    rows = await db.query_raw(
        f"""
        SELECT * FROM pmak_tools {where_sql}
         ORDER BY date DESC, created_at DESC
         LIMIT ${p} OFFSET ${p+1}
        """,
        *params,
        take,
        skip,
    )

    ts = await _fetch_account_timestamps(db, account_id)

    return {
        "account": {
            "id":          account.id,
            "accountName": account.accountName,
            "isActive":    account.isActive,
            "createdAt":   ts["createdAt"],
            "updatedAt":   ts["updatedAt"],
        },
        "totalDebit":  float(agg.get("total_debit",  0)),
        "totalCredit": float(agg.get("total_credit", 0)),
        "latestTotal": float(agg.get("latest_total", 0)),
        "pagination": {
            "page":       pagination.page if pagination else 1,
            "pageSize":   take,
            "total":      total,
            "totalPages": max(1, -(-total // take)),
        },
        "tools": [_serialize_tool_row(r, account.accountName) for r in rows],
    }


async def update_tool(db: Prisma, tool_id: str, data: PmakToolUpdate):
    """
    PATCH /tools/{tool_id} — Partial update of any field.

    Same auto-recompute logic as update_transaction for debit/credit changes.
    """
    await _ensure_schema(db)

    rows = await db.query_raw(
        "SELECT * FROM pmak_tools WHERE id = $1",
        tool_id,
    )
    if not rows:
        raise HTTPException(status_code=404, detail="PMAK tool entry not found")
    current = rows[0]

    # Resolve new account if requested
    new_account = None
    if data.account_name is not None:
        new_account = await _resolve_active_account(db, data.account_name)

    # Build the field-level patch dict
    patch: dict = {}

    if new_account:
        patch["account_id"] = new_account.id
    if data.date is not None:
        patch["date"] = data.date.isoformat()   # "YYYY-MM-DD" string — will be cast ::date in SET clause
    if data.details is not None:
        patch["details"] = data.details
    if data.debit is not None:
        patch["debit"] = float(data.debit)
    if data.credit is not None:
        patch["credit"] = float(data.credit)

    # ── total resolution ─────────────────────────────────────────────────────
    computed_total: Optional[float] = None

    caller_override = (
        data.total is not None
        and data.total != Decimal("0")
    )

    if caller_override:
        computed_total = float(data.total)
        patch["total"] = computed_total

    elif "debit" in patch or "credit" in patch:
        old_total  = float(current["total"])
        old_debit  = float(current["debit"])
        old_credit = float(current["credit"])
        # Reverse the stored formula to get balance before this entry
        balance_before = old_total + old_debit - old_credit
        new_debit  = patch.get("debit",  old_debit)
        new_credit = patch.get("credit", old_credit)
        computed_total = round(balance_before - new_debit + new_credit, 2)
        patch["total"] = computed_total

    # Idempotent — nothing changed
    if not patch:
        acct = await db.pmakaccount.find_unique(where={"id": current["account_id"]})
        return _serialize_tool_row(current, acct.accountName if acct else "")

    patch["updated_at"] = _now_utc().isoformat()  # ISO-8601 string — will be cast ::timestamptz in SET clause

    # Build SET clauses with explicit type casts for date/timestamptz columns.
    # All other columns (TEXT, NUMERIC) need no cast — bare $N is fine.
    _DATE_CAST  = {"date": "::date", "created_at": "::timestamptz", "updated_at": "::timestamptz"}
    set_clauses = [
        f"{col} = ${i+2}{_DATE_CAST.get(col, '')}"
        for i, col in enumerate(patch)
    ]
    sql_params  = [tool_id] + list(patch.values())
    await db.execute_raw(
        f"UPDATE pmak_tools SET {', '.join(set_clauses)} WHERE id = $1",
        *sql_params,
    )

    updated_rows = await db.query_raw(
        "SELECT * FROM pmak_tools WHERE id = $1",
        tool_id,
    )
    row = updated_rows[0]
    account_id   = new_account.id if new_account else current["account_id"]
    acct         = await db.pmakaccount.find_unique(where={"id": account_id})
    account_name = acct.accountName if acct else ""

    serialized = _serialize_tool_row(row, account_name)
    if computed_total is not None:
        serialized["total"] = Decimal(str(computed_total))
    return serialized


async def delete_tool(db: Prisma, tool_id: str):
    """DELETE /tools/{tool_id} — Hard delete."""
    await _ensure_schema(db)
    rows = await db.query_raw(
        "SELECT id FROM pmak_tools WHERE id = $1",
        tool_id,
    )
    if not rows:
        raise HTTPException(status_code=404, detail="PMAK tool entry not found")
    await db.execute_raw("DELETE FROM pmak_tools WHERE id = $1", tool_id)


# ═════════════════════════════════════════════════════════════════════════════
# § 8  Excel Export — single account  (3 sheets: Ledger + Inhouse + Tools)
# ═════════════════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _apply_header(ws, headers: list, col_widths: list) -> None:
    ws.row_dimensions[1].height = 22
    for idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=1, column=idx, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(idx)].width = w


async def export_account_excel(
    db:         Prisma,
    account_id: str,
    filters:    DateRangeFilter,
) -> tuple:
    """
    Three-sheet workbook:
      Sheet 1 — Ledger Transactions
      Sheet 2 — Inhouse Deals
      Sheet 3 — Tools
    """
    await _ensure_schema(db)
    account = await db.pmakaccount.find_unique(where={"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="PMAK account not found")

    date_filter = filters.to_prisma_filter()
    txn_where:  dict = {"accountId": account_id}
    deal_where: dict = {"accountId": account_id}
    if date_filter:
        txn_where["date"]  = date_filter
        deal_where["date"] = date_filter

    txns  = await db.pmaktransaction.find_many(where=txn_where,  order={"date": "asc"})
    deals = await db.pmakinhouse.find_many(    where=deal_where, order={"date": "asc"})

    # Fetch tools for this account with optional date filter
    tool_conditions: List[str] = ["account_id = $1"]
    tool_params:     List     = [account_id]
    tp = 2
    if date_filter:
        if "gte" in date_filter and "lte" in date_filter:
            tool_conditions.append(f"date >= ${tp}::date AND date <= ${tp+1}::date")
            tool_params.extend([
                date_filter["gte"].isoformat()[:10],
                date_filter["lte"].isoformat()[:10],
            ])
            tp += 2
        elif "gte" in date_filter:
            tool_conditions.append(f"date >= ${tp}::date")
            tool_params.append(date_filter["gte"].isoformat()[:10])
            tp += 1
        elif "lte" in date_filter:
            tool_conditions.append(f"date <= ${tp}::date")
            tool_params.append(date_filter["lte"].isoformat()[:10])
            tp += 1

    tool_where_sql = "WHERE " + " AND ".join(tool_conditions)
    tool_rows = await db.query_raw(
        f"SELECT * FROM pmak_tools {tool_where_sql} ORDER BY date ASC, created_at ASC",
        *tool_params,
    )

    wb  = openpyxl.Workbook()

    # ── Sheet 1: Ledger Transactions ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ledger"
    _apply_header(ws1, [
        "Date", "Details", "Account From", "Account To",
        "Debit", "Credit", "Balance", "Status",
    ], [12, 36, 20, 20, 12, 12, 14, 12])

    for row_idx, t in enumerate(txns, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        vals = [
            t.date.date().isoformat() if hasattr(t.date, "date") else str(t.date),
            t.details or "",
            t.accountFrom or "",
            t.accountTo   or "",
            float(t.debit),
            float(t.credit),
            float(t.remainingBalance),
            t.status if isinstance(t.status, str) else t.status.value,
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell           = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _CENTER if col_idx in (1, 5, 6, 7, 8) else _LEFT
            if fill:
                cell.fill = fill
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Inhouse Deals ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Inhouse Deals")
    _apply_header(ws2, [
        "Date", "Buyer", "Seller", "Amount", "Status", "Details",
    ], [12, 24, 24, 14, 14, 40])

    for row_idx, d in enumerate(deals, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        vals = [
            d.date.date().isoformat() if hasattr(d.date, "date") else str(d.date),
            d.buyerName,
            d.sellerName,
            float(d.orderAmount),
            d.orderStatus if isinstance(d.orderStatus, str) else d.orderStatus.value,
            d.details or "",
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell           = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _CENTER if col_idx in (1, 4, 5) else _LEFT
            if fill:
                cell.fill = fill
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Tools ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Tools")
    _apply_header(ws3, [
        "Date", "Details", "Debit", "Credit", "Total",
    ], [12, 42, 12, 12, 14])

    for row_idx, r in enumerate(tool_rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        raw_date = r.get("date")
        if isinstance(raw_date, datetime):
            date_str = raw_date.date().isoformat()
        elif isinstance(raw_date, dt_date):
            date_str = raw_date.isoformat()
        elif isinstance(raw_date, str):
            date_str = raw_date[:10]
        else:
            date_str = ""

        vals = [
            date_str,
            r.get("details") or "",
            float(r.get("debit",  0)),
            float(r.get("credit", 0)),
            float(r.get("total",  0)),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell           = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _CENTER if col_idx in (1, 3, 4, 5) else _LEFT
            if fill:
                cell.fill = fill
    ws3.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = account.accountName.replace(" ", "_").replace("/", "-")
    period    = filters.meta()["dateRange"]["from"] or "all"
    filename  = f"pmak_{safe_name}_{period}.xlsx"
    return buffer.read(), filename



async def list_all_transactions(
    db:           Prisma,
    filters:      DateRangeFilter,
    pagination:   PageParams,
    account_name: Optional[str] = None,
    search:       Optional[str] = None,
    status:       Optional[str] = None,
):
    """
    GET /transactions — Cross-account flat transaction list.

    Filters (all combinable):
      account_name — case-insensitive substring on PmakAccount.accountName
      search       — case-insensitive keyword search on PmakTransaction.details
      status       — exact enum: PENDING | CLEARED | ON_HOLD | REJECTED
      period / from / to / year / month — standard DateRangeFilter
    """
    await _ensure_schema(db)
    date_filter = filters.to_prisma_filter()

    where: dict = {}

    if date_filter:
        where["date"] = date_filter

    if search:
        where["details"] = {"contains": search, "mode": "insensitive"}

    if status:
        valid_statuses = {"PENDING", "CLEARED", "ON_HOLD", "REJECTED"}
        normalised = status.upper()
        if normalised not in valid_statuses:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status '{status}'. "
                       f"Valid values: {', '.join(sorted(valid_statuses))}",
            )
        where["status"] = normalised

    account_filter: dict = {"isActive": True}
    if account_name:
        account_filter["accountName"] = {"contains": account_name, "mode": "insensitive"}
    where["account"] = {"is": account_filter}

    total = await db.pmaktransaction.count(where=where)

    txns = await db.pmaktransaction.find_many(
        where=where,
        order={"date": "desc"},
        skip=pagination.skip,
        take=pagination.take,
        include={"account": True},
    )

    # Aggregates across the full matching set (not just this page)
    all_txns_for_totals = await db.pmaktransaction.find_many(where=where)
    total_debit  = sum(float(t.debit)  for t in all_txns_for_totals)
    total_credit = sum(float(t.credit) for t in all_txns_for_totals)

    def _serialize_txn_with_account(t) -> dict:
        return {
            "id":               t.id,
            "accountId":        t.accountId,
            "accountName":      t.account.accountName if t.account else "",
            "date":             t.date.date() if hasattr(t.date, "date") else t.date,
            "details":          t.details,
            "accountFrom":      t.accountFrom,
            "accountTo":        t.accountTo,
            "debit":            t.debit,
            "credit":           t.credit,
            "remainingBalance": t.remainingBalance,
            "status":           t.status if isinstance(t.status, str) else t.status.value,
            "createdAt":        t.createdAt,
        }

    return {
        "filter": filters.meta(),
        "totals": {
            "totalTransactions": total,
            "totalDebit":        round(total_debit,  2),
            "totalCredit":       round(total_credit, 2),
        },
        "pagination": {
            "page":       pagination.page,
            "pageSize":   pagination.take,
            "total":      total,
            "totalPages": max(1, -(-total // pagination.take)),
        },
        "transactions": [_serialize_txn_with_account(t) for t in txns],
    }

# ═════════════════════════════════════════════════════════════════════════════
# § 9  Name aliases — router uses short names; keep both sides in sync
# ═════════════════════════════════════════════════════════════════════════════

list_transactions   = list_all_transactions
add_inhouse         = create_inhouse_deal
update_inhouse      = update_inhouse_deal
delete_inhouse      = delete_inhouse_deal
get_all_inhouse     = list_all_inhouse_deals
get_account_inhouse = get_account_inhouse_deals