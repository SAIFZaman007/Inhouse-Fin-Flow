"""
app/modules/upwork/service.py
════════════════════════════════════════════════════════════════════════════════
v14 — Enterprise Edition
════════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

import io
import logging
import math
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Optional

from fastapi import HTTPException
from prisma import Prisma

from app.core import trash_store
from app.shared.filters import DateRangeFilter
from app.shared.pagination import PageParams
from .schema import (
    UpworkOrderCreate,
    UpworkOrderUpdate,
    UpworkProfileCreate,
    UpworkProfileUpdate,
    UpworkSnapshotCreate,
    UpworkSnapshotUpdate,
)

logger = logging.getLogger(__name__)

_UPWORK_FEE = Decimal("0.10")
_AFTER_RATE = Decimal("1") - _UPWORK_FEE   # 0.90
_ZERO       = Decimal("0")

# Snapshot field names present on UpworkProfileUpdate (used for presence check)
_SNAPSHOT_FIELDS = frozenset({
    "available_withdraw",
    "pending",
    "in_review",
    "work_in_progress",
    "withdrawn",
    "connects",
    "upwork_plus",
})


# ── Private helpers ───────────────────────────────────────────────────────────

def _d(v: Any) -> Decimal:
    return _ZERO if v is None else Decimal(str(v))


def _after_fee(amount: Any) -> Decimal:
    return (_d(amount) * _AFTER_RATE).quantize(Decimal("0.01"))


def _revenue_from_snapshot(entry: Any) -> Decimal:
    """
    Compute revenueInPeriod from a snapshot row.

    Formula (spec):
        revenueInPeriod = availableWithdraw + pending + inReview - withdrawn
    """
    if entry is None:
        return _ZERO
    return (
        _d(entry.availableWithdraw)
        + _d(entry.pending)
        + _d(entry.inReview)
        - _d(entry.withdrawn)
    )


def _entry_to_dict(entry: Any, profile_name: str) -> dict:
    """
    Serialise a UpworkEntry ORM object to a plain dict.

    ``activeAmount`` always mirrors ``workInProgress`` (same DB column).
    ``revenueInPeriod`` = availableWithdraw + pending + inReview - withdrawn.
    """
    aw  = _d(entry.availableWithdraw)
    wip = _d(entry.workInProgress)
    rev = _revenue_from_snapshot(entry)
    return {
        "id":                        entry.id,
        "profileId":                 entry.profileId,
        "profileName":               profile_name,
        "date":                      entry.date.date() if isinstance(entry.date, datetime) else entry.date,
        "availableWithdraw":         float(aw),
        "availableWithdrawAfterFee": float(_after_fee(aw)),
        "pending":                   float(_d(entry.pending)),
        "inReview":                  float(_d(entry.inReview)),
        "workInProgress":            float(wip),
        "activeAmount":              float(wip),          # always equals workInProgress
        "withdrawn":                 float(_d(entry.withdrawn)),
        "revenueInPeriod":           float(rev),          # aw + pending + inReview - withdrawn
        "connects":                  entry.connects,
        "upworkPlus":                entry.upworkPlus,
        "createdAt":                 entry.createdAt,
    }


def _order_to_dict(order: Any) -> dict:
    return {
        "id":          order.id,
        "profileId":   order.profileId,
        "date":        order.date.date() if isinstance(order.date, datetime) else order.date,
        "clientName":  order.clientName,
        "orderId":     order.orderId,
        "amount":      float(_d(order.amount)),
        "afterUpwork": float(_d(order.afterUpwork)),
        "createdAt":   order.createdAt,
    }


def _pagination_meta(pagination: Optional[PageParams], total: int) -> dict:
    if pagination is None:
        return {"page": 1, "pageSize": total, "total": total, "totalPages": 1}
    return {
        "page":       pagination.page,
        "pageSize":   pagination.page_size,
        "total":      total,
        "totalPages": math.ceil(total / pagination.page_size) if total > 0 else 1,
    }


async def _get_profile_or_404(db: Prisma, profile_id: str):
    profile = await db.upworkprofile.find_unique(where={"id": profile_id})
    if not profile:
        raise HTTPException(status_code=404, detail="Upwork profile not found.")
    return profile


async def _resolve_profile_by_name(db: Prisma, profile_name: str):
    profile = await db.upworkprofile.find_first(
        where={
            "profileName": {"equals": profile_name, "mode": "insensitive"},
            "isActive":    True,
        }
    )
    if not profile:
        raise HTTPException(
            status_code=404,
            detail=f"Upwork profile '{profile_name}' not found.",
        )
    return profile


# ── Snapshot sync helper ──────────────────────────────────────────────────────

async def _sync_snapshot_wip(
    db: Prisma,
    profile_id: str,
    snap_dt: datetime,
    delta: Decimal,
) -> None:
    """
    Additively increment ``workInProgress`` on the snapshot for
    ``(profile_id, snap_dt)`` by ``delta``.

    If no snapshot row exists yet for that date it is created with
    ``workInProgress = delta`` and all other numeric fields at zero.
    """
    existing = await db.upworkentry.find_unique(
        where={"profileId_date": {"profileId": profile_id, "date": snap_dt}}
    )

    if existing is None:
        await db.upworkentry.create(
            data={
                "profileId":         profile_id,
                "date":              snap_dt,
                "availableWithdraw": _ZERO,
                "workInProgress":    delta,
            }
        )
    else:
        new_wip = _d(existing.workInProgress) + delta
        await db.upworkentry.update(
            where={"profileId_date": {"profileId": profile_id, "date": snap_dt}},
            data={"workInProgress": new_wip},
        )


# ── Profile CRUD ──────────────────────────────────────────────────────────────

async def create_profile(db: Prisma, data: UpworkProfileCreate) -> dict:
    if not data.profileName:
        raise HTTPException(status_code=422, detail="profileName is required.")

    existing = await db.upworkprofile.find_unique(
        where={"profileName": data.profileName}
    )
    if existing:
        raise HTTPException(status_code=409, detail="Profile name already exists.")

    profile = await db.upworkprofile.create(
        data={"profileName": data.profileName}
    )

    snapshot: Optional[dict] = None
    if data.available_withdraw is not None:
        snap_date = data.snapshot_date or date.today()
        entry = await db.upworkentry.upsert(
            where={"profileId_date": {
                "profileId": profile.id,
                "date":      datetime.combine(snap_date, time.min),
            }},
            data={
                "create": {
                    "profileId":         profile.id,
                    "date":              datetime.combine(snap_date, time.min),
                    "availableWithdraw": data.available_withdraw,
                    "pending":           data.pending,
                    "inReview":          data.in_review,
                    "workInProgress":    data.work_in_progress,
                    "withdrawn":         data.withdrawn,
                    "connects":          data.connects,
                    "upworkPlus":        data.upwork_plus,
                },
                "update": {
                    "availableWithdraw": data.available_withdraw,
                    "pending":           data.pending,
                    "inReview":          data.in_review,
                    "workInProgress":    data.work_in_progress,
                    "withdrawn":         data.withdrawn,
                    "connects":          data.connects,
                    "upworkPlus":        data.upwork_plus,
                },
            },
        )
        snapshot = _entry_to_dict(entry, profile.profileName)

    return {
        "id":              profile.id,
        "profileName":     profile.profileName,
        "isActive":        profile.isActive,
        "initialSnapshot": snapshot,
    }


async def update_profile(
    db: Prisma,
    profile_id: str,
    data: UpworkProfileUpdate,
) -> dict:
    """
    PATCH /profiles/{id} — partial profile update (v7).

    1. Profile metadata  — rename (with uniqueness check) and/or isActive toggle.
    2. Snapshot upsert   — if any snapshot field is supplied, upsert the
                           UpworkEntry for ``snapshot_date`` (defaults to today).
    """
    profile = await _get_profile_or_404(db, profile_id)

    profile_patch: dict = {}

    if data.profileName is not None and data.profileName != profile.profileName:
        conflict = await db.upworkprofile.find_first(
            where={
                "profileName": {"equals": data.profileName, "mode": "insensitive"},
                "id":          {"not": profile_id},
            }
        )
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Profile name '{data.profileName}' is already taken.",
            )
        profile_patch["profileName"] = data.profileName

    if data.isActive is not None:
        profile_patch["isActive"] = data.isActive

    if profile_patch:
        profile = await db.upworkprofile.update(
            where={"id": profile_id},
            data=profile_patch,
        )

    sent            = data.model_fields_set
    snapshot_fields = sent & _SNAPSHOT_FIELDS
    upserted_snapshot: Optional[dict] = None

    if snapshot_fields:
        snap_date = data.snapshot_date or date.today()

        existing_entry = await db.upworkentry.find_unique(
            where={"profileId_date": {
                "profileId": profile_id,
                "date":      datetime.combine(snap_date, time.min),
            }}
        )

        def _pick(field: str, existing_attr: str, default: Any) -> Any:
            if field in sent:
                return getattr(data, field)
            if existing_entry is not None:
                return getattr(existing_entry, existing_attr)
            return default

        aw               = _pick("available_withdraw", "availableWithdraw", _ZERO)
        pending          = _pick("pending",            "pending",           _ZERO)
        in_review        = _pick("in_review",          "inReview",          _ZERO)
        work_in_progress = _pick("work_in_progress",   "workInProgress",    _ZERO)
        withdrawn        = _pick("withdrawn",           "withdrawn",         _ZERO)
        connects         = _pick("connects",            "connects",          0)
        upwork_plus      = _pick("upwork_plus",         "upworkPlus",        False)

        entry_data = {
            "availableWithdraw": aw,
            "pending":           pending,
            "inReview":          in_review,
            "workInProgress":    work_in_progress,
            "withdrawn":         withdrawn,
            "connects":          connects,
            "upworkPlus":        upwork_plus,
        }

        entry = await db.upworkentry.upsert(
            where={"profileId_date": {
                "profileId": profile_id,
                "date":      datetime.combine(snap_date, time.min),
            }},
            data={
                "create": {"profileId": profile_id, "date": datetime.combine(snap_date, time.min), **entry_data},
                "update": entry_data,
            },
        )
        upserted_snapshot = _entry_to_dict(entry, profile.profileName)

    return {
        "id":               profile.id,
        "profileName":      profile.profileName,
        "isActive":         profile.isActive,
        "snapshotUpserted": upserted_snapshot,
    }


async def deactivate_profile(db: Prisma, profile_id: str) -> None:
    """Legacy hard soft-delete — sets isActive=False only (no trash registry)."""
    profile = await db.upworkprofile.find_unique(where={"id": profile_id})
    if not profile:
        raise HTTPException(status_code=404, detail="Upwork profile not found.")
    await db.upworkprofile.update(
        where={"id": profile_id}, data={"isActive": False}
    )


async def soft_delete_profile(db: Prisma, profile_id: str) -> dict:
    """
    Full soft-delete for a profile (v14).

    1. Fetch profile + all its entries + all its orders.
    2. Write the profile record to trash_store (type="profile").
    3. Write each entry to trash_store (type="snapshot").
    4. Write each order to trash_store (type="order").
    5. Set isActive=False on the profile.
    6. Return confirmation with counts.
    """
    profile = await db.upworkprofile.find_unique(
        where={"id": profile_id},
        include={"entries": True, "orders": True},
    )
    if not profile:
        raise HTTPException(status_code=404, detail="Upwork profile not found.")

    # Capture full profile snapshot for trash
    profile_snap = {
        "id":          profile.id,
        "profileName": profile.profileName,
        "isActive":    profile.isActive,
    }
    await trash_store.add(
        record_id=profile.id,
        module="upwork",
        record_type="profile",
        snapshot=profile_snap,
    )

    # Archive each snapshot
    for entry in profile.entries:
        entry_snap = _entry_to_dict(entry, profile.profileName)
        await trash_store.add(
            record_id=entry.id,
            module="upwork",
            record_type="snapshot",
            snapshot=entry_snap,
        )

    # Archive each order
    for order in profile.orders:
        order_snap = _order_to_dict(order)
        await trash_store.add(
            record_id=order.id,
            module="upwork",
            record_type="order",
            snapshot=order_snap,
        )

    # Soft-delete at DB level
    await db.upworkprofile.update(
        where={"id": profile_id}, data={"isActive": False}
    )

    order_count = len(profile.orders)
    snap_count  = len(profile.entries)

    return {
        "success":         True,
        "message":         "Upwork profile has been soft-deleted.",
        "profileId":       profile_id,
        "profileName":     profile.profileName,
        "snapshotsArchived": snap_count,
        "ordersArchived":    order_count,
    }


# ── Snapshot CRUD ─────────────────────────────────────────────────────────────

async def create_snapshot(db: Prisma, data: UpworkSnapshotCreate) -> dict:
    """
    POST /snapshots — additive daily snapshot (v13).

    ROOT CAUSE FIX
    ──────────────
    Uses fetch-then-additive-update pattern:
      • If no row exists for (profile, date) → plain INSERT.
      • If a row already exists             → ADD incoming numeric values.
        ``upworkPlus`` retains OR semantics (sticky True).

    revenueInPeriod formula (spec):
        revenueInPeriod = availableWithdraw + pending + inReview - withdrawn
    """
    if not data.profile_name:
        raise HTTPException(status_code=422, detail="profile_name is required.")
    if not data.date:
        raise HTTPException(status_code=422, detail="date is required.")

    profile = await _resolve_profile_by_name(db, data.profile_name)
    snap_dt = datetime.combine(data.date, time.min)

    # ── Fetch existing row (if any) ───────────────────────────────────────────
    existing = await db.upworkentry.find_unique(
        where={"profileId_date": {
            "profileId": profile.id,
            "date":      snap_dt,
        }}
    )

    # Guard: if this entry is in trash, reject
    if existing and trash_store.is_deleted(existing.id):
        raise HTTPException(
            status_code=409,
            detail="A snapshot for this date was soft-deleted. Restore it first via POST /restore-trash.",
        )

    aw_in  = data.available_withdraw or _ZERO
    pd_in  = data.pending            or _ZERO
    ir_in  = data.in_review          or _ZERO
    wip_in = data.work_in_progress   or _ZERO
    wd_in  = data.withdrawn          or _ZERO
    cn_in  = data.connects           or 0
    pl_in  = data.upwork_plus        or False

    if existing is None:
        # ── First entry for this (profile, date) — plain INSERT ───────────────
        entry = await db.upworkentry.create(
            data={
                "profileId":         profile.id,
                "date":              snap_dt,
                "availableWithdraw": aw_in,
                "pending":           pd_in,
                "inReview":          ir_in,
                "workInProgress":    wip_in,
                "withdrawn":         wd_in,
                "connects":          cn_in,
                "upworkPlus":        pl_in,
            }
        )
    else:
        # ── Row already exists — ADD (accumulate) the incoming values ─────────
        entry = await db.upworkentry.update(
            where={"profileId_date": {
                "profileId": profile.id,
                "date":      snap_dt,
            }},
            data={
                "availableWithdraw": _d(existing.availableWithdraw) + aw_in,
                "pending":           _d(existing.pending)           + pd_in,
                "inReview":          _d(existing.inReview)          + ir_in,
                "workInProgress":    _d(existing.workInProgress)    + wip_in,
                "withdrawn":         _d(existing.withdrawn)         + wd_in,
                "connects":          existing.connects              + cn_in,
                "upworkPlus":        existing.upworkPlus            or pl_in,
            },
        )

    # ── Rebuild aggregated totals ─────────────────────────────────────────────
    all_orders          = await db.upworkorder.find_many(where={"profileId": profile.id})
    # Exclude soft-deleted orders from totals
    live_orders         = [o for o in all_orders if not trash_store.is_deleted(o.id)]
    total_revenue       = sum((_d(o.afterUpwork) for o in live_orders), _ZERO)
    total_active_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)

    updated_aw  = _d(entry.availableWithdraw)
    updated_wip = _d(entry.workInProgress)
    rev         = _revenue_from_snapshot(entry)

    snapshot_dict = _entry_to_dict(entry, profile.profileName)

    return {
        **snapshot_dict,
        "syncedTotals": {
            "revenueAllTime":       float(total_revenue),
            "activeAmountAllTime":  float(total_active_amount),
            "latestSnapshot": {
                "availableWithdraw":         float(updated_aw),
                "availableWithdrawAfterFee": float(_after_fee(updated_aw)),
                "pending":                   float(_d(entry.pending)),
                "inReview":                  float(_d(entry.inReview)),
                "workInProgress":            float(updated_wip),
                "activeAmount":              float(updated_wip),
                "withdrawn":                 float(_d(entry.withdrawn)),
                "revenueInPeriod":           float(rev),
                "connects":                  entry.connects,
            },
        },
    }


async def update_snapshot(
    db: Prisma,
    snapshot_id: str,
    data: UpworkSnapshotUpdate,
) -> dict:
    """
    PATCH /snapshots/{id} — partial update with SET semantics (v14).

    Only supplied fields are overwritten. This is correction-mode, not
    accumulation-mode — existing values are replaced, not added to.
    """
    entry = await db.upworkentry.find_unique(where={"id": snapshot_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Upwork snapshot not found.")

    if trash_store.is_deleted(snapshot_id):
        raise HTTPException(
            status_code=410,
            detail="This snapshot has been soft-deleted. Restore it first via POST /restore-trash.",
        )

    patch: dict = {}
    sent = data.model_fields_set

    if "available_withdraw" in sent and data.available_withdraw is not None:
        patch["availableWithdraw"] = data.available_withdraw
    if "pending" in sent and data.pending is not None:
        patch["pending"] = data.pending
    if "in_review" in sent and data.in_review is not None:
        patch["inReview"] = data.in_review
    if "work_in_progress" in sent and data.work_in_progress is not None:
        patch["workInProgress"] = data.work_in_progress
    if "withdrawn" in sent and data.withdrawn is not None:
        patch["withdrawn"] = data.withdrawn
    if "connects" in sent and data.connects is not None:
        patch["connects"] = data.connects
    if "upwork_plus" in sent and data.upwork_plus is not None:
        patch["upworkPlus"] = data.upwork_plus

    if not patch:
        # Idempotent — return current state
        profile = await _get_profile_or_404(db, entry.profileId)
        return _entry_to_dict(entry, profile.profileName)

    updated = await db.upworkentry.update(
        where={"id": snapshot_id},
        data=patch,
    )
    profile = await _get_profile_or_404(db, updated.profileId)
    return _entry_to_dict(updated, profile.profileName)


async def soft_delete_snapshot(db: Prisma, snapshot_id: str) -> dict:
    """
    DELETE /snapshots/{id} — soft-delete a snapshot (v14).

    1. Fetch snapshot + owning profile.
    2. Write full snapshot dict to trash_store.
    3. The row remains in DB — it is excluded from all live calculations
       via ``trash_store.is_deleted()`` guards in list/detail builders.
    4. Return the trash item dict for confirmation.
    """
    entry = await db.upworkentry.find_unique(where={"id": snapshot_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Upwork snapshot not found.")

    if trash_store.is_deleted(snapshot_id):
        raise HTTPException(
            status_code=409,
            detail="Snapshot is already in trash.",
        )

    profile = await _get_profile_or_404(db, entry.profileId)
    snap_dict = _entry_to_dict(entry, profile.profileName)

    trash_item = await trash_store.add(
        record_id=snapshot_id,
        module="upwork",
        record_type="snapshot",
        snapshot=snap_dict,
    )

    return {
        "success":   True,
        "message":   "Upwork snapshot has been soft-deleted.",
        "trashItem": trash_item,
    }


async def get_profile_snapshots(
    db: Prisma,
    profile_id: str,
    date_filter: dict,
    pagination: Optional[PageParams] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    where: dict = {"profileId": profile_id}
    if date_filter:
        where["date"] = date_filter

    all_entries = await db.upworkentry.find_many(where=where, order={"date": "desc"})
    # Exclude soft-deleted
    live_entries = [e for e in all_entries if not trash_store.is_deleted(e.id)]
    total = len(live_entries)

    if pagination:
        live_entries = live_entries[pagination.skip: pagination.skip + pagination.take]

    return {
        "profileId":   profile_id,
        "profileName": profile.profileName,
        "pagination":  _pagination_meta(pagination, total),
        "snapshots":   [_entry_to_dict(e, profile.profileName) for e in live_entries],
    }


# ── Order CRUD ────────────────────────────────────────────────────────────────

async def add_order(db: Prisma, data: UpworkOrderCreate) -> dict:
    """
    POST /orders — log a new Upwork order and sync the daily snapshot (v14).

    1. Resolve the active profile by name (case-insensitive).
    2. Guard against duplicate orderId (409 on conflict).
    3. Persist the UpworkOrder row with server-computed ``afterUpwork``.
    4. Additively sync the UpworkEntry (snapshot) for the same date:
         workInProgress += order.amount
       Snapshot is created if it doesn't exist for that date.
    5. Return the persisted order dict + snapshotSync summary + syncedTotals.

    orderCount is dynamic — computed from live (non-trashed) orders.
    """
    if not data.profile_name:
        raise HTTPException(status_code=422, detail="profile_name is required.")
    if not data.date:
        raise HTTPException(status_code=422, detail="date is required.")
    if not data.client_name:
        raise HTTPException(status_code=422, detail="client_name is required.")
    if not data.order_id:
        raise HTTPException(status_code=422, detail="order_id is required.")
    if data.amount is None:
        raise HTTPException(status_code=422, detail="amount is required.")

    profile = await _resolve_profile_by_name(db, data.profile_name)

    existing = await db.upworkorder.find_unique(where={"orderId": data.order_id})
    if existing:
        raise HTTPException(status_code=409, detail="Order ID already exists.")

    snap_dt = datetime.combine(data.date, time.min)

    order = await db.upworkorder.create(
        data={
            "profileId":   profile.id,
            "date":        snap_dt,
            "clientName":  data.client_name,
            "orderId":     data.order_id,
            "amount":      data.amount,
            "afterUpwork": _after_fee(data.amount),
        }
    )

    await _sync_snapshot_wip(db, profile.id, snap_dt, _d(data.amount))

    updated_entry = await db.upworkentry.find_unique(
        where={"profileId_date": {"profileId": profile.id, "date": snap_dt}}
    )

    # Dynamic orderCount — live orders only
    all_orders          = await db.upworkorder.find_many(where={"profileId": profile.id})
    live_orders         = [o for o in all_orders if not trash_store.is_deleted(o.id)]
    total_revenue       = sum((_d(o.afterUpwork) for o in live_orders), _ZERO)
    total_active_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)

    rev = _revenue_from_snapshot(updated_entry) if updated_entry else _ZERO

    return {
        **_order_to_dict(order),
        "snapshotSync": {
            "date":            str(data.date),
            "workInProgress":  float(_d(updated_entry.workInProgress)) if updated_entry else float(_d(data.amount)),
            "activeAmount":    float(_d(updated_entry.workInProgress)) if updated_entry else float(_d(data.amount)),
            "latestSnapshot": {
                "availableWithdraw":         float(_d(updated_entry.availableWithdraw))            if updated_entry else 0.0,
                "availableWithdrawAfterFee": float(_after_fee(_d(updated_entry.availableWithdraw))) if updated_entry else 0.0,
                "pending":                   float(_d(updated_entry.pending))                       if updated_entry else 0.0,
                "inReview":                  float(_d(updated_entry.inReview))                      if updated_entry else 0.0,
                "workInProgress":            float(_d(updated_entry.workInProgress))                if updated_entry else float(_d(data.amount)),
                "activeAmount":              float(_d(updated_entry.workInProgress))                if updated_entry else float(_d(data.amount)),
                "withdrawn":                 float(_d(updated_entry.withdrawn))                     if updated_entry else 0.0,
                "revenueInPeriod":           float(rev),
                "connects":                  updated_entry.connects                                 if updated_entry else 0,
            } if updated_entry else None,
        },
        "syncedTotals": {
            "orderCount":           len(live_orders),
            "revenueAllTime":       float(total_revenue),
            "activeAmountAllTime":  float(total_active_amount),
        },
    }


async def update_order(
    db: Prisma,
    order_id: str,
    data: UpworkOrderUpdate,
) -> dict:
    order = await db.upworkorder.find_unique(where={"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Upwork order not found.")

    if trash_store.is_deleted(order_id):
        raise HTTPException(
            status_code=410,
            detail="This order has been soft-deleted. Restore it via POST /restore-trash.",
        )

    patch: dict = {}
    sent = data.model_fields_set

    if "date" in sent and data.date is not None:
        patch["date"] = datetime.combine(data.date, time.min)

    if data.client_name is not None:
        patch["clientName"] = data.client_name

    if data.order_id is not None and data.order_id != order.orderId:
        conflict = await db.upworkorder.find_unique(where={"orderId": data.order_id})
        if conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Order ID '{data.order_id}' is already taken.",
            )
        patch["orderId"] = data.order_id

    if data.amount is not None:
        patch["amount"]      = data.amount
        patch["afterUpwork"] = _after_fee(data.amount)

    if not patch:
        return _order_to_dict(order)

    updated = await db.upworkorder.update(where={"id": order_id}, data=patch)
    return _order_to_dict(updated)


async def get_profile_orders(
    db: Prisma,
    profile_id: str,
    date_filter: dict,
    pagination: Optional[PageParams] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    where: dict = {"profileId": profile_id}
    if date_filter:
        where["date"] = date_filter

    all_orders  = await db.upworkorder.find_many(where=where, order={"date": "desc"})
    live_orders = [o for o in all_orders if not trash_store.is_deleted(o.id)]
    total       = len(live_orders)

    if pagination:
        live_orders = live_orders[pagination.skip: pagination.skip + pagination.take]

    return {
        "profileId":   profile_id,
        "profileName": profile.profileName,
        "orderCount":  total,
        "pagination":  _pagination_meta(pagination, total),
        "orders":      [_order_to_dict(o) for o in live_orders],
    }


# ── Trash / Restore ───────────────────────────────────────────────────────────

async def get_trash(record_type: Optional[str] = None) -> dict:
    """
    GET /trash — return all Upwork soft-deleted records.

    Optionally filter by ``record_type`` ("profile" | "snapshot" | "order").
    Results are sorted newest-deleted-first.
    """
    items = await trash_store.get_all(module="upwork", record_type=record_type)
    return {
        "total": len(items),
        "items": items,
    }


async def restore_trash(db: Prisma, ids: list[str]) -> dict:
    """
    POST /restore-trash — restore one or more soft-deleted Upwork records (v14).

    For each ID:
    1. Look up the trash item.
    2. If type="profile"  → set isActive=True in DB.
    3. If type="snapshot" → the DB row still exists; simply remove from trash.
    4. If type="order"    → the DB row still exists; simply remove from trash.
    5. Remove from trash registry.

    Returns lists of successfully restored IDs and failed IDs.
    """
    restored: list[str] = []
    failed:   list[str] = []

    for record_id in ids:
        try:
            item = await trash_store.get_by_id(record_id)
            if not item or item.get("module") != "upwork":
                failed.append(record_id)
                continue

            record_type = item.get("type")

            if record_type == "profile":
                # Restore isActive in DB
                profile = await db.upworkprofile.find_unique(where={"id": record_id})
                if profile:
                    await db.upworkprofile.update(
                        where={"id": record_id},
                        data={"isActive": True},
                    )

            elif record_type in ("snapshot", "order"):
                # DB row was never deleted — just remove from trash registry
                pass

            # Remove from trash registry
            removed = await trash_store.remove(record_id)
            if removed:
                restored.append(record_id)
            else:
                failed.append(record_id)

        except Exception as exc:
            logger.error("restore_trash: failed to restore %s — %s", record_id, exc)
            failed.append(record_id)

    total = len(restored)
    return {
        "restored": restored,
        "failed":   failed,
        "message":  f"{total} record(s) restored successfully." if total else "No records were restored.",
    }


# ── List / detail ─────────────────────────────────────────────────────────────

async def list_profiles_summary(
    db: Prisma,
    filters: DateRangeFilter,
    name: Optional[str] = None,
    pagination: Optional[PageParams] = None,
) -> dict:
    """
    GET /profiles — combined totals + paginated per-profile breakdown (v14).

    revenueInPeriod formula (spec):
        revenueInPeriod = availableWithdraw + pending + inReview - withdrawn

    Soft-deleted snapshots and orders are excluded from all calculations
    via ``trash_store.is_deleted()`` guards.

    ``orderCount`` is dynamic — counts only live (non-trashed) orders.
    """
    date_f = filters.to_prisma_filter()

    where: dict = {"isActive": True}
    if name:
        where["profileName"] = {"contains": name, "mode": "insensitive"}

    total_profiles = await db.upworkprofile.count(where=where)

    # ── Global totalOrderCount ─────────────────────────────────────────────────
    # Must be computed across ALL active profiles — not just the current page —
    # so we fetch all matching profile IDs first, then count live (non-trashed)
    # orders in a single query.  This keeps the value pagination-independent.
    all_active_profiles_for_count = await db.upworkprofile.find_many(where=where)
    active_profile_ids            = [p.id for p in all_active_profiles_for_count]
    all_orders_for_count          = await db.upworkorder.find_many(
        where={"profileId": {"in": active_profile_ids}}
    ) if active_profile_ids else []
    t_order_count = sum(
        1 for o in all_orders_for_count if not trash_store.is_deleted(o.id)
    )
    # ─────────────────────────────────────────────────────────────────────────

    find_kw: dict = dict(
        where=where,
        include={
            "entries": {
                "where":    {"date": date_f} if date_f else {},
                "order_by": {"date": "desc"},
            },
            "orders": {
                "where":    {"date": date_f} if date_f else {},
                "order_by": {"date": "desc"},
            },
        },
    )
    if pagination:
        find_kw["skip"] = pagination.skip
        find_kw["take"] = pagination.take

    profiles = await db.upworkprofile.find_many(**find_kw)

    t_avail         = t_pending = t_in_review = t_wip = t_withdrawn = _ZERO
    t_connects      = 0
    t_revenue       = _ZERO   # Σ revenueInPeriod per spec formula
    t_order_revenue = _ZERO   # Σ afterUpwork (order flow)
    t_active_amount = _ZERO   # Σ order.amount

    summaries = []
    for p in profiles:
        # Exclude trashed snapshots
        live_entries = [e for e in p.entries if not trash_store.is_deleted(e.id)]
        # Exclude trashed orders
        live_orders  = [o for o in p.orders  if not trash_store.is_deleted(o.id)]

        latest = live_entries[0] if live_entries else None
        aw     = _d(latest.availableWithdraw) if latest else _ZERO
        wdrawn = _d(latest.withdrawn)          if latest else _ZERO

        t_avail     += aw
        t_pending   += _d(latest.pending)        if latest else _ZERO
        t_in_review += _d(latest.inReview)       if latest else _ZERO
        t_wip       += _d(latest.workInProgress) if latest else _ZERO
        t_withdrawn += wdrawn
        t_connects  += latest.connects            if latest else 0

        period_revenue       = _revenue_from_snapshot(latest)
        period_order_revenue = sum((_d(o.afterUpwork) for o in live_orders), _ZERO)
        period_active_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)

        t_revenue       += period_revenue
        t_order_revenue += period_order_revenue
        t_active_amount += period_active_amount

        wip_val = float(_d(latest.workInProgress) if latest else _ZERO)

        period_totals = {
            "availableWithdraw":         float(aw),
            "availableWithdrawAfterFee": float(_after_fee(aw)),
            "pending":                   float(_d(latest.pending)        if latest else _ZERO),
            "inReview":                  float(_d(latest.inReview)       if latest else _ZERO),
            "workInProgress":            wip_val,
            "activeAmount":              wip_val,
            "withdrawn":                 float(wdrawn),
            "revenueInPeriod":           float(period_revenue),
            "orderRevenueInPeriod":      float(period_order_revenue),
            "totalActiveAmount":         float(period_active_amount),
            "connects":                  latest.connects                  if latest else 0,
        }

        summaries.append({
            "id":              p.id,
            "profileName":     p.profileName,
            "isActive":        p.isActive,
            "latestSnapshot":  _entry_to_dict(latest, p.profileName) if latest else None,
            "periodTotals":    period_totals,
            "snapshotCount":   len(live_entries),
            "orderCount":      len(live_orders),          # dynamic — excludes trashed
            "revenueInPeriod": float(period_revenue),
            "orders":          [_order_to_dict(o) for o in live_orders],
        })

    return {
        "filter": filters.meta(),
        "totals": {
            "totalAvailableWithdraw":         float(t_avail),
            "totalAvailableWithdrawAfterFee": float(_after_fee(t_avail)),
            "totalPending":                   float(t_pending),
            "totalInReview":                  float(t_in_review),
            "totalWorkInProgress":            float(t_wip),
            "totalWithdrawn":                 float(t_withdrawn),
            "totalConnects":                  t_connects,
            "totalRevenueInPeriod":           float(t_revenue),
            "totalOrderRevenueInPeriod":      float(t_order_revenue),
            "totalActiveAmount":              float(t_active_amount),
            "totalOrderCount":                t_order_count,   # dynamic — live orders across all active profiles
            "activeProfileCount":             total_profiles,
        },
        "pagination": _pagination_meta(pagination, total_profiles),
        "profiles":   summaries,
    }


async def get_profile_detail(
    db: Prisma,
    profile_id: str,
    filters: DateRangeFilter,
    pagination: Optional[PageParams] = None,
    name: Optional[str] = None,
) -> dict:
    profile = await _get_profile_or_404(db, profile_id)

    if name and name.lower() not in profile.profileName.lower():
        raise HTTPException(
            status_code=404,
            detail=f"Upwork profile not found matching name '{name}'.",
        )

    date_f = filters.to_prisma_filter()

    snap_where:  dict = {"profileId": profile_id}
    order_where: dict = {"profileId": profile_id}
    if date_f:
        snap_where["date"]  = date_f
        order_where["date"] = date_f

    all_entries = await db.upworkentry.find_many(where=snap_where, order={"date": "desc"})
    all_orders  = await db.upworkorder.find_many(where=order_where, order={"date": "desc"})

    # Exclude soft-deleted rows
    live_entries = [e for e in all_entries if not trash_store.is_deleted(e.id)]
    live_orders  = [o for o in all_orders  if not trash_store.is_deleted(o.id)]

    snap_total  = len(live_entries)
    order_total = len(live_orders)

    if pagination:
        live_entries = live_entries[pagination.skip: pagination.skip + pagination.take]
        live_orders  = live_orders[pagination.skip:  pagination.skip + pagination.take]

    latest        = live_entries[0] if live_entries else None
    aw            = _d(latest.availableWithdraw) if latest else _ZERO
    wip           = _d(latest.workInProgress)    if latest else _ZERO
    revenue       = _revenue_from_snapshot(latest)
    order_revenue = sum((_d(o.afterUpwork) for o in live_orders), _ZERO)
    active_amount = sum((_d(o.amount)      for o in live_orders), _ZERO)

    return {
        "filter": filters.meta(),
        "profile": {
            "id":          profile.id,
            "profileName": profile.profileName,
            "isActive":    profile.isActive,
        },
        "periodTotals": {
            "availableWithdraw":         float(aw),
            "availableWithdrawAfterFee": float(_after_fee(aw)),
            "pending":                   float(_d(latest.pending)        if latest else _ZERO),
            "inReview":                  float(_d(latest.inReview)       if latest else _ZERO),
            "workInProgress":            float(wip),
            "activeAmount":              float(wip),
            "withdrawn":                 float(_d(latest.withdrawn)      if latest else _ZERO),
            "revenueInPeriod":           float(revenue),
            "orderRevenueInPeriod":      float(order_revenue),
            "totalActiveAmount":         float(active_amount),
            "connects":                  latest.connects                  if latest else 0,
            "snapshotCount":             snap_total,
            "orderCount":                order_total,          # dynamic
        },
        "pagination": _pagination_meta(pagination, max(snap_total, order_total)),
        "snapshots":  [_entry_to_dict(e, profile.profileName) for e in live_entries],
        "orders":     [_order_to_dict(o) for o in live_orders],
    }


# ── Excel export ──────────────────────────────────────────────────────────────

async def export_profile_excel(
    db: Prisma,
    profile_id: str,
    filters: DateRangeFilter,
) -> tuple[bytes, str]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — cannot generate Excel export.",
        )

    detail   = await get_profile_detail(db, profile_id, filters)
    pname    = detail["profile"]["profileName"]
    start, end = filters.window()

    wb          = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")

    def _header(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    def _autofit(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    ws1 = wb.active
    ws1.title = "Snapshots"
    _header(ws1, [
        "Date", "Available Withdraw ($)", "After Fee ($)", "Pending ($)",
        "In Review ($)", "Work in Progress ($)", "Active Amount ($)",
        "Withdrawn ($)", "Revenue in Period ($)", "Connects", "Upwork Plus",
    ])
    for ri, s in enumerate(detail["snapshots"], 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([
            str(s["date"]),
            s["availableWithdraw"], s["availableWithdrawAfterFee"],
            s["pending"], s["inReview"], s["workInProgress"],
            s["activeAmount"], s["withdrawn"], s["revenueInPeriod"],
            s["connects"], s["upworkPlus"],
        ], 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            if fill:
                cell.fill = fill
    _autofit(ws1)

    ws2 = wb.create_sheet("Orders")
    _header(ws2, ["Date", "Client", "Order ID", "Amount ($)", "After Upwork ($)"])
    for ri, o in enumerate(detail["orders"], 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, v in enumerate([
            str(o["date"]), o["clientName"], o["orderId"],
            o["amount"], o["afterUpwork"],
        ], 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            if fill:
                cell.fill = fill
    _autofit(ws2)

    buf      = io.BytesIO()
    wb.save(buf)
    tag      = f"{start}_{end}" if start else "all"
    filename = f"upwork_{pname.replace(' ', '_')}_{tag}.xlsx"
    return buf.getvalue(), filename